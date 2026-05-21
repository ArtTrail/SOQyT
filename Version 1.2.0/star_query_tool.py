"""
Stellar Object Query Tool
Queries 5 astronomical databases: SIMBAD, AAVSO VSX, VizieR/2MASS, Gaia DR3, NASA Exoplanet Archive (NEA).
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import math
import statistics
import requests
import json
import os
import urllib.parse
import re
import tkinter.font as tkfont
import webbrowser
import datetime
from collections import defaultdict

# ──────────────────────────────────────────────────────────────
# Config persistence
# ──────────────────────────────────────────────────────────────

_CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')

def _load_config():
    try:
        with open(_CONFIG_PATH, 'r') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}

def _save_config(data: dict):
    try:
        existing = _load_config()
        existing.update(data)
        with open(_CONFIG_PATH, 'w') as f:
            json.dump(existing, f, indent=2)
    except Exception:
        pass

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ──────────────────────────────────────────────────────────────
# Column definitions
# ──────────────────────────────────────────────────────────────

SIMBAD_COLS = [
    ('Name',    'Star Name',  500),
    ('RA_hms',  'RA (hms)',   200),
    ('Dec_dms', 'Dec (dms)',  200),
]

VSX_COLS = [
    ('Name',    'Star Name',  300),
    ('AUID',    'AUID',       120),
    ('Summary', 'Type  /  Period  /  Mag Range', 560),
]

TMASS_COLS = [
    ('Name',    '2MASS ID',   400),
    ('RA_hms',  'RA (hms)',   200),
    ('Dec_dms', 'Dec (dms)',  200),
]

WISE_COLS   = [('Name','AllWISE ID',400),('RA_hms','RA (hms)',200),('Dec_dms','Dec (dms)',200)]
APASS_COLS  = [('Name','APASS ID',  400),('RA_hms','RA (hms)',200),('Dec_dms','Dec (dms)',200)]
TYCHO2_COLS = [('Name','Tycho-2 ID',400),('RA_hms','RA (hms)',200),('Dec_dms','Dec (dms)',200)]
WDS_COLS    = [('Name','WDS ID',    400),('RA_hms','RA (hms)',200),('Dec_dms','Dec (dms)',200)]
ORB6_COLS   = [('Name','Orb6 ID',  400),('RA_hms','RA (hms)',200),('Dec_dms','Dec (dms)',200)]

GAIA_COLS = [
    ('source_id', 'Source ID',  220),
    ('Summary',   'G  /  BP-RP  /  Parallax  /  PM  /  RV  /  RUWE', 700),
]

NEA_OVERVIEW_COLS = [
    ('pl_name',        'Planet',      220),
    ('hostname',       'Host Star',   200),
    ('discoverymethod','Method',      140),
    ('disc_year',      'Year',         70),
    ('disc_facility',  'Facility',    180),
]

NUMERIC_COLS = {
    'Period', 'MaxMag', 'MinMag', 'Dist_arcsec', 'Score', 'GMag', 'N_refs',
    'pl_orbper', 'pl_orbsmax', 'pl_orbeccen', 'pl_orbincl', 'pl_rade', 'pl_bmasse',
    'pl_eqt', 'pl_trandep', 'pl_trandur', 'st_teff', 'st_logg', 'st_met',
    'st_rad', 'st_mass', 'sy_dist', 'sy_vmag', 'sy_kmag', 'sy_gaiamag',
    'disc_year', 'sy_pnum',
}

SPINNER_FRAMES = ['⠋', '⠙', '⠹', '⠸', '⠼', '⠴', '⠦', '⠧', '⠇', '⠏']


# ──────────────────────────────────────────────────────────────
# Coordinate helpers
# ──────────────────────────────────────────────────────────────

def ra_hms_to_deg(s):
    """Convert RA string 'HH MM SS.ss' or 'HH:MM:SS.ss' to decimal degrees."""
    if s is None:
        return None
    s = str(s).strip().replace(':', ' ')
    parts = s.split()
    if len(parts) < 2:
        return None
    try:
        h = float(parts[0])
        m = float(parts[1]) if len(parts) > 1 else 0.0
        sec = float(parts[2]) if len(parts) > 2 else 0.0
        return (h + m / 60.0 + sec / 3600.0) * 15.0
    except (ValueError, IndexError):
        return None

def dec_dms_to_deg(s):
    """Convert Dec string '±DD MM SS.s' or '±DD:MM:SS.s' to decimal degrees."""
    if s is None:
        return None
    s = str(s).strip().replace(':', ' ')
    neg = s.startswith('-')
    s = s.lstrip('+-')
    parts = s.split()
    if not parts:
        return None
    try:
        d = float(parts[0])
        m = float(parts[1]) if len(parts) > 1 else 0.0
        sec = float(parts[2]) if len(parts) > 2 else 0.0
        val = d + m / 60.0 + sec / 3600.0
        return -val if neg else val
    except (ValueError, IndexError):
        return None

def parse_ra(s):
    """Parse RA in hms or decimal degrees."""
    if s is None:
        return None
    s = str(s).strip()
    # If it contains letters or colons or spaces with 3 parts → hms
    if ':' in s or (len(s.split()) == 3 and not s.replace('.', '').replace('-', '').replace(' ', '').isdigit()):
        return ra_hms_to_deg(s)
    try:
        return float(s)
    except ValueError:
        return ra_hms_to_deg(s)

def parse_dec(s):
    """Parse Dec in dms or decimal degrees."""
    if s is None:
        return None
    s = str(s).strip()
    if ':' in s or (len(s.split()) >= 2 and not s.replace('.', '').replace('-', '').replace('+', '').replace(' ', '').isdigit()):
        return dec_dms_to_deg(s)
    try:
        return float(s)
    except ValueError:
        return dec_dms_to_deg(s)

def deg_to_hms(deg):
    """Convert decimal degrees RA to HH MM SS.ss string."""
    if deg is None:
        return ''
    deg = float(deg) % 360.0
    h_total = deg / 15.0
    h = int(h_total)
    m_total = (h_total - h) * 60.0
    m = int(m_total)
    sec = (m_total - m) * 60.0
    return f"{h:02d} {m:02d} {sec:05.2f}"

def deg_to_dms(deg):
    """Convert decimal degrees Dec to ±DD MM SS.s string."""
    if deg is None:
        return ''
    neg = deg < 0
    deg = abs(float(deg))
    d = int(deg)
    m_total = (deg - d) * 60.0
    m = int(m_total)
    sec = (m_total - m) * 60.0
    sign = '-' if neg else '+'
    return f"{sign}{d:02d} {m:02d} {sec:04.1f}"

def _ra_with_deg(hms):
    """Return 'HH MM SS.ss  (DDD.dddd°)' or the original string if conversion fails."""
    if not hms:
        return hms
    deg = ra_hms_to_deg(hms)
    if deg is None:
        return hms
    return f"{hms}  ({deg:.4f}°)"

def _dec_with_deg(dms):
    """Return '±DD MM SS.s  (±DD.dddd°)' or the original string if conversion fails."""
    if not dms:
        return dms
    deg = dec_dms_to_deg(dms)
    if deg is None:
        return dms
    sign = '+' if deg >= 0 else ''
    return f"{dms}  ({sign}{deg:.4f}°)"

def arcsec_dist(ra1, dec1, ra2, dec2):
    """Great-circle distance in arcseconds between two (RA, Dec) points in degrees."""
    if None in (ra1, dec1, ra2, dec2):
        return None
    r1 = math.radians(ra1)
    r2 = math.radians(ra2)
    d1 = math.radians(dec1)
    d2 = math.radians(dec2)
    cos_angle = (math.sin(d1) * math.sin(d2) +
                 math.cos(d1) * math.cos(d2) * math.cos(r1 - r2))
    cos_angle = max(-1.0, min(1.0, cos_angle))
    return math.degrees(math.acos(cos_angle)) * 3600.0

def _parse_vsx_ra(s):
    """Parse VSX RA — decimal degrees (single number) or 'HH MM SS.ss' format."""
    if not s:
        return None
    s = str(s).strip()
    parts = s.replace(':', ' ').split()
    if len(parts) == 1:
        try:
            return float(s)
        except ValueError:
            return None
    return ra_hms_to_deg(s)

def _parse_vsx_dec(s):
    """Parse VSX Dec — decimal degrees or '±DD MM SS.s' format."""
    return dec_dms_to_deg(s)


# ──────────────────────────────────────────────────────────────
# Filter helpers
# ──────────────────────────────────────────────────────────────

def _try_float(val):
    """Return float or None."""
    if val is None or val == '':
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def _strip_flags(val):
    """Strip flag characters from a magnitude/period string, return numeric string or ''."""
    if val is None:
        return ''
    s = str(val).strip().lstrip('<>(: ').rstrip('):> ')
    # Remove single trailing/leading colon or parenthesis
    s = s.strip('():< ')
    try:
        float(s)
        return s
    except ValueError:
        # Try removing individual flag chars
        cleaned = ''
        for c in s:
            if c.isdigit() or c in '.+-':
                cleaned += c
        try:
            float(cleaned)
            return cleaned
        except ValueError:
            return ''

def _passes_filters(period_val, mag_val, period_min, period_max, mag_min, mag_max):
    """
    Apply period and magnitude filters. If the star has no data for that field,
    do NOT exclude it — show it without filtering on that field.
    """
    p = _try_float(period_val)
    m = _try_float(mag_val)

    if p is not None:
        if period_min is not None and p < period_min:
            return False
        if period_max is not None and p > period_max:
            return False

    if m is not None:
        if mag_min is not None and m < mag_min:
            return False
        if mag_max is not None and m > mag_max:
            return False

    return True


# ──────────────────────────────────────────────────────────────
# SIMBAD queries
# ──────────────────────────────────────────────────────────────

SIMBAD_TAP = "https://simbad.u-strasbg.fr/simbad/sim-tap/sync"

OTYPE_LABELS = {
    'RR*': 'RR Lyrae',
    'Ce*': 'Cepheid',
    'dS*': 'Delta Sct',
    'RV*': 'RV Tau',
    'LP*': 'Long-period',
    'SR*': 'Semi-regular',
    'Mi*': 'Mira',
    'Ell*': 'Ellipsoidal',
    'Ro*': 'Rotating',
    'EB*': 'Eclipsing bin.',
    'WV*': 'W Vir',
}

def _build_simbad_results(rows, period_min, period_max, mag_min, mag_max,
                           ra_center=None, dec_center=None, name_mode=False):
    """
    Process raw SIMBAD TAP rows into result dicts.
    SELECT columns: main_id(0), ra(1), dec(2), otype(3), vartyp(4),
                    period(5), vmax(6), vmin(7), magtyp(8), bibcode(9)
    Groups by main_id, computes median period/mag, counts bibcodes for N_refs.
    """
    grouped = defaultdict(list)
    for row in rows:
        grouped[row[0]].append(row)

    results = []
    for main_id, star_rows in grouped.items():
        r0      = star_rows[0]
        ra_val  = r0[1]
        dec_val = r0[2]
        otype   = (r0[3] or '').strip()
        vartype = (r0[4] or '').strip()

        periods_clean = []
        vmaxes_clean  = []
        vmins_clean   = []
        magtyp_val    = ''
        bibcodes      = set()

        for r in star_rows:
            if r[5] is not None:
                try: periods_clean.append(float(r[5]))
                except (ValueError, TypeError): pass
            if r[6] is not None:
                try: vmaxes_clean.append(float(r[6]))
                except (ValueError, TypeError): pass
            if r[7] is not None:
                try: vmins_clean.append(float(r[7]))
                except (ValueError, TypeError): pass
            if r[8]:
                magtyp_val = str(r[8])
            if r[9]:
                bibcodes.add(str(r[9]))

        period_med = statistics.median(periods_clean) if periods_clean else None
        vmax_med   = statistics.median(vmaxes_clean)  if vmaxes_clean  else None
        vmin_med   = statistics.median(vmins_clean)   if vmins_clean   else None

        period_str = f"{period_med:.4f}" if period_med is not None else ''
        maxmag_str = f"{vmax_med:.3f}"   if vmax_med   is not None else ''
        minmag_str = f"{vmin_med:.3f}"   if vmin_med   is not None else ''

        if not _passes_filters(period_str, maxmag_str, period_min, period_max, mag_min, mag_max):
            continue

        try:
            ra_f = float(ra_val) if ra_val is not None else None
        except (ValueError, TypeError):
            ra_f = None
        try:
            dec_f = float(dec_val) if dec_val is not None else None
        except (ValueError, TypeError):
            dec_f = None

        dist = arcsec_dist(ra_center, dec_center, ra_f, dec_f) if ra_center is not None else None
        dist_str = f"{dist:.1f}" if dist is not None else ''

        otype_label = OTYPE_LABELS.get(otype, otype)

        results.append({
            'Name':        main_id or '',
            'RA_hms':      deg_to_hms(ra_f),
            'Dec_dms':     deg_to_dms(dec_f),
            'OType':       otype,
            'OType_label': otype_label,
            'VarType':     vartype,
            'Period':      period_str,
            'MaxMag':      maxmag_str,
            'MinMag':      minmag_str,
            'MagBand':     magtyp_val,
            'Dist_arcsec': dist_str,
            'N_refs':      str(len(bibcodes)),
        })

    return results


def query_simbad(ra_deg, dec_deg, radius_arcmin, otype_filter,
                 period_min, period_max, mag_min, mag_max, status_callback):
    """Coordinate-based SIMBAD cone search (any stellar object, LEFT JOIN mesVar)."""
    radius_deg = radius_arcmin / 60.0

    otype_clause = ""
    if otype_filter and otype_filter != 'All':
        otype_clause = f"AND b.otype = '{otype_filter}'"

    adql = f"""
SELECT b.main_id, b.ra, b.dec, b.otype, v.vartyp, v.period, v.vmax, v.vmin, v.magtyp, v.bibcode
FROM basic b
LEFT JOIN mesVar v ON b.oid = v.oidref
WHERE CONTAINS(POINT('ICRS', b.ra, b.dec),
               CIRCLE('ICRS', {ra_deg}, {dec_deg}, {radius_deg})) = 1
  {otype_clause}
""".strip()

    if status_callback:
        status_callback("Querying SIMBAD...")
    params = {
        'REQUEST': 'doQuery',
        'LANG': 'ADQL',
        'FORMAT': 'json',
        'QUERY': adql,
    }
    r = requests.post(SIMBAD_TAP, data=params, timeout=60)
    r.raise_for_status()
    data = r.json().get('data', [])

    return _build_simbad_results(data, period_min, period_max, mag_min, mag_max,
                                  ra_center=ra_deg, dec_center=dec_deg)


def query_simbad_by_name(name, period_min, period_max, mag_min, mag_max,
                          status_callback, otype_filter=None):
    """Name-based SIMBAD query (LEFT JOIN mesVar so any star can be found)."""

    def _run_adql(adql):
        params = {
            'REQUEST': 'doQuery',
            'LANG': 'ADQL',
            'FORMAT': 'json',
            'QUERY': adql,
        }
        r = requests.post(SIMBAD_TAP, data=params, timeout=60)
        r.raise_for_status()
        return r.json().get('data', [])

    name = _normalize_name(name)

    if status_callback:
        status_callback("Querying SIMBAD...")

    if '*' in name:
        pattern = name.replace('*', '%').replace("'", "''")
        adql_wild = f"""
SELECT TOP 100 b.main_id, b.ra, b.dec, b.otype, v.vartyp, v.period, v.vmax, v.vmin, v.magtyp, v.bibcode
FROM basic b
LEFT JOIN mesVar v ON b.oid = v.oidref
JOIN ident i ON b.oid = i.oidref
WHERE i.id LIKE '{pattern}'
""".strip()
        data = _run_adql(adql_wild)
        if otype_filter and otype_filter != 'All' and data:
            data = [row for row in data if row[3] and row[3].strip() == otype_filter]
        return _build_simbad_results(data, period_min, period_max, mag_min, mag_max,
                                      ra_center=None, dec_center=None, name_mode=True)

    # Exact match first
    adql_exact = f"""
SELECT b.main_id, b.ra, b.dec, b.otype, v.vartyp, v.period, v.vmax, v.vmin, v.magtyp, v.bibcode
FROM basic b
LEFT JOIN mesVar v ON b.oid = v.oidref
JOIN ident i ON b.oid = i.oidref
WHERE i.id = '{name.replace("'", "''")}'
""".strip()

    data = _run_adql(adql_exact)

    if not data:
        # Fallback: LIKE search
        adql_like = f"""
SELECT b.main_id, b.ra, b.dec, b.otype, v.vartyp, v.period, v.vmax, v.vmin, v.magtyp, v.bibcode
FROM basic b
LEFT JOIN mesVar v ON b.oid = v.oidref
JOIN ident i ON b.oid = i.oidref
WHERE i.id LIKE '%{name.replace("'", "''")}%'
""".strip()
        data = _run_adql(adql_like)

    # Apply otype filter if specified
    if otype_filter and otype_filter != 'All' and data:
        data = [row for row in data if row[3] and row[3].strip() == otype_filter]

    return _build_simbad_results(data, period_min, period_max, mag_min, mag_max,
                                  ra_center=None, dec_center=None, name_mode=True)


# ──────────────────────────────────────────────────────────────
# AAVSO VSX queries
# ──────────────────────────────────────────────────────────────

VSX_BASE = "https://www.aavso.org/vsx/index.php"

def _parse_vsx_section(html_text, section_name):
    """Generic parser for a VSX detail-page section (Remarks or References).
    Finds the <td class="datasheethead"> with section_name, then extracts the
    first width="100%" inner table, returning rows as lists of stripped cell texts.
    Also returns the href from the first <a> tag in each cell (or None).
    Each item: list of (text, href|None) tuples, one per detaildata cell.
    """
    import re

    def _strip_tags(s):
        s = re.sub(r'<br\s*/?>', ' ', s, flags=re.IGNORECASE)
        s = re.sub(r'<[^>]+>', '', s)
        s = s.replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>') \
             .replace('&#039;', "'").replace('&quot;', '"').replace('&nbsp;', ' ')
        return ' '.join(s.split())

    def _href(cell_html):
        m = re.search(r'href="([^"]+)"', cell_html, re.IGNORECASE)
        return m.group(1) if m else None

    header_m = re.search(
        rf'class="datasheethead"[^>]*>\s*{re.escape(section_name)}\s*</td>',
        html_text, re.IGNORECASE
    )
    if not header_m:
        return []

    table_m = re.search(
        r'<table[^>]*width="100%"[^>]*>(.*?)</table>',
        html_text[header_m.end():],
        re.DOTALL | re.IGNORECASE
    )
    if not table_m:
        return []

    rows = []
    for row_html in re.findall(r'<tr[^>]*>(.*?)</tr>', table_m.group(1),
                                re.DOTALL | re.IGNORECASE):
        cells = re.findall(r'<td[^>]*class="detaildata"[^>]*>(.*?)</td>',
                           row_html, re.DOTALL | re.IGNORECASE)
        if not cells:
            continue
        rows.append([(_strip_tags(c), _href(c)) for c in cells])
    return rows


def _parse_vsx_page_refs(html_text):
    """Scrape the References section. Returns list of dicts: {citation, bibcode, url}."""
    rows = _parse_vsx_section(html_text, 'References')
    refs = []
    for cells in rows:
        if not cells:
            continue
        citation, cite_url = cells[0]
        bibcode, bib_url   = cells[1] if len(cells) > 1 else ('', None)
        if not citation:
            continue
        if bibcode in ('--', ''):
            bibcode = ''
        url = bib_url if bibcode else cite_url
        refs.append({'citation': citation, 'bibcode': bibcode, 'url': url})
    return refs


def _parse_vsx_page_remarks(html_text):
    """Scrape the Remarks section. Returns list of dicts: {submitter, text}."""
    rows = _parse_vsx_section(html_text, 'Remarks')
    remarks = []
    for cells in rows:
        if len(cells) < 2:
            continue
        submitter, _ = cells[0]
        text, _      = cells[1]
        if text:
            remarks.append({'submitter': submitter, 'text': text})
    return remarks


def _parse_vsx_object(obj, ra_center=None, dec_center=None):
    """Parse a single VSX object dict into a result dict."""
    name = obj.get('Name', '') or ''
    auid = obj.get('AUID', '') or ''

    ra_raw  = obj.get('RA2000', '') or ''
    dec_raw = obj.get('Declination2000', '') or ''

    ra_deg  = _parse_vsx_ra(str(ra_raw).strip())
    dec_deg = _parse_vsx_dec(str(dec_raw).strip())

    vartype       = obj.get('VariabilityType', '') or ''
    period_raw    = obj.get('Period', '') or ''
    maxmag_raw    = obj.get('MaxMag', '') or ''
    minmag_raw    = obj.get('MinMag', '') or ''
    magband       = (obj.get('Bands') or obj.get('Band') or obj.get('FilterBand') or '')
    constellation = obj.get('Constellation', '') or ''
    spectral_type = obj.get('SpectralType', '') or ''
    discoverer    = obj.get('Discoverer', '') or ''
    epoch_raw     = obj.get('Epoch', '') or ''
    rise_dur      = obj.get('RiseDuration', '') or ''
    # Other names — may be comma or newline separated string
    other_raw = obj.get('OtherDesignations', '') or obj.get('OtherNames', '') or ''
    other_names = [n.strip() for n in re.split(r'[,\n]', str(other_raw)) if n.strip()] if other_raw else []

    period_str = _strip_flags(period_raw)
    maxmag_str = _strip_flags(maxmag_raw)
    minmag_str = _strip_flags(minmag_raw)
    epoch_str  = _strip_flags(epoch_raw)

    dist = arcsec_dist(ra_center, dec_center, ra_deg, dec_deg) if ra_center is not None else None
    dist_str = f"{dist:.1f}" if dist is not None else ''

    # Build summary for the treeview column
    vsx_parts = []
    if vartype:
        vsx_parts.append(str(vartype))
    if period_str:
        vsx_parts.append(f"P={period_str} d")
    if maxmag_str and minmag_str:
        mag_str = f"{maxmag_str}–{minmag_str}"
        if magband:
            mag_str += f" {magband}"
        vsx_parts.append(mag_str)
    elif maxmag_str:
        mag_str = f"max {maxmag_str}"
        if magband:
            mag_str += f" {magband}"
        vsx_parts.append(mag_str)

    oid = str(obj.get('OID', '') or '').strip()

    return {
        'Name':          name,
        'AUID':          auid,
        'RA_hms':        deg_to_hms(ra_deg),
        'Dec_dms':       deg_to_dms(dec_deg),
        'Summary':       '   '.join(vsx_parts),
        'VarType':       str(vartype),
        'Period':        period_str,
        'MaxMag':        maxmag_str,
        'MinMag':        minmag_str,
        'MagBand':       str(magband).strip(),
        'Constellation': str(constellation).strip(),
        'SpectralType':  str(spectral_type).strip(),
        'Discoverer':    str(discoverer).strip(),
        'Epoch':         epoch_str,
        'RiseDuration':  str(rise_dur).strip(),
        'OtherNames':    other_names,
        'Dist_arcsec':   dist_str,
        '_vsx_oid':      oid,
        '_ra_deg':       ra_deg,
        '_dec_deg':      dec_deg,
    }


def query_vsx_by_name(name, period_min, period_max, mag_min, mag_max, status_callback):
    """Name-based VSX query."""
    if status_callback:
        status_callback("Querying AAVSO VSX...")

    url = VSX_BASE
    params = {
        'view': 'api.object',
        'format': 'json',
        'ident': name,
    }
    r = requests.get(url, params=params, timeout=30)
    r.raise_for_status()
    j = r.json()
    obj = j.get('VSXObject')
    if not obj:
        return []

    result = _parse_vsx_object(obj)
    if not _passes_filters(result['Period'], result['MaxMag'],
                           period_min, period_max, mag_min, mag_max):
        return []
    # Remove internal keys
    result.pop('_ra_deg', None)
    result.pop('_dec_deg', None)
    return [result]


def query_vsx(ra_deg, dec_deg, radius_arcmin, period_min, period_max,
              mag_min, mag_max, status_callback):
    """Coordinate-based VSX query."""
    if status_callback:
        status_callback("Querying AAVSO VSX...")

    radius_deg = radius_arcmin / 60.0
    params = {
        'view': 'api.list',
        'format': 'json',
        'ra': ra_deg,
        'dec': dec_deg,
        'radius': radius_deg,
        'coords': 'decimal',
    }
    r = requests.get(VSX_BASE, params=params, timeout=30)
    r.raise_for_status()
    j = r.json()

    raw = j.get('VSXObjects', {}).get('VSXObject', [])
    if isinstance(raw, dict):
        raw = [raw]

    results = []
    for obj in raw:
        result = _parse_vsx_object(obj, ra_center=ra_deg, dec_center=dec_deg)
        if not _passes_filters(result['Period'], result['MaxMag'],
                               period_min, period_max, mag_min, mag_max):
            continue
        result.pop('_ra_deg', None)
        result.pop('_dec_deg', None)
        results.append(result)

    return results


# ──────────────────────────────────────────────────────────────
# VizieR / GCVS queries
# ──────────────────────────────────────────────────────────────

VIZIER_TAP = "https://tapvizier.cds.unistra.fr/TAPVizieR/tap/sync"

def _tmass_name_from_coords(ra_f, dec_f):
    """Build a 2MASS-style designation from decimal degrees RA/Dec."""
    if ra_f is None or dec_f is None:
        return ''
    # RA → HHMMSS.ss
    h_tot = ra_f / 15.0
    h = int(h_tot); m_tot = (h_tot - h) * 60; m = int(m_tot); s = (m_tot - m) * 60
    ra_s = f"{h:02d}{m:02d}{s:05.2f}"
    # Dec → ±DDMMSS.s
    neg = dec_f < 0; dabs = abs(dec_f)
    d = int(dabs); m2_tot = (dabs - d) * 60; m2 = int(m2_tot); s2 = (m2_tot - m2) * 60
    dec_s = f"{'-' if neg else '+'}{d:02d}{m2:02d}{s2:04.1f}"
    return f"2MASS J{ra_s}{dec_s}"


def _tmass_row_to_result(row, ra_center=None, dec_center=None):
    """Convert a 2MASS TAP row to result dict.
    Row: (RAJ2000, DEJ2000, 2MASS, Jmag, Hmag, Kmag,
          e_Jmag, e_Hmag, e_Kmag, Qflg, Rflg, Cflg)
    """
    def _f(v):
        try: return float(v) if v is not None else None
        except: return None
    def _mag(v):
        try: return f"{float(v):.3f}" if v is not None else ''
        except: return ''
    def _s(v):
        return str(v).strip() if v is not None else ''

    ra_f  = _f(row[0])
    dec_f = _f(row[1])
    desig = _s(row[2]) if len(row) > 2 else ''
    name  = f"2MASS J{desig}" if desig else _tmass_name_from_coords(ra_f, dec_f)
    dist  = arcsec_dist(ra_center, dec_center, ra_f, dec_f) if ra_center is not None else None

    return {
        'Name':        name,
        'RA_hms':      deg_to_hms(ra_f),
        'Dec_dms':     deg_to_dms(dec_f),
        'Designation': desig,
        'Jmag':        _mag(row[3] if len(row) > 3 else None),
        'Hmag':        _mag(row[4] if len(row) > 4 else None),
        'Kmag':        _mag(row[5] if len(row) > 5 else None),
        'e_Jmag':      _mag(row[6] if len(row) > 6 else None),
        'e_Hmag':      _mag(row[7] if len(row) > 7 else None),
        'e_Kmag':      _mag(row[8] if len(row) > 8 else None),
        'Qflg':        _s(row[9])  if len(row) > 9  else '',
        'Rflg':        _s(row[10]) if len(row) > 10 else '',
        'Cflg':        _s(row[11]) if len(row) > 11 else '',
        'Dist_arcsec': f"{dist:.1f}" if dist is not None else '',
        '_ra_deg':     ra_f,
        '_dec_deg':    dec_f,
    }


def _parse_votable(text, service='TAP'):
    """
    Parse a VOTable XML response from a TAP service.
    Raises RuntimeError with the server's error message if QUERY_STATUS=ERROR.
    Returns a list of row-lists (TD text values, None for empty cells) on success.
    """
    import xml.etree.ElementTree as ET
    try:
        root = ET.fromstring(text)
    except ET.ParseError as e:
        raise RuntimeError(f"{service}: could not parse response XML: {e}")

    # Check for QUERY_STATUS ERROR (namespace-agnostic tag match)
    for elem in root.iter():
        local = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
        if local == 'INFO' and elem.get('name') == 'QUERY_STATUS':
            if elem.get('value', '').upper() == 'ERROR':
                msg = (elem.text or '').strip() or 'Unknown query error'
                raise RuntimeError(f"{service}: {msg}")

    # Extract TABLEDATA rows
    rows = []
    for elem in root.iter():
        local = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
        if local == 'TR':
            row = []
            for child in elem:
                cloc = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                if cloc == 'TD':
                    row.append(child.text)   # None if cell is empty
            rows.append(row)
    return rows


def _parse_gaia_csv(text):
    """
    Parse a CSV response from the Gaia ESAC TAP server.
    Returns a list of dicts keyed by column name (from the CSV header row).
    Empty cells become None. Raises RuntimeError on server errors.
    """
    import csv, io
    lines = text.strip().splitlines()
    if not lines:
        return []
    first = lines[0].strip()
    if first.startswith('<'):
        raise RuntimeError(f"Gaia: {first[:200]}")
    reader = csv.DictReader(io.StringIO(text))
    return [{k: (v if v != '' else None) for k, v in row.items()} for row in reader]


def _vizier_post(adql):
    """POST an ADQL query to VizieR TAP, parse VOTable, return data rows."""
    params = {'REQUEST': 'doQuery', 'LANG': 'ADQL', 'FORMAT': 'votable', 'QUERY': adql}
    r = requests.post(VIZIER_TAP, data=params, timeout=60)
    # Don't raise_for_status first — VizieR returns 400 with a VOTable body
    # that contains the actual error message; let _parse_votable extract it.
    try:
        return _parse_votable(r.text, service='VizieR')
    except RuntimeError:
        raise
    except Exception:
        r.raise_for_status()   # fallback for truly non-XML responses
        return []


def _sort_by_dist(results):
    """Sort result rows by Dist_arcsec ascending; rows with no distance go last."""
    return sorted(results, key=lambda r: float(r['Dist_arcsec']) if r.get('Dist_arcsec') else 999)


def _tmass_adql_cone(ra_deg, dec_deg, radius_deg):
    return f"""
SELECT t.RAJ2000, t.DEJ2000, t."2MASS", t.Jmag, t.Hmag, t.Kmag,
       t.e_Jmag, t.e_Hmag, t.e_Kmag, t.Qflg, t.Rflg, t.Cflg
FROM "II/246/out" AS t
WHERE CONTAINS(POINT('ICRS', t.RAJ2000, t.DEJ2000),
               CIRCLE('ICRS', {ra_deg}, {dec_deg}, {radius_deg})) = 1
""".strip()


def query_tmass_by_name(name, radius_arcmin, status_callback):
    """Name-based 2MASS query: resolve via SIMBAD then cone search."""
    if status_callback:
        status_callback("Resolving name for VizieR/2MASS...")
    ra_deg, dec_deg = _resolve_name_simbad(name)
    if ra_deg is None:
        return []
    if status_callback:
        status_callback("Querying VizieR / 2MASS...")
    data = _vizier_post(_tmass_adql_cone(ra_deg, dec_deg, radius_arcmin / 60.0))
    results = []
    for row in data:
        r = _tmass_row_to_result(row, ra_center=ra_deg, dec_center=dec_deg)
        r.pop('_ra_deg', None); r.pop('_dec_deg', None)
        results.append(r)
    return _sort_by_dist(results)


def query_tmass(ra_deg, dec_deg, radius_arcmin, status_callback):
    """Coordinate-based 2MASS query via VizieR TAP."""
    if status_callback:
        status_callback("Querying VizieR / 2MASS...")
    radius_deg = radius_arcmin / 60.0
    data = _vizier_post(_tmass_adql_cone(ra_deg, dec_deg, radius_deg))
    results = []
    for row in data:
        r = _tmass_row_to_result(row, ra_center=ra_deg, dec_center=dec_deg)
        r.pop('_ra_deg', None); r.pop('_dec_deg', None)
        results.append(r)
    return _sort_by_dist(results)


# ── AllWISE ───────────────────────────────────────────────────

def _wise_name_from_coords(ra_f, dec_f):
    if ra_f is None or dec_f is None:
        return ''
    ra_h  = int(ra_f / 15)
    ra_m  = int((ra_f / 15 - ra_h) * 60)
    ra_s  = ((ra_f / 15 - ra_h) * 60 - ra_m) * 60
    sign  = '+' if dec_f >= 0 else '-'
    ad    = abs(dec_f)
    dd    = int(ad)
    dm    = int((ad - dd) * 60)
    ds    = ((ad - dd) * 60 - dm) * 60
    return f"WISE J{ra_h:02d}{ra_m:02d}{ra_s:05.2f}{sign}{dd:02d}{dm:02d}{ds:04.1f}"


def _wise_adql_cone(ra_deg, dec_deg, radius_deg):
    return f"""
SELECT t.RAJ2000, t.DEJ2000, t.AllWISE,
       t.W1mag, t.W2mag, t.W3mag, t.W4mag,
       t.e_W1mag, t.e_W2mag, t.e_W3mag, t.e_W4mag,
       t.Jmag, t.Hmag, t.Kmag,
       t.qph, t.ccf, t.ex, t.var, t.pmRA, t.pmDE
FROM "II/328/allwise" AS t
WHERE CONTAINS(POINT('ICRS', t.RAJ2000, t.DEJ2000),
               CIRCLE('ICRS', {ra_deg}, {dec_deg}, {radius_deg})) = 1
""".strip()


def _wise_row_to_result(row, ra_center=None, dec_center=None):
    def _mag(v):
        try: return f"{float(v):.3f}" if v is not None else ''
        except: return ''
    def _pm(v):
        try: return f"{float(v):.1f}" if v is not None else ''
        except: return ''
    def _s(v):
        return str(v).strip() if v is not None else ''
    try: ra_f  = float(row[0])
    except: ra_f = None
    try: dec_f = float(row[1])
    except: dec_f = None
    desig = _s(row[2]) if len(row) > 2 else ''
    name  = f"WISE J{desig}" if desig else _wise_name_from_coords(ra_f, dec_f)
    dist  = arcsec_dist(ra_center, dec_center, ra_f, dec_f) if ra_center is not None else None
    return {
        'Name':        name,
        'RA_hms':      deg_to_hms(ra_f),
        'Dec_dms':     deg_to_dms(dec_f),
        'Designation': desig,
        'W1mag':       _mag(row[3]  if len(row) > 3  else None),
        'W2mag':       _mag(row[4]  if len(row) > 4  else None),
        'W3mag':       _mag(row[5]  if len(row) > 5  else None),
        'W4mag':       _mag(row[6]  if len(row) > 6  else None),
        'e_W1mag':     _mag(row[7]  if len(row) > 7  else None),
        'e_W2mag':     _mag(row[8]  if len(row) > 8  else None),
        'e_W3mag':     _mag(row[9]  if len(row) > 9  else None),
        'e_W4mag':     _mag(row[10] if len(row) > 10 else None),
        'J2m':         _mag(row[11] if len(row) > 11 else None),
        'H2m':         _mag(row[12] if len(row) > 12 else None),
        'K2m':         _mag(row[13] if len(row) > 13 else None),
        'qph':         _s(row[14]  if len(row) > 14 else None),
        'ccf':         _s(row[15]  if len(row) > 15 else None),
        'ex':          _s(row[16]  if len(row) > 16 else None),
        'var':         _s(row[17]  if len(row) > 17 else None),
        'pmRA':        _pm(row[18] if len(row) > 18 else None),
        'pmDE':        _pm(row[19] if len(row) > 19 else None),
        'Dist_arcsec': f"{dist:.1f}" if dist is not None else '',
    }


def query_wise_by_name(name, radius_arcmin, status_callback):
    if status_callback:
        status_callback("Resolving name for AllWISE...")
    ra_deg, dec_deg = _resolve_name_simbad(name)
    if ra_deg is None:
        return []
    if status_callback:
        status_callback("Querying AllWISE...")
    data = _vizier_post(_wise_adql_cone(ra_deg, dec_deg, radius_arcmin / 60.0))
    return _sort_by_dist([_wise_row_to_result(row, ra_center=ra_deg, dec_center=dec_deg) for row in data])


def query_wise(ra_deg, dec_deg, radius_arcmin, status_callback):
    if status_callback:
        status_callback("Querying AllWISE...")
    data = _vizier_post(_wise_adql_cone(ra_deg, dec_deg, radius_arcmin / 60.0))
    return _sort_by_dist([_wise_row_to_result(row, ra_center=ra_deg, dec_center=dec_deg) for row in data])


# ── APASS DR9 ─────────────────────────────────────────────────

def _apass_name_from_coords(ra_f, dec_f):
    if ra_f is None or dec_f is None:
        return ''
    sign = '+' if dec_f >= 0 else '-'
    return f"APASS J{ra_f:09.5f}{sign}{abs(dec_f):08.5f}"


def _apass_adql_cone(ra_deg, dec_deg, radius_deg):
    # NOTE: Sloan columns (g'mag, r'mag, i'mag) cause VizieR 400 errors — never add them.
    return f"""
SELECT t.RAJ2000, t.DEJ2000, t.Vmag, t.Bmag,
       t.e_Vmag, t.e_Bmag, t."B-V", t."e_B-V", t.nobs, t.mobs
FROM "II/336/apass9" AS t
WHERE CONTAINS(POINT('ICRS', t.RAJ2000, t.DEJ2000),
               CIRCLE('ICRS', {ra_deg}, {dec_deg}, {radius_deg})) = 1
""".strip()


def _apass_row_to_result(row, ra_center=None, dec_center=None):
    def _mag(v):
        try: return f"{float(v):.3f}" if v is not None else ''
        except: return ''
    def _int(v):
        try: return str(int(float(v))) if v is not None else ''
        except: return ''
    try: ra_f  = float(row[0])
    except: ra_f = None
    try: dec_f = float(row[1])
    except: dec_f = None
    dist = arcsec_dist(ra_center, dec_center, ra_f, dec_f) if ra_center is not None else None
    return {
        'Name':        _apass_name_from_coords(ra_f, dec_f),
        'RA_hms':      deg_to_hms(ra_f),
        'Dec_dms':     deg_to_dms(dec_f),
        'Vmag':        _mag(row[2] if len(row) > 2 else None),
        'Bmag':        _mag(row[3] if len(row) > 3 else None),
        'e_Vmag':      _mag(row[4] if len(row) > 4 else None),
        'e_Bmag':      _mag(row[5] if len(row) > 5 else None),
        'BV':          _mag(row[6] if len(row) > 6 else None),
        'e_BV':        _mag(row[7] if len(row) > 7 else None),
        'nobs':        _int(row[8] if len(row) > 8 else None),
        'mobs':        _int(row[9] if len(row) > 9 else None),
        'Dist_arcsec': f"{dist:.1f}" if dist is not None else '',
    }


def query_apass_by_name(name, radius_arcmin, status_callback):
    if status_callback:
        status_callback("Resolving name for APASS...")
    ra_deg, dec_deg = _resolve_name_simbad(name)
    if ra_deg is None:
        return []
    if status_callback:
        status_callback("Querying APASS DR9...")
    data = _vizier_post(_apass_adql_cone(ra_deg, dec_deg, radius_arcmin / 60.0))
    return _sort_by_dist([_apass_row_to_result(row, ra_center=ra_deg, dec_center=dec_deg) for row in data])


def query_apass(ra_deg, dec_deg, radius_arcmin, status_callback):
    if status_callback:
        status_callback("Querying APASS DR9...")
    data = _vizier_post(_apass_adql_cone(ra_deg, dec_deg, radius_arcmin / 60.0))
    return _sort_by_dist([_apass_row_to_result(row, ra_center=ra_deg, dec_center=dec_deg) for row in data])


# ── Tycho-2 ───────────────────────────────────────────────────

def _tycho2_adql_cone(ra_deg, dec_deg, radius_deg):
    return f"""
SELECT t.RAmdeg, t.DEmdeg, t.BTmag, t.VTmag, t.pmRA, t.pmDE,
       t.TYC1, t.TYC2, t.TYC3,
       t.e_BTmag, t.e_VTmag, t.e_pmRA, t.e_pmDE, t.HIP, t.prox
FROM "I/259/tyc2" AS t
WHERE CONTAINS(POINT('ICRS', t.RAmdeg, t.DEmdeg),
               CIRCLE('ICRS', {ra_deg}, {dec_deg}, {radius_deg})) = 1
""".strip()


def _tycho2_row_to_result(row, ra_center=None, dec_center=None):
    def _mag(v):
        try: return f"{float(v):.3f}" if v is not None else ''
        except: return ''
    def _pm(v):
        try: return f"{float(v):.1f}" if v is not None else ''
        except: return ''
    def _s(v):
        return str(v).strip() if v is not None else ''
    try: ra_f  = float(row[0])
    except: ra_f = None
    try: dec_f = float(row[1])
    except: dec_f = None
    try:
        tyc1 = str(int(float(row[6]))) if row[6] is not None else ''
        tyc2 = str(int(float(row[7]))) if row[7] is not None else ''
        tyc3 = str(int(float(row[8]))) if row[8] is not None else ''
        tyc_name = f"TYC {tyc1}-{tyc2}-{tyc3}" if tyc1 else ''
    except Exception:
        tyc_name = ''
    dist = arcsec_dist(ra_center, dec_center, ra_f, dec_f) if ra_center is not None else None
    return {
        'Name':        tyc_name,
        'RA_hms':      deg_to_hms(ra_f),
        'Dec_dms':     deg_to_dms(dec_f),
        'BTmag':       _mag(row[2]  if len(row) > 2  else None),
        'VTmag':       _mag(row[3]  if len(row) > 3  else None),
        'pmRA':        _pm(row[4]   if len(row) > 4  else None),
        'pmDE':        _pm(row[5]   if len(row) > 5  else None),
        'e_BTmag':     _mag(row[9]  if len(row) > 9  else None),
        'e_VTmag':     _mag(row[10] if len(row) > 10 else None),
        'e_pmRA':      _pm(row[11]  if len(row) > 11 else None),
        'e_pmDE':      _pm(row[12]  if len(row) > 12 else None),
        'HIP':         _s(row[13]   if len(row) > 13 else None),
        'prox':        _s(row[14]   if len(row) > 14 else None),
        'Dist_arcsec': f"{dist:.1f}" if dist is not None else '',
    }


def query_tycho2_by_name(name, radius_arcmin, status_callback):
    if status_callback:
        status_callback("Resolving name for Tycho-2...")
    ra_deg, dec_deg = _resolve_name_simbad(name)
    if ra_deg is None:
        return []
    if status_callback:
        status_callback("Querying Tycho-2...")
    data = _vizier_post(_tycho2_adql_cone(ra_deg, dec_deg, radius_arcmin / 60.0))
    return _sort_by_dist([_tycho2_row_to_result(row, ra_center=ra_deg, dec_center=dec_deg) for row in data])


def query_tycho2(ra_deg, dec_deg, radius_arcmin, status_callback):
    if status_callback:
        status_callback("Querying Tycho-2...")
    data = _vizier_post(_tycho2_adql_cone(ra_deg, dec_deg, radius_arcmin / 60.0))
    return _sort_by_dist([_tycho2_row_to_result(row, ra_center=ra_deg, dec_center=dec_deg) for row in data])


# ── Washington Double Star Catalog (WDS) ──────────────────────

def _parse_wds_coords(name):
    """Parse RA/Dec from a WDS designation (HHMMM±DDMM).
    Accepts bare designations, 'WDS HHMMM±DDMM', or 'WDS J...' forms.
    Returns (ra_deg, dec_deg) or (None, None) if not a WDS designation."""
    s = re.sub(r'^WDS\s*J?', '', name.strip(), flags=re.IGNORECASE).strip()
    m = re.match(r'^(\d{2})(\d{2})(\d)([+-])(\d{2})(\d{2})', s)
    if not m:
        return None, None
    hh, mm, dm, sign, dd, dmm = m.groups()
    ra_deg  = (int(hh) + (int(mm) + int(dm) / 10.0) / 60.0) * 15.0
    dec_deg = (int(dd) + int(dmm) / 60.0) * (-1 if sign == '-' else 1)
    return ra_deg, dec_deg


def _parse_wds_designation(name):
    """Extract the bare HHMMM±DDMM WDS designation from any input form.
    Returns the 10-char string (e.g. '03069+3952'), or None if not a WDS name."""
    s = re.sub(r'^WDS\s*J?', '', name.strip(), flags=re.IGNORECASE).strip()
    m = re.match(r'^(\d{5}[+-]\d{4})', s)
    return m.group(1) if m else None


def _wds_adql_by_designation(desig):
    safe = desig.replace("'", "''")
    return f"""
SELECT t.RAJ2000, t.DEJ2000, t.WDS, t.Disc, t.Comp,
       t.Obs1, t.Obs2, t.Nobs, t.pa1, t.pa2, t.sep1, t.sep2,
       t.mag1, t.mag2, t.SpType
FROM "B/wds/wds" AS t
WHERE t.WDS = '{safe}'
""".strip()


def _wds_adql_cone(ra_deg, dec_deg, radius_deg):
    # RAJ2000/DEJ2000 are in degrees despite catalog description saying "hours"
    return f"""
SELECT t.RAJ2000, t.DEJ2000, t.WDS, t.Disc, t.Comp,
       t.Obs1, t.Obs2, t.Nobs, t.pa1, t.pa2, t.sep1, t.sep2,
       t.mag1, t.mag2, t.SpType
FROM "B/wds/wds" AS t
WHERE CONTAINS(POINT('ICRS', t.RAJ2000, t.DEJ2000),
               CIRCLE('ICRS', {ra_deg}, {dec_deg}, {radius_deg})) = 1
""".strip()


def _wds_row_to_result(row, ra_center=None, dec_center=None):
    def _f(v):
        try: return float(v) if v is not None else None
        except: return None
    def _s(v):
        return str(v).strip() if v is not None else ''
    def _mag(v):
        return str(v).strip() if v is not None else ''
    def _yr(v):
        try: return str(int(float(v))) if v is not None else ''
        except: return ''
    def _sep(v):
        return str(v).strip() if v is not None else ''
    def _pa(v):
        return (str(v).strip() + '°') if v is not None else ''

    ra_f   = _f(row[0])
    dec_f  = _f(row[1])
    wds    = _s(row[2])
    disc   = _s(row[3])
    comp   = _s(row[4])
    obs1   = _yr(row[5])
    obs2   = _yr(row[6])
    nobs   = _s(row[7])
    pa1    = _pa(row[8])
    pa2    = _pa(row[9])
    sep1   = _sep(row[10])
    sep2   = _sep(row[11])
    mag1   = _mag(row[12])
    mag2   = _mag(row[13])
    sptype = _s(row[14])

    name = ('WDS J' + wds if wds else '') + (' ' + comp if comp else '')
    dist = arcsec_dist(ra_center, dec_center, ra_f, dec_f) if ra_center is not None else None

    return {
        'Name':        name.strip(),
        'RA_hms':      deg_to_hms(ra_f),
        'Dec_dms':     deg_to_dms(dec_f),
        'WDS':         wds,
        'Disc':        disc,
        'Comp':        comp,
        'Obs1':        obs1,
        'Obs2':        obs2,
        'Nobs':        nobs,
        'PA1':         pa1,
        'PA2':         pa2,
        'Sep1':        sep1,
        'Sep2':        sep2,
        'Mag1':        mag1,
        'Mag2':        mag2,
        'SpType':      sptype,
        'Dist_arcsec': f"{dist:.1f}" if dist is not None else '',
        '_ra_deg':     ra_f,
        '_dec_deg':    dec_f,
    }


def query_wds_by_name(name, radius_arcmin, status_callback):
    if status_callback:
        status_callback("Querying WDS...")
    # Primary path: if name looks like a WDS designation, look up directly by WDS column.
    # This is exact and avoids any coordinate/radius ambiguity.
    desig = _parse_wds_designation(name)
    if desig is not None:
        data = _vizier_post(_wds_adql_by_designation(desig))
        if data:
            return [_wds_row_to_result(row) for row in data]
    # Fallback: resolve general star names via SIMBAD then cone search.
    if status_callback:
        status_callback("Resolving name for WDS...")
    ra_deg, dec_deg = _resolve_name_simbad(name)
    if ra_deg is None:
        return []
    if status_callback:
        status_callback("Querying WDS...")
    data = _vizier_post(_wds_adql_cone(ra_deg, dec_deg, radius_arcmin / 60.0))
    return _sort_by_dist([_wds_row_to_result(row, ra_center=ra_deg, dec_center=dec_deg) for row in data])


def query_wds(ra_deg, dec_deg, radius_arcmin, status_callback):
    if status_callback:
        status_callback("Querying WDS...")
    data = _vizier_post(_wds_adql_cone(ra_deg, dec_deg, radius_arcmin / 60.0))
    return _sort_by_dist([_wds_row_to_result(row, ra_center=ra_deg, dec_center=dec_deg) for row in data])


# ── Sixth Orbit Catalog (orb6) ────────────────────────────────

_ORB6_CACHE    = None
_ORB6_URL      = 'https://www.astro.gsu.edu/wds/orb6/orb6orbits.sql'
_ORB6_PLOT_BASE = 'https://www.astro.gsu.edu/wds/orb6/PNG/'

_ORB6_GRADES = {'1':'Definitive','2':'Good','3':'Reliable','4':'Preliminary','5':'Indeterminate'}

# Column indices in the pipe-delimited orb6 file
_O_RA,_O_DEC,_O_WDS,_O_DISC,_O_ADS,_O_HD,_O_HIP = 0,1,2,3,4,5,6
_O_MAG1,_O_MAG2 = 7,9
_O_P,_O_PUNIT,_O_EP = 11,12,13
_O_A,_O_EA = 14,16
_O_I,_O_EI = 17,18
_O_OM,_O_EOM = 19,21
_O_T0,_O_ET0 = 22,24
_O_E,_O_EE = 25,26
_O_OM2,_O_EOM2 = 27,28
_O_EQX,_O_LAST,_O_GRADE,_O_REF,_O_PNG = 29,30,31,33,34


def _parse_orb6_ra(s):
    s = s.strip()
    try:
        hh, mm, ss = int(s[0:2]), int(s[2:4]), float(s[4:])
        return (hh + mm/60.0 + ss/3600.0) * 15.0
    except Exception:
        return None


def _parse_orb6_dec(s):
    s = s.strip()
    try:
        sign = -1 if s[0] == '-' else 1
        dd, mm, ss = int(s[1:3]), int(s[3:5]), float(s[5:])
        return sign * (dd + mm/60.0 + ss/3600.0)
    except Exception:
        return None


def _load_orb6():
    global _ORB6_CACHE
    if _ORB6_CACHE is not None:
        return _ORB6_CACHE
    r = requests.get(_ORB6_URL, timeout=30)
    r.raise_for_status()
    cache = {}
    for line in r.text.splitlines()[2:]:
        parts = line.split('|')
        if len(parts) < 3:
            continue
        wds = parts[_O_WDS].strip()
        if wds:
            cache[wds] = parts
    _ORB6_CACHE = cache
    return cache


def _orb6_row_to_result(parts, ra_center=None, dec_center=None):
    def _s(i): return parts[i].strip() if i < len(parts) else ''
    ra_f  = _parse_orb6_ra(_s(_O_RA))
    dec_f = _parse_orb6_dec(_s(_O_DEC))
    wds   = _s(_O_WDS)
    disc  = _s(_O_DISC)
    name  = (('WDS J' + wds) if wds else '') + ((' ' + disc) if disc else '')
    dist  = arcsec_dist(ra_center, dec_center, ra_f, dec_f) if ra_center is not None else None
    return {
        'Name':        name.strip(),
        'RA_hms':      deg_to_hms(ra_f),
        'Dec_dms':     deg_to_dms(dec_f),
        'WDS':         wds,
        'Disc':        disc,
        'ADS':         _s(_O_ADS),
        'HD':          _s(_O_HD),
        'HIP':         _s(_O_HIP),
        'Mag1':        _s(_O_MAG1),
        'Mag2':        _s(_O_MAG2),
        'P':           _s(_O_P),
        'P_unit':      _s(_O_PUNIT),
        'e_P':         _s(_O_EP),
        'a':           _s(_O_A),
        'e_a':         _s(_O_EA),
        'i':           _s(_O_I),
        'e_i':         _s(_O_EI),
        'Omega':       _s(_O_OM),
        'e_Omega':     _s(_O_EOM),
        'T0':          _s(_O_T0),
        'e_T0':        _s(_O_ET0),
        'e_ecc':       _s(_O_E),
        'e_e_ecc':     _s(_O_EE),
        'omega':       _s(_O_OM2),
        'e_omega':     _s(_O_EOM2),
        'Equinox':     _s(_O_EQX),
        'LastObs':     _s(_O_LAST),
        'Grade':       _s(_O_GRADE),
        'Ref':         _s(_O_REF),
        '_pngfile':    _s(_O_PNG),
        'PlotURL':     (_ORB6_PLOT_BASE + _s(_O_PNG)) if _s(_O_PNG).strip() else '',
        'Dist_arcsec': f"{dist:.1f}" if dist is not None else '',
        '_ra_deg':     ra_f,
        '_dec_deg':    dec_f,
    }


def _orb6_cone_search(cache, ra_deg, dec_deg, radius_arcmin):
    results = []
    radius_arcsec = radius_arcmin * 60.0
    for parts in cache.values():
        ra_f  = _parse_orb6_ra(parts[_O_RA].strip() if len(parts) > _O_RA else '')
        dec_f = _parse_orb6_dec(parts[_O_DEC].strip() if len(parts) > _O_DEC else '')
        if ra_f is None or dec_f is None:
            continue
        dist = arcsec_dist(ra_deg, dec_deg, ra_f, dec_f)
        if dist is not None and dist <= radius_arcsec:
            results.append(_orb6_row_to_result(parts, ra_center=ra_deg, dec_center=dec_deg))
    return sorted(results, key=lambda r: float(r['Dist_arcsec']) if r['Dist_arcsec'] else 999)


def query_orb6_by_name(name, radius_arcmin, status_callback):
    import re
    if status_callback:
        status_callback("Loading Sixth Orbit Catalog...")
    cache = _load_orb6()
    designation = re.sub(r'^WDS\s*J?', '', name.strip(), flags=re.IGNORECASE).strip()
    if designation in cache:
        return [_orb6_row_to_result(cache[designation])]
    if status_callback:
        status_callback("Resolving name for Orb6...")
    ra_deg, dec_deg = _resolve_name_simbad(name)
    if ra_deg is None:
        return []
    return _orb6_cone_search(cache, ra_deg, dec_deg, radius_arcmin)


def query_orb6(ra_deg, dec_deg, radius_arcmin, status_callback):
    if status_callback:
        status_callback("Loading Sixth Orbit Catalog...")
    cache = _load_orb6()
    return _orb6_cone_search(cache, ra_deg, dec_deg, radius_arcmin)


# ──────────────────────────────────────────────────────────────
# Gaia DR3 queries
# ──────────────────────────────────────────────────────────────

GAIA_TAP = "https://gea.esac.esa.int/tap-server/tap/sync"

# SIMBAD measurement sections shown in the detail panel — each entry: (key, label)
SIMBAD_MEAS_OPTS = [
    ('spt',  'Spectral Type'),
    ('plx',  'Parallax'),
    ('dist', 'Distance'),
    ('pm',   'Proper Motion'),
    ('rv',   'Radial Velocity'),
    ('rot',  'Rotation (v sin i)'),
    ('flux', 'Photometry'),
    ('fe_h', 'Stellar Parameters'),
]

# Optional parameter groups — each entry: (key, label, [gaia_source columns])
GAIA_PARAM_OPTS = [
    ('bp_mag',   'BP Magnitude',     ['phot_bp_mean_mag']),
    ('rp_mag',   'RP Magnitude',     ['phot_rp_mean_mag']),
    ('bp_rp',    'BP-RP Color',      ['bp_rp']),
    ('parallax', 'Parallax',         ['parallax', 'parallax_error']),
    ('pm',       'Proper Motion',    ['pmra', 'pmra_error', 'pmdec', 'pmdec_error']),
    ('rv',       'Radial Velocity',  ['radial_velocity', 'radial_velocity_error']),
    ('ruwe',     'RUWE',             ['ruwe']),
]

def _normalize_name(name):
    """Apply catalog-specific name normalizations before SIMBAD queries."""
    import re
    # WDS identifiers: SIMBAD requires "WDS J..." — auto-insert J if missing
    name = re.sub(r'^(WDS\s+)(\d)', lambda m: m.group(1) + 'J' + m.group(2), name, flags=re.IGNORECASE)
    return name


def _resolve_name_simbad(name):
    """Resolve object name to (ra_deg, dec_deg).
    Primary: SIMBAD ident lookup. Fallback for 'Gaia DR3 XXXXXX' names: direct Gaia position lookup.
    """
    name = _normalize_name(name)
    safe = name.replace("'", "''")

    def _run(adql):
        params = {'REQUEST': 'doQuery', 'LANG': 'ADQL', 'FORMAT': 'json', 'QUERY': adql}
        r = requests.post(SIMBAD_TAP, data=params, timeout=30)
        r.raise_for_status()
        data = r.json().get('data', [])
        if data and data[0][0] is not None:
            return float(data[0][0]), float(data[0][1])
        return None, None

    try:
        ra, dec = _run(
            f"SELECT b.ra, b.dec FROM basic b JOIN ident i ON b.oid = i.oidref WHERE i.id = '{safe}'"
        )
        if ra is not None:
            return ra, dec
        # LIKE fallback (catches e.g. WDS J20176+2622 stored as WDS J20176+2622AB)
        ra, dec = _run(
            f"SELECT b.ra, b.dec FROM basic b JOIN ident i ON b.oid = i.oidref WHERE i.id LIKE '{safe}%'"
        )
        if ra is not None:
            return ra, dec
    except Exception:
        pass

    # Gaia DR3 position fallback — avoids SIMBAD dependency for Gaia source-id names
    clean = name.strip()
    source_id_int = None
    if clean.lower().startswith('gaia dr3 '):
        candidate = clean[9:].strip()
        if candidate.isdigit():
            source_id_int = int(candidate)
    elif clean.isdigit() and len(clean) >= 10:
        source_id_int = int(clean)
    if source_id_int is not None:
        try:
            params = {
                'REQUEST': 'doQuery', 'LANG': 'ADQL', 'FORMAT': 'csv',
                'MAXREC': '1',
                'QUERY': f"SELECT ra, dec FROM gaiadr3.gaia_source WHERE source_id = {source_id_int}",
            }
            r = requests.post(GAIA_TAP, data=params, timeout=30)
            r.raise_for_status()
            rows = _parse_gaia_csv(r.text)
            if rows and rows[0].get('ra') is not None:
                return float(rows[0]['ra']), float(rows[0]['dec'])
        except Exception:
            pass

    return None, None


def query_gaia(ra_deg, dec_deg, radius_arcmin, status_callback, extra_cols=None):
    """Coordinate-based Gaia DR3 query — any stellar object in gaia_source."""
    if status_callback:
        status_callback("Querying Gaia DR3...")

    radius_deg = radius_arcmin / 60.0
    base_cols = ['source_id', 'ra', 'dec', 'phot_g_mean_mag']
    select = ', '.join(base_cols + (extra_cols or []))
    adql = f"""
SELECT {select}
FROM gaiadr3.gaia_source
WHERE ra  BETWEEN {ra_deg  - radius_deg} AND {ra_deg  + radius_deg}
  AND dec BETWEEN {dec_deg - radius_deg} AND {dec_deg + radius_deg}
""".strip()

    params = {
        'REQUEST': 'doQuery',
        'LANG': 'ADQL',
        'FORMAT': 'csv',
        'MAXREC': '2000',
        'QUERY': adql,
    }
    r = requests.post(GAIA_TAP, data=params, timeout=120)
    r.raise_for_status()
    data = _parse_gaia_csv(r.text)
    return _gaia_rows_to_results(data, ra_center=ra_deg, dec_center=dec_deg)


def _gaia_rows_to_results(rows, ra_center=None, dec_center=None):
    """Convert list of Gaia CSV dicts (keyed by column name) to result dicts for display."""
    def _f(v, fmt='.3f'):
        try:
            return format(float(v), fmt) if v is not None else ''
        except (ValueError, TypeError):
            return ''

    results = []
    for row in rows:
        ra_f  = float(row['ra'])  if row.get('ra')  is not None else None
        dec_f = float(row['dec']) if row.get('dec') is not None else None
        dist  = arcsec_dist(ra_center, dec_center, ra_f, dec_f) if ra_center is not None else None
        gmag    = _f(row.get('phot_g_mean_mag'))
        bprp    = _f(row.get('bp_rp'), '.3f')
        plx     = _f(row.get('parallax'),       '.4f')
        plx_err = _f(row.get('parallax_error'), '.4f')
        pmra    = _f(row.get('pmra'),   '.3f')
        pmdec   = _f(row.get('pmdec'),  '.3f')
        rv      = _f(row.get('radial_velocity'),       '.2f')
        ruwe    = _f(row.get('ruwe'), '.3f')

        gaia_parts = []
        if gmag:
            gaia_parts.append(f"G={gmag}")
        if bprp:
            gaia_parts.append(f"BP-RP={bprp}")
        if plx:
            gaia_parts.append(f"plx={plx}" + (f"±{plx_err}" if plx_err else '') + " mas")
        if pmra and pmdec:
            gaia_parts.append(f"PM=({pmra}, {pmdec}) mas/yr")
        if rv:
            gaia_parts.append(f"RV={rv} km/s")
        if ruwe:
            gaia_parts.append(f"RUWE={ruwe}")

        results.append({
            'source_id':   str(row.get('source_id') or ''),
            'RA_hms':      deg_to_hms(ra_f),
            'Dec_dms':     deg_to_dms(dec_f),
            'Summary':     '   '.join(gaia_parts),
            'GMag':        gmag,
            'BPMag':       _f(row.get('phot_bp_mean_mag')),
            'RPMag':       _f(row.get('phot_rp_mean_mag')),
            'BPRP':        bprp,
            'Parallax':    plx,
            'PlxErr':      plx_err,
            'PMRA':        pmra,
            'PMRAErr':     _f(row.get('pmra_error'),  '.3f'),
            'PMDec':       pmdec,
            'PMDecErr':    _f(row.get('pmdec_error'), '.3f'),
            'RV':          rv,
            'RVErr':       _f(row.get('radial_velocity_error'), '.2f'),
            'RUWE':        ruwe,
            'Dist_arcsec': f"{dist:.1f}" if dist is not None else '',
        })
    return results


def query_gaia_by_name(name, radius_arcmin, status_callback, extra_cols=None):
    """Name-based Gaia DR3 query.
    If name looks like a Gaia source_id (all digits or 'Gaia DR3 XXXXX'), do a direct
    source_id lookup.  Otherwise resolve via SIMBAD and do a cone search.
    """
    if status_callback:
        status_callback("Querying Gaia DR3...")

    extra = extra_cols or []

    # Detect raw source_id or "Gaia DR3 XXXXX" prefix
    clean = name.strip()
    source_id_int = None
    if clean.lower().startswith('gaia dr3 '):
        candidate = clean[9:].strip()
        if candidate.isdigit():
            source_id_int = int(candidate)
    elif clean.isdigit() and len(clean) >= 10:
        source_id_int = int(clean)

    if source_id_int is not None:
        base_cols = ['gs.source_id', 'gs.ra', 'gs.dec', 'gs.phot_g_mean_mag']
        extra_prefixed = [f'gs.{c}' for c in extra]
        select = ', '.join(base_cols + extra_prefixed)
        adql = f"""
SELECT {select}
FROM gaiadr3.gaia_source AS gs
WHERE gs.source_id = {source_id_int}
""".strip()
        params = {
            'REQUEST': 'doQuery', 'LANG': 'ADQL', 'FORMAT': 'csv',
            'MAXREC': '10', 'QUERY': adql,
        }
        r = requests.post(GAIA_TAP, data=params, timeout=60)
        r.raise_for_status()
        data = _parse_gaia_csv(r.text)
        return _gaia_rows_to_results(data)

    # Resolve via SIMBAD, then box filter against gaia_source.
    if status_callback:
        status_callback("Resolving name via SIMBAD for Gaia...")
    ra_deg, dec_deg = _resolve_name_simbad(name)
    if ra_deg is None:
        return []

    radius_deg = radius_arcmin / 60.0
    base_cols = ['source_id', 'ra', 'dec', 'phot_g_mean_mag']
    select = ', '.join(base_cols + extra)
    adql = f"""
SELECT {select}
FROM gaiadr3.gaia_source
WHERE ra  BETWEEN {ra_deg  - radius_deg} AND {ra_deg  + radius_deg}
  AND dec BETWEEN {dec_deg - radius_deg} AND {dec_deg + radius_deg}
""".strip()
    params = {
        'REQUEST': 'doQuery', 'LANG': 'ADQL', 'FORMAT': 'csv',
        'MAXREC': '200', 'QUERY': adql,
    }
    r = requests.post(GAIA_TAP, data=params, timeout=60)
    r.raise_for_status()
    data = _parse_gaia_csv(r.text)
    return _gaia_rows_to_results(data, ra_center=ra_deg, dec_center=dec_deg)


# ──────────────────────────────────────────────────────────────
# NASA Exoplanet Archive queries
# ──────────────────────────────────────────────────────────────

NEA_TAP = "https://exoplanetarchive.ipac.caltech.edu/TAP/sync"

_NEA_SELECT_BASE = """
pl_name, hostname, discoverymethod, disc_year, disc_facility,
pl_orbper, pl_orbsmax, pl_orbeccen, pl_orbincl, pl_tranmid, pl_imppar,
pl_rade, pl_bmasse, pl_eqt, pl_trandep, pl_trandur, pl_dens,
st_teff, st_logg, st_met, st_rad, st_mass, st_lum, st_age,
sy_snum, sy_pnum, sy_dist, sy_vmag, sy_kmag, sy_gaiamag,
disc_refname,
pl_orbper_reflink, pl_tranmid_reflink, pl_orbsmax_reflink,
pl_orbincl_reflink, pl_orbeccen_reflink, pl_imppar_reflink,
pl_rade_reflink, pl_bmasse_reflink, pl_eqt_reflink,
pl_trandep_reflink, pl_trandur_reflink, pl_dens_reflink,
st_teff_reflink, st_logg_reflink, st_met_reflink,
st_rad_reflink, st_mass_reflink, st_lum_reflink, st_age_reflink,
sy_dist_reflink, sy_vmag_reflink,
ra, dec
""".strip()

# Maps reflink column → human-readable parameter label (in display order)
_NEA_REF_LABELS = [
    ('disc_refname',       'Discovery'),
    ('pl_orbper_reflink',  'Period'),
    ('pl_tranmid_reflink', 'Transit epoch'),
    ('pl_orbsmax_reflink', 'Semi-major axis'),
    ('pl_orbincl_reflink', 'Inclination'),
    ('pl_orbeccen_reflink','Eccentricity'),
    ('pl_imppar_reflink',  'Impact param'),
    ('pl_rade_reflink',    'Planet radius'),
    ('pl_bmasse_reflink',  'Planet mass'),
    ('pl_eqt_reflink',     'Teq'),
    ('pl_trandep_reflink', 'Transit depth'),
    ('pl_trandur_reflink', 'Transit duration'),
    ('pl_dens_reflink',    'Planet density'),
    ('st_teff_reflink',    'Teff'),
    ('st_logg_reflink',    'log g'),
    ('st_met_reflink',     '[Fe/H]'),
    ('st_rad_reflink',     'R*'),
    ('st_mass_reflink',    'M*'),
    ('st_lum_reflink',     'Luminosity'),
    ('st_age_reflink',     'Stellar age'),
    ('sy_dist_reflink',    'Distance'),
    ('sy_vmag_reflink',    'V mag'),
]


def _nea_reflink_text(s):
    """Extract author-year display text from an NEA reflink HTML anchor."""
    import re as _re
    if not s:
        return ''
    m = _re.search(r'>([^<]+)</a>', str(s))
    return m.group(1).strip() if m else str(s).strip()


def _nea_reflink_url(s):
    """Extract the ADS URL from an NEA reflink HTML anchor."""
    import re as _re
    if not s:
        return None
    m = _re.search(r'href=["\']?([^"\'\s>]+)', str(s))
    return m.group(1) if m else None

def _parse_nea_csv(text):
    """Parse NEA TAP CSV response. Returns list of dicts keyed by column name."""
    import csv, io
    lines = text.strip().splitlines()
    if not lines:
        return []
    first = lines[0].strip()
    if first.startswith('<'):
        raise RuntimeError(f"NEA: {first[:300]}")
    reader = csv.DictReader(io.StringIO(text))
    return [{k.strip(): (v.strip() if v and v.strip() != '' else None)
             for k, v in row.items()} for row in reader]


def _nea_fmt_row(raw):
    """Normalize a raw NEA CSV dict for display (numbers formatted to reasonable precision)."""
    def _fmt(v, decimals=3):
        if v is None:
            return ''
        try:
            f = float(v)
            return f'{f:.{decimals}f}'
        except (ValueError, TypeError):
            return str(v)

    return {
        'pl_name':        raw.get('pl_name') or '',
        'hostname':       raw.get('hostname') or '',
        'discoverymethod':raw.get('discoverymethod') or '',
        'disc_year':      raw.get('disc_year') or '',
        'disc_facility':  raw.get('disc_facility') or '',
        'pl_orbper':      _fmt(raw.get('pl_orbper'), 5),
        'pl_orbsmax':     _fmt(raw.get('pl_orbsmax'), 5),
        'pl_orbeccen':    _fmt(raw.get('pl_orbeccen'), 4),
        'pl_orbincl':     _fmt(raw.get('pl_orbincl'), 2),
        'pl_tranmid':     _fmt(raw.get('pl_tranmid'), 4),
        'pl_imppar':      _fmt(raw.get('pl_imppar'), 3),
        'pl_rade':        _fmt(raw.get('pl_rade'), 3),
        'pl_bmasse':      _fmt(raw.get('pl_bmasse'), 3),
        'pl_eqt':         _fmt(raw.get('pl_eqt'), 0),
        'pl_trandep':     _fmt(raw.get('pl_trandep'), 6),
        'pl_trandur':     _fmt(raw.get('pl_trandur'), 4),
        'pl_dens':        _fmt(raw.get('pl_dens'), 3),
        'st_teff':        _fmt(raw.get('st_teff'), 0),
        'st_logg':        _fmt(raw.get('st_logg'), 3),
        'st_met':         _fmt(raw.get('st_met'), 3),
        'st_rad':         _fmt(raw.get('st_rad'), 3),
        'st_mass':        _fmt(raw.get('st_mass'), 3),
        'st_lum':         _fmt(raw.get('st_lum'), 3),
        'st_age':         _fmt(raw.get('st_age'), 2),
        'sy_snum':        raw.get('sy_snum') or '',
        'sy_pnum':        raw.get('sy_pnum') or '',
        'sy_dist':        _fmt(raw.get('sy_dist'), 1),
        'sy_vmag':        _fmt(raw.get('sy_vmag'), 2),
        'sy_kmag':        _fmt(raw.get('sy_kmag'), 2),
        'sy_gaiamag':     _fmt(raw.get('sy_gaiamag'), 2),
        'ra':             raw.get('ra') or '',
        'dec':            raw.get('dec') or '',
        **{col: _nea_reflink_text(raw.get(col)) for col, _ in _NEA_REF_LABELS},
        '_reflink_urls':  {col: _nea_reflink_url(raw.get(col))
                           for col, _ in _NEA_REF_LABELS if raw.get(col)},
    }


def _nea_query(where_clause, status_callback, top=None):
    """Run a NEA TAP query and return list of formatted planet dicts."""
    top_str = f"TOP {top} " if top else ""
    adql = f"SELECT {top_str}{_NEA_SELECT_BASE} FROM pscomppars WHERE {where_clause}"
    params = {
        'REQUEST': 'doQuery',
        'LANG':    'ADQL',
        'FORMAT':  'csv',
        'QUERY':   adql,
    }
    r = requests.post(NEA_TAP, data=params, timeout=60)
    r.raise_for_status()
    raw_rows = _parse_nea_csv(r.text)
    return [_nea_fmt_row(row) for row in raw_rows]


def query_nea_by_name(name, status_callback):
    """Search NEA by hostname or planet name."""
    if status_callback:
        status_callback("Querying NASA Exoplanet Archive...")
    if '*' in name:
        pattern = name.replace('*', '%').replace("'", "''")
        where = f"hostname LIKE '{pattern}' OR pl_name LIKE '{pattern}'"
        return _nea_query(where, status_callback, top=100)
    safe = name.replace("'", "''")
    where = f"hostname = '{safe}' OR pl_name = '{safe}'"
    results = _nea_query(where, status_callback)
    if not results:
        # Try partial match on hostname
        where2 = f"hostname LIKE '%{safe}%'"
        results = _nea_query(where2, status_callback)
    return results


def query_nea(ra_deg, dec_deg, radius_arcmin, status_callback):
    """Search NEA by coordinate box."""
    if status_callback:
        status_callback("Querying NASA Exoplanet Archive...")
    r = radius_arcmin / 60.0
    where = (f"ra BETWEEN {ra_deg - r} AND {ra_deg + r} "
             f"AND dec BETWEEN {dec_deg - r} AND {dec_deg + r}")
    return _nea_query(where, status_callback)


# ──────────────────────────────────────────────────────────────
# Main App
# ──────────────────────────────────────────────────────────────

BG    = "#1e1e2e"
PANEL = "#2a2a3e"
ACC   = "#7c9fd4"
FG    = "#cdd6f4"
ENT   = "#313244"
SEL   = "#45475a"

BANNER_WORK_BG = "#3a3210"
BANNER_WORK_FG = "#f9e2af"
BANNER_OK_BG   = "#1a3a1a"
BANNER_OK_FG   = "#a6e3a1"
BANNER_ERR_BG  = "#3a1f10"
BANNER_ERR_FG  = "#fab387"
BANNER_NS_BG   = PANEL
BANNER_NS_FG   = "#888aaa"

VERSION   = "1.2.0"
COPYRIGHT = "© Art Trail 2026"

# Auto-run server cooldown: pause for _COOLDOWN_SECS every _COOLDOWN_EVERY targets
_COOLDOWN_EVERY = 150
_COOLDOWN_SECS  = 45


class StarQueryApp(tk.Tk):

    TABS = [
        ('simbad', 'SIMBAD',    SIMBAD_COLS),
        ('vsx',    'AAVSO VSX', VSX_COLS),
        ('vizier', 'VizieR',    None),
        ('gaia',   'Gaia DR3',  GAIA_COLS),
        ('nea',    'NEA',       NEA_OVERVIEW_COLS),
    ]

    VIZIER_SUBTABS = [
        ('tmass',  '2MASS',    TMASS_COLS),
        ('wise',   'AllWISE',  WISE_COLS),
        ('apass',  'APASS',    APASS_COLS),
        ('tycho2', 'Tycho-2',  TYCHO2_COLS),
        ('wds',    'WDS',      WDS_COLS),
        ('orb6',   'Orb6',     ORB6_COLS),
    ]

    # Full field export definitions for each catalog tab (used by batch export)
    _BATCH_EXPORT_COLS = {
        'simbad': [
            ('Name','Name'), ('RA_hms','RA (hms)'), ('Dec_dms','Dec (dms)'),
            ('OType','Object Type Code'), ('OType_label','Object Type'), ('VarType','Var Sub-type'),
            ('Period','Period (d)'), ('MaxMag','Max Mag'), ('MinMag','Min Mag'), ('MagBand','Band'),
            ('Dist_arcsec','Dist (")'), ('N_refs','N Refs'),
        ],
        'vsx': [
            ('Name','Name'), ('AUID','AUID'), ('RA_hms','RA (hms)'), ('Dec_dms','Dec (dms)'),
            ('VarType','Var Type'), ('SpectralType','Spectral Type'), ('Period','Period (d)'),
            ('MaxMag','Max Mag'), ('MinMag','Min Mag'), ('MagBand','Band'),
            ('Epoch','Epoch (HJD)'), ('Constellation','Constellation'), ('Discoverer','Discoverer'),
            ('Dist_arcsec','Dist (")'),
        ],
        'tmass': [
            ('Name','2MASS ID'), ('RA_hms','RA (hms)'), ('Dec_dms','Dec (dms)'),
            ('Jmag','J mag'), ('e_Jmag','e_J'), ('Hmag','H mag'), ('e_Hmag','e_H'),
            ('Kmag','K mag'), ('e_Kmag','e_K'),
            ('Qflg','Quality'), ('Rflg','Read Flags'), ('Cflg','Contamination'),
            ('Dist_arcsec','Dist (")'),
        ],
        'wise': [
            ('Name','AllWISE ID'), ('RA_hms','RA (hms)'), ('Dec_dms','Dec (dms)'),
            ('W1mag','W1'), ('e_W1mag','e_W1'), ('W2mag','W2'), ('e_W2mag','e_W2'),
            ('W3mag','W3'), ('e_W3mag','e_W3'), ('W4mag','W4'), ('e_W4mag','e_W4'),
            ('J2m','2MASS J'), ('H2m','2MASS H'), ('K2m','2MASS K'),
            ('qph','Quality'), ('ccf','CC Flags'), ('ex','Ext Flag'), ('var','Var Flag'),
            ('pmRA','PM RA (mas/yr)'), ('pmDE','PM Dec (mas/yr)'), ('Dist_arcsec','Dist (")'),
        ],
        'apass': [
            ('Name','APASS ID'), ('RA_hms','RA (hms)'), ('Dec_dms','Dec (dms)'),
            ('Vmag','V mag'), ('e_Vmag','e_V'), ('Bmag','B mag'), ('e_Bmag','e_B'),
            ('BV','B-V'), ('e_BV','e_(B-V)'), ('nobs','N Fields'), ('mobs','N Meas'),
            ('Dist_arcsec','Dist (")'),
        ],
        'tycho2': [
            ('Name','Tycho-2 ID'), ('RA_hms','RA (hms)'), ('Dec_dms','Dec (dms)'),
            ('BTmag','BT mag'), ('e_BTmag','e_BT'), ('VTmag','VT mag'), ('e_VTmag','e_VT'),
            ('pmRA','PM RA (mas/yr)'), ('e_pmRA','e_PM RA'),
            ('pmDE','PM Dec (mas/yr)'), ('e_pmDE','e_PM Dec'),
            ('HIP','HIP'), ('prox',"Prox (0.1')"), ('Dist_arcsec','Dist (")'),
        ],
        'wds': [
            ('Name','WDS ID'), ('RA_hms','RA (hms)'), ('Dec_dms','Dec (dms)'),
            ('WDS','WDS Desig'), ('Disc','Discoverer'), ('Comp','Components'),
            ('Obs1','First Obs'), ('Obs2','Last Obs'), ('Nobs','N Obs'),
            ('PA1','PA First (°)'), ('PA2','PA Last (°)'),
            ('Sep1','Sep First (")'), ('Sep2','Sep Last (")'),
            ('Mag1','Mag 1'), ('Mag2','Mag 2'), ('SpType','Spectral Type'),
            ('Dist_arcsec','Dist (")'),
        ],
        'orb6': [
            ('Name','Orb6 ID'), ('RA_hms','RA (hms)'), ('Dec_dms','Dec (dms)'),
            ('WDS','WDS Desig'), ('Disc','Discoverer'), ('ADS','ADS'), ('HD','HD'), ('HIP','HIP'),
            ('Mag1','Mag 1'), ('Mag2','Mag 2'),
            ('P','Period'), ('P_unit','P Unit'), ('e_P','e_P'),
            ('T0','Epoch T0 (yr)'), ('e_T0','e_T0'),
            ('a','Semi-major axis (")'), ('e_a','e_a'),
            ('i','Inclination (°)'), ('e_i','e_i'),
            ('Omega','Omega (°)'), ('e_Omega','e_Omega'),
            ('e_ecc','Eccentricity'), ('e_e_ecc','e_Eccentricity'),
            ('omega','omega (°)'), ('e_omega','e_omega'),
            ('Equinox','Equinox'), ('LastObs','Last Obs'), ('Grade','Grade'), ('Ref','Reference'),
            ('Dist_arcsec','Dist (")'), ('PlotURL','Plot URL'),
        ],
        'gaia': [
            ('source_id','Source ID'), ('RA_hms','RA (hms)'), ('Dec_dms','Dec (dms)'),
            ('GMag','G mag'), ('BPMag','BP mag'), ('RPMag','RP mag'), ('BPRP','BP-RP'),
            ('Parallax','Parallax (mas)'), ('PlxErr','e_Parallax'),
            ('PMRA','PM RA (mas/yr)'), ('PMRAErr','e_PM RA'),
            ('PMDec','PM Dec (mas/yr)'), ('PMDecErr','e_PM Dec'),
            ('RV','RV (km/s)'), ('RVErr','e_RV'), ('RUWE','RUWE'),
            ('Dist_arcsec','Dist (")'),
        ],
        'nea': [
            ('pl_name','Planet'), ('hostname','Host Star'), ('discoverymethod','Method'),
            ('disc_year','Year'), ('disc_facility','Facility'),
            ('pl_orbper','Orb Period (d)'), ('pl_orbsmax','Semi-major axis (AU)'),
            ('pl_orbeccen','Eccentricity'), ('pl_orbincl','Inclination (°)'),
            ('pl_imppar','Impact Param'), ('pl_rade','Radius (R⊕)'), ('pl_bmasse','Mass (M⊕)'),
            ('pl_eqt','Eq Temp (K)'), ('pl_trandep','Transit Depth (ppm)'),
            ('pl_trandur','Transit Dur (hr)'), ('pl_dens','Density (g/cm³)'),
            ('st_teff','Teff (K)'), ('st_logg','log g'), ('st_met','[Fe/H]'),
            ('st_rad','Stellar Rad (R☉)'), ('st_mass','Stellar Mass (M☉)'),
            ('st_lum','Luminosity (L☉)'), ('st_age','Age (Gyr)'),
            ('sy_dist','Distance (pc)'), ('sy_vmag','V mag'), ('sy_pnum','N Planets'),
        ],
    }

    def __init__(self):
        super().__init__()
        self.title(f"SOQyT — Stellar Object Query Tool  v{VERSION}")
        self.geometry("1700x1450")
        self.minsize(1300, 1100)
        self.configure(bg=BG)

        _all_keys = (
            [k for k, _, _ in self.TABS if k != 'vizier'] +
            [k for k, _, _ in self.VIZIER_SUBTABS]
        )
        self._results      = {k: [] for k in _all_keys}
        self._sort_col     = {k: None for k in _all_keys}
        self._sort_asc     = {k: True for k in _all_keys}
        self._anim_job     = {k: None for k in _all_keys}
        self._anim_idx     = {k: 0    for k in _all_keys}
        self._banner_label = {}
        self._banner_outer = {}
        self._progress_bar = {}
        self._trees        = {}
        self._tab_cols     = {}
        self._notebook     = None
        self._vizier_notebook = None
        self._cfg          = _load_config()
        self._active_simbad_row = None
        self._active_vsx_row    = None
        self._wildcard_mode     = False
        self._year_filter_job   = None
        # Batch query state
        self._batch_rows     = []
        self._batch_index    = -1
        self._batch_results  = {}
        self._batch_mode     = None
        self._batch_name_col  = None
        self._batch_ra_col    = None
        self._batch_dec_col   = None
        self._batch_label_col = None
        self._batch_radius    = 1.0
        self._auto_run           = False
        self._auto_run_after_id  = None
        self._batch_filename     = ''
        self._query_pending  = 0

        self._build_ui()
        self.bind('<Return>', lambda e: self._run_query())
        self.after(200, self._auto_size_pane)

    # ──────────────────────────────────────────────────────────
    # UI construction
    # ──────────────────────────────────────────────────────────

    def _setup_styles(self):
        style = ttk.Style(self)
        style.theme_use('default')

        style.configure('TFrame',       background=BG)
        style.configure('Panel.TFrame', background=PANEL)

        style.configure('TLabel',  background=BG,    foreground=FG, font=('Segoe UI', 11))
        style.configure('Panel.TLabel', background=PANEL, foreground=FG, font=('Segoe UI', 11))
        style.configure('Header.TLabel', background=PANEL, foreground=ACC,
                        font=('Segoe UI', 11, 'bold'))

        style.configure('TRadiobutton', background=PANEL, foreground=FG,
                        font=('Segoe UI', 11), indicatorcolor=ENT, selectcolor=ACC)
        style.map('TRadiobutton',
                  background=[('active', PANEL)],
                  foreground=[('active', FG)])

        style.configure('TCheckbutton', background=PANEL, foreground=FG,
                        font=('Segoe UI', 11))
        style.map('TCheckbutton',
                  background=[('active', PANEL)],
                  foreground=[('active', FG)])

        style.configure('TEntry', fieldbackground=ENT, foreground=FG,
                        insertcolor=FG, font=('Segoe UI', 11))

        style.configure('TCombobox', fieldbackground=ENT, foreground=FG,
                        background=ENT, selectbackground=SEL, insertcolor='white',
                        font=('Segoe UI', 11))
        style.map('TCombobox',
                  fieldbackground=[('readonly', ENT)],
                  foreground=[('readonly', FG)])

        style.configure('TSeparator', background=SEL)

        style.configure('TNotebook', background=BG, borderwidth=0)
        style.configure('TNotebook.Tab', background=PANEL, foreground=FG,
                        padding=[12, 6], font=('Segoe UI', 11))
        style.map('TNotebook.Tab',
                  background=[('selected', ACC)],
                  foreground=[('selected', BG)])

        style.configure('Treeview', background=BG, fieldbackground=BG,
                        foreground=FG, font=('Segoe UI', 18),
                        rowheight=44, borderwidth=0)
        style.configure('Treeview.Heading', background=PANEL, foreground=ACC,
                        font=('Segoe UI', 11, 'bold'), relief='flat')
        style.map('Treeview',
                  background=[('selected', SEL)],
                  foreground=[('selected', FG)])
        style.map('Treeview.Heading',
                  background=[('active', SEL)])

        style.configure('Vertical.TScrollbar',   background=PANEL, troughcolor=BG,
                        arrowcolor=FG,  borderwidth=0)
        style.configure('Horizontal.TScrollbar', background=PANEL, troughcolor=BG,
                        arrowcolor=FG,  borderwidth=0)

        style.configure('Accent.TButton', background=ACC, foreground=BG,
                        font=('Segoe UI', 11, 'bold'), padding=[8, 6])
        style.map('Accent.TButton',
                  background=[('active', FG)],
                  foreground=[('active', BG)])

        style.configure('TButton', background=PANEL, foreground=FG,
                        font=('Segoe UI', 11), padding=[8, 6])
        style.map('TButton',
                  background=[('active', SEL)],
                  foreground=[('active', FG)])

        style.configure('TProgressbar', troughcolor=ENT, background=ACC, borderwidth=0)

    def _build_ui(self):
        self._setup_styles()
        self._build_menu()
        self.columnconfigure(1, weight=1)
        self.rowconfigure(0, weight=1)

        # Left panel (fixed ~230px)
        left = tk.Frame(self, bg=PANEL, width=230)
        left.grid(row=0, column=0, sticky='nsew', padx=(8, 0), pady=8)
        left.grid_propagate(False)
        left.pack_propagate(False)
        self._build_left_panel(left)

        # Right panel
        right = ttk.Frame(self)
        right.grid(row=0, column=1, sticky='nsew', padx=8, pady=8)
        right.columnconfigure(0, weight=1)
        right.rowconfigure(0, weight=1)
        self._build_right_panel(right)

        # Footer: status bar (left) + copyright (right)
        footer = tk.Frame(self, bg=PANEL)
        footer.grid(row=1, column=0, columnspan=2, sticky='ew', pady=(0, 8), padx=8)
        footer.columnconfigure(0, weight=1)

        self._status_var = tk.StringVar(value="Ready.")
        tk.Label(footer, textvariable=self._status_var,
                 bg=PANEL, fg=FG, font=('Segoe UI', 20),
                 anchor='w', padx=16).grid(row=0, column=0, sticky='ew')

        tk.Label(footer, text=f"{COPYRIGHT}  |  github.com/ArtTrail/SOQyT",
                 bg=PANEL, fg=BANNER_NS_FG, font=('Segoe UI', 18),
                 anchor='e', padx=16).grid(row=0, column=1, sticky='e')

    def _collapsible_section(self, parent, title, default_open=True):
        """Return a body frame under a clickable toggle header."""
        open_state = [default_open]

        # Container always stays in parent's pack order; only body inside it is toggled.
        container = tk.Frame(parent, bg=PANEL)
        container.pack(fill='x')

        header = tk.Frame(container, bg=PANEL, cursor='hand2')
        header.pack(fill='x', padx=10, pady=(6, 0))

        arrow = tk.Label(header, text='▼' if default_open else '▶',
                         bg=PANEL, fg=ACC, font=('Segoe UI', 10, 'bold'), cursor='hand2')
        arrow.pack(side='left')

        tk.Label(header, text=f'  {title}', bg=PANEL, fg=ACC,
                 font=('Segoe UI', 11, 'bold'), cursor='hand2').pack(side='left')

        body = tk.Frame(container, bg=PANEL)
        if default_open:
            body.pack(fill='x')

        def toggle(e=None):
            open_state[0] = not open_state[0]
            if open_state[0]:
                body.pack(fill='x')
                arrow.config(text='▼')
            else:
                body.pack_forget()
                arrow.config(text='▶')

        for w in (header, arrow, header.winfo_children()[-1]):
            w.bind('<Button-1>', toggle)

        return body

    def _center_popup(self, pop):
        """Center a Toplevel on the main window and give it focus."""
        pop.update_idletasks()
        pw = pop.winfo_reqwidth()
        ph = pop.winfo_reqheight()
        x = self.winfo_rootx() + (self.winfo_width()  - pw) // 2
        y = self.winfo_rooty() + (self.winfo_height() - ph) // 2
        pop.geometry(f"+{x}+{y}")
        pop.focus_force()

    def _show_gaia_params_popup(self):
        """Open the Gaia DR3 parameters configuration popup."""
        if hasattr(self, '_gaia_popup') and self._gaia_popup.winfo_exists():
            self._gaia_popup.lift()
            self._gaia_popup.focus_force()
            return

        pop = tk.Toplevel(self)
        pop.title("Gaia DR3 Parameters")
        pop.configure(bg=PANEL)
        pop.transient(self)
        self._gaia_popup = pop

        tk.Label(pop, text="Gaia DR3 Parameters", bg=PANEL, fg=ACC,
                 font=('Segoe UI', 11, 'bold')).pack(padx=16, pady=(12, 6))

        ttk.Separator(pop, orient='horizontal').pack(fill='x', padx=16, pady=4)

        for key, label, _ in GAIA_PARAM_OPTS:
            tk.Checkbutton(pop, text=label, variable=self._gaia_param_vars[key],
                           bg=PANEL, fg=FG, selectcolor=ENT,
                           activebackground=PANEL, activeforeground=FG,
                           anchor='w', font=('Segoe UI', 11)).pack(fill='x', padx=16, pady=2)

        ttk.Separator(pop, orient='horizontal').pack(fill='x', padx=16, pady=(8, 4))
        ttk.Button(pop, text="Close", command=pop.destroy).pack(padx=16, pady=(0, 12))

        self._center_popup(pop)

    def _show_simbad_meas_popup(self):
        """Open the SIMBAD detail measurement sections configuration popup."""
        if hasattr(self, '_simbad_meas_popup') and self._simbad_meas_popup.winfo_exists():
            self._simbad_meas_popup.lift()
            self._simbad_meas_popup.focus_force()
            return

        pop = tk.Toplevel(self)
        pop.title("SIMBAD Detail Sections")
        pop.configure(bg=PANEL)
        pop.transient(self)
        self._simbad_meas_popup = pop

        tk.Label(pop, text="SIMBAD Detail Sections", bg=PANEL, fg=ACC,
                 font=('Segoe UI', 11, 'bold')).pack(padx=16, pady=(12, 6))
        ttk.Separator(pop, orient='horizontal').pack(fill='x', padx=16, pady=4)

        for key, label in SIMBAD_MEAS_OPTS:
            tk.Checkbutton(pop, text=label, variable=self._simbad_meas_vars[key],
                           bg=PANEL, fg=FG, selectcolor=ENT,
                           activebackground=PANEL, activeforeground=FG,
                           anchor='w', font=('Segoe UI', 11),
                           command=self._rerender_simbad_detail).pack(fill='x', padx=16, pady=2)

        ttk.Separator(pop, orient='horizontal').pack(fill='x', padx=16, pady=(8, 4))
        ttk.Button(pop, text="Close", command=pop.destroy).pack(padx=16, pady=(0, 12))

        self._center_popup(pop)

    def _rerender_simbad_detail(self):
        """Re-render the current SIMBAD detail row (called when measurement vars change)."""
        if self._active_tab_key() != 'simbad':
            return
        tv = self._trees.get('simbad')
        if tv is None:
            return
        sel = tv.selection()
        if not sel:
            return
        idx = int(sel[0])
        rows = self._results.get('simbad', [])
        if 0 <= idx < len(rows):
            self._render_simbad_detail(rows[idx])

    def _build_left_panel(self, parent):
        pad = {'padx': 10, 'pady': 3}

        # Header
        tk.Label(parent, text="SEARCH PARAMETERS", bg=PANEL, fg=ACC,
                 font=('Segoe UI', 11, 'bold')).pack(fill='x', padx=10, pady=(10, 4))

        ttk.Separator(parent, orient='horizontal').pack(fill='x', padx=10, pady=2)

        # Mode radio buttons
        self._mode_var = tk.StringVar(value='name')
        mode_frame = tk.Frame(parent, bg=PANEL)
        mode_frame.pack(fill='x', **pad)
        tk.Radiobutton(mode_frame, text="By Name", variable=self._mode_var, value='name',
                       bg=PANEL, fg=FG, selectcolor=ENT, activebackground=PANEL,
                       activeforeground=FG, font=('Segoe UI', 11),
                       command=self._toggle_mode).pack(side='left', padx=(0, 12))
        tk.Radiobutton(mode_frame, text="By Coordinates", variable=self._mode_var, value='coords',
                       bg=PANEL, fg=FG, selectcolor=ENT, activebackground=PANEL,
                       activeforeground=FG, font=('Segoe UI', 11),
                       command=self._toggle_mode).pack(side='left')

        # Name entry (with recent-search history dropdown)
        tk.Label(parent, text="Star Name:", bg=PANEL, fg=FG,
                 font=('Segoe UI', 10)).pack(fill='x', padx=10, pady=(6, 0))
        self._name_var = tk.StringVar()
        self._name_entry = ttk.Combobox(parent, textvariable=self._name_var,
                                        values=self._cfg.get('name_history', []))
        self._name_entry.pack(fill='x', **pad)
        self._name_hint = tk.Label(
            parent,
            text='e.g.  RR Lyr   WASP-24   GJ 3470   HD 216963   Gaia DR3 1234567890',
            bg=PANEL, fg=BANNER_NS_FG, font=('Segoe UI', 9, 'italic'),
            anchor='w', justify='left', wraplength=260)
        self._name_hint.pack(fill='x', padx=10, pady=(0, 1))

        self._resolved_var = tk.StringVar()
        self._resolved_label = tk.Label(
            parent, textvariable=self._resolved_var,
            bg=PANEL, fg=ACC, font=('Segoe UI', 9),
            anchor='w', padx=0)
        self._resolved_label.pack(fill='x', padx=10, pady=(0, 2))

        # RA / Dec / Radius entries
        tk.Label(parent, text="RA (deg or hms):", bg=PANEL, fg=FG,
                 font=('Segoe UI', 10)).pack(fill='x', padx=10, pady=(6, 0))
        self._ra_var = tk.StringVar()
        self._ra_entry = ttk.Entry(parent, textvariable=self._ra_var)
        self._ra_entry.pack(fill='x', **pad)

        tk.Label(parent, text="Dec (deg or dms):", bg=PANEL, fg=FG,
                 font=('Segoe UI', 10)).pack(fill='x', padx=10, pady=(4, 0))
        self._dec_var = tk.StringVar()
        self._dec_entry = ttk.Entry(parent, textvariable=self._dec_var)
        self._dec_entry.pack(fill='x', **pad)

        tk.Label(parent, text="Search Radius (arcmin):", bg=PANEL, fg=FG,
                 font=('Segoe UI', 10)).pack(fill='x', padx=10, pady=(4, 0))
        self._radius_var = tk.StringVar(value=self._cfg.get('last_radius', '0.1'))
        self._radius_entry = ttk.Entry(parent, textvariable=self._radius_var)
        self._radius_entry.pack(fill='x', **pad)

        ttk.Separator(parent, orient='horizontal').pack(fill='x', padx=10, pady=6)

        # Gaia param vars must exist before the DATA SOURCES rows are built
        self._gaia_param_vars = {}
        for key, label, _ in GAIA_PARAM_OPTS:
            self._gaia_param_vars[key] = tk.BooleanVar(value=True)

        # DATA SOURCES (collapsible)
        body_src = self._collapsible_section(parent, "DATA SOURCES")

        self._src_vars = {}
        for key, label, _ in self.TABS:
            var = tk.BooleanVar(value=True)
            self._src_vars[key] = var
            row_f = tk.Frame(body_src, bg=PANEL)
            row_f.pack(fill='x', padx=4, pady=1)
            tk.Checkbutton(row_f, text=label, variable=var,
                           bg=PANEL, fg=FG, selectcolor=ENT,
                           activebackground=PANEL, activeforeground=FG,
                           anchor='w', font=('Segoe UI', 11)).pack(side='left', fill='x', expand=True)
            if key == 'simbad':
                tk.Button(row_f, text='⚙', bg=PANEL, fg=ACC,
                          relief='flat', bd=0, font=('Segoe UI', 12),
                          activebackground=PANEL, activeforeground=FG,
                          cursor='hand2',
                          command=self._show_simbad_meas_popup).pack(side='right', padx=(0, 6))
                # SIMBAD filters — indented under SIMBAD row
                flt_f = tk.Frame(body_src, bg=PANEL)
                flt_f.pack(fill='x', padx=(24, 4), pady=(0, 6))
                tk.Label(flt_f, text="Object Type:", bg=PANEL, fg=FG,
                         font=('Segoe UI', 10)).pack(fill='x', pady=(2, 0))
                otype_vals = ['All', 'RR*', 'Ce*', 'dS*', 'RV*', 'LP*', 'SR*', 'Mi*', 'Ell*', 'Ro*', 'EB*', 'WV*']
                self._otype_var = tk.StringVar(value='All')
                self._otype_combo = ttk.Combobox(flt_f, textvariable=self._otype_var,
                                                  values=otype_vals, state='readonly')
                self._otype_combo.pack(fill='x', pady=2)
                tk.Label(flt_f, text="Period range (days):", bg=PANEL, fg=FG,
                         font=('Segoe UI', 10)).pack(fill='x', pady=(4, 0))
                pf = tk.Frame(flt_f, bg=PANEL)
                pf.pack(fill='x', pady=2)
                self._period_min_var = tk.StringVar()
                self._period_max_var = tk.StringVar()
                ttk.Entry(pf, textvariable=self._period_min_var, width=7).pack(side='left')
                tk.Label(pf, text=' – ', bg=PANEL, fg=FG, font=('Segoe UI', 10)).pack(side='left')
                ttk.Entry(pf, textvariable=self._period_max_var, width=7).pack(side='left')
                tk.Label(flt_f, text="Max Mag range:", bg=PANEL, fg=FG,
                         font=('Segoe UI', 10)).pack(fill='x', pady=(4, 0))
                mf = tk.Frame(flt_f, bg=PANEL)
                mf.pack(fill='x', pady=2)
                self._mag_min_var = tk.StringVar()
                self._mag_max_var = tk.StringVar()
                ttk.Entry(mf, textvariable=self._mag_min_var, width=7).pack(side='left')
                tk.Label(mf, text=' – ', bg=PANEL, fg=FG, font=('Segoe UI', 10)).pack(side='left')
                ttk.Entry(mf, textvariable=self._mag_max_var, width=7).pack(side='left')
                tk.Label(flt_f, text="Ref year range:", bg=PANEL, fg=FG,
                         font=('Segoe UI', 10)).pack(fill='x', pady=(4, 0))
                yf = tk.Frame(flt_f, bg=PANEL)
                yf.pack(fill='x', pady=2)
                self._ref_year_from_var = tk.StringVar()
                self._ref_year_to_var   = tk.StringVar(value=str(datetime.date.today().year))
                ttk.Entry(yf, textvariable=self._ref_year_from_var, width=7).pack(side='left')
                tk.Label(yf, text=' – ', bg=PANEL, fg=FG, font=('Segoe UI', 10)).pack(side='left')
                ttk.Entry(yf, textvariable=self._ref_year_to_var, width=7).pack(side='left')
                self._ref_year_from_var.trace_add('write', self._on_ref_year_filter_change)
                self._ref_year_to_var.trace_add('write', self._on_ref_year_filter_change)
            if key == 'gaia':
                tk.Button(row_f, text='⚙', bg=PANEL, fg=ACC,
                          relief='flat', bd=0, font=('Segoe UI', 12),
                          activebackground=PANEL, activeforeground=FG,
                          cursor='hand2',
                          command=self._show_gaia_params_popup).pack(side='right', padx=(0, 6))

        ttk.Separator(parent, orient='horizontal').pack(fill='x', padx=10, pady=8)

        # Buttons
        ttk.Button(parent, text="▶  Run Query", style='Accent.TButton',
                   command=self._run_query).pack(fill='x', padx=10, pady=2)
        ttk.Button(parent, text="⬇  Export to Excel",
                   command=self._export_excel).pack(fill='x', padx=10, pady=2)
        ttk.Button(parent, text="⬇  Export to CSV",
                   command=self._export_csv).pack(fill='x', padx=10, pady=2)
        ttk.Button(parent, text="✕  Clear Results",
                   command=self._clear_all).pack(fill='x', padx=10, pady=2)
        ttk.Button(parent, text="📂  Import Batch…",
                   command=self._import_batch).pack(fill='x', padx=10, pady=(2, 8))

        # Batch navigator (hidden until a batch is loaded)
        self._batch_nav_frame = tk.Frame(parent, bg=PANEL, relief='flat', bd=0)

        tk.Label(self._batch_nav_frame, text="─── BATCH QUERY ───────────────────",
                 bg=PANEL, fg=ACC, font=('Segoe UI', 9, 'bold')).pack(fill='x', padx=10, pady=(6, 2))

        nav_row = tk.Frame(self._batch_nav_frame, bg=PANEL)
        nav_row.pack(fill='x', padx=10, pady=2)
        self._batch_prev_btn = tk.Button(
            nav_row, text="◀", bg=PANEL, fg=FG, relief='flat', bd=0,
            font=('Segoe UI', 13), activebackground=PANEL, cursor='hand2',
            command=lambda: self._batch_go(-1))
        self._batch_prev_btn.pack(side='left')
        self._batch_nav_label = tk.Label(
            nav_row, text="", bg=PANEL, fg=FG,
            font=('Segoe UI', 10), anchor='center', wraplength=180)
        self._batch_nav_label.pack(side='left', fill='x', expand=True)
        self._batch_next_btn = tk.Button(
            nav_row, text="▶", bg=PANEL, fg=FG, relief='flat', bd=0,
            font=('Segoe UI', 13), activebackground=PANEL, cursor='hand2',
            command=lambda: self._batch_go(1))
        self._batch_next_btn.pack(side='right')

        self._batch_export_btn = ttk.Button(
            self._batch_nav_frame, text="⬇  Export Batch Results",
            command=self._export_batch)
        self._batch_export_btn.pack(fill='x', padx=10, pady=(4, 2))

        self._batch_cancel_auto_btn = tk.Button(
            self._batch_nav_frame, text="⏹  Cancel Auto-Run",
            bg=PANEL, fg='#e08080', relief='flat', bd=0,
            font=('Segoe UI', 10), cursor='hand2',
            activebackground=PANEL,
            command=self._cancel_auto_run)
        # not packed initially — shown only during auto-run

        self._batch_close_btn = tk.Button(
            self._batch_nav_frame, text="✕  Close Batch",
            bg=PANEL, fg='#888', relief='flat', bd=0,
            font=('Segoe UI', 10), cursor='hand2',
            activebackground=PANEL,
            command=self._close_batch)
        self._batch_close_btn.pack(padx=10, pady=(0, 8))

        self._toggle_mode()

    def _build_right_panel(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=1)

        pane = tk.PanedWindow(parent, orient='vertical', bg=BG,
                              sashwidth=5, sashrelief='flat', bd=0)
        pane.grid(row=0, column=0, sticky='nsew')
        self._right_pane = pane

        # Top pane: results notebook
        nb_frame = ttk.Frame(pane)
        pane.add(nb_frame, stretch='always', minsize=80)
        nb_frame.columnconfigure(0, weight=1)
        nb_frame.rowconfigure(0, weight=1)

        nb = ttk.Notebook(nb_frame)
        nb.grid(row=0, column=0, sticky='nsew')
        self._notebook = nb
        nb.bind('<<NotebookTabChanged>>', self._on_tab_changed)

        _NEA_NOTE = (
            "Source: NASA Exoplanet Archive — Planetary Systems Composite Parameters (pscomppars). "
            "One row per planet; each parameter drawn from the most authoritative reference "
            "as curated by the NEA team."
        )
        for key, label, cols in self.TABS:
            if key == 'vizier':
                self._build_vizier_tab(label)
            elif key == 'nea':
                self._build_tab(key, label, cols, note_text=_NEA_NOTE)
            else:
                self._build_tab(key, label, cols)

        # SIMBAD measurement section toggle vars
        self._simbad_meas_vars = {}
        for key, _ in SIMBAD_MEAS_OPTS:
            self._simbad_meas_vars[key] = tk.BooleanVar(value=True)

        # Bottom pane: detail panel
        detail_outer = tk.Frame(pane, bg=PANEL, bd=0)
        pane.add(detail_outer, stretch='always', minsize=150)
        detail_outer.columnconfigure(0, weight=1)
        detail_outer.rowconfigure(1, weight=1)

        tk.Label(detail_outer, text="Selected Star Detail", bg=PANEL, fg=ACC,
                 font=('Segoe UI', 10, 'bold')).grid(row=0, column=0, sticky='w', padx=8, pady=(4, 0))

        detail_inner = tk.Frame(detail_outer, bg=ENT)
        detail_inner.grid(row=1, column=0, sticky='nsew', padx=8, pady=(2, 6))
        detail_inner.columnconfigure(0, weight=1)
        detail_inner.rowconfigure(0, weight=1)

        self._detail_text = tk.Text(detail_inner, bg=ENT, fg=FG,
                                     font=('Consolas', 17), relief='flat',
                                     state='disabled', wrap='word', padx=8, pady=4,
                                     insertbackground=FG)
        detail_vsb = ttk.Scrollbar(detail_inner, orient='vertical',
                                    command=self._detail_text.yview)
        self._detail_text.configure(yscrollcommand=detail_vsb.set)
        self._detail_text.grid(row=0, column=0, sticky='nsew')
        detail_vsb.grid(row=0, column=1, sticky='ns')

    def _build_vizier_tab(self, outer_label):
        """Build the VizieR outer tab containing a sub-notebook for each VizieR catalog."""
        tab_frame = ttk.Frame(self._notebook)
        tab_frame.columnconfigure(0, weight=1)
        tab_frame.rowconfigure(0, weight=1)
        self._notebook.add(tab_frame, text=outer_label)

        sub_nb = ttk.Notebook(tab_frame)
        sub_nb.grid(row=0, column=0, sticky='nsew')
        self._vizier_notebook = sub_nb
        sub_nb.bind('<<NotebookTabChanged>>', self._on_vizier_subtab_changed)

        for key, label, cols in self.VIZIER_SUBTABS:
            self._build_tab(key, label, cols, notebook=sub_nb)

    def _build_tab(self, key, label, cols, notebook=None, note_text=None):
        nb = notebook if notebook is not None else self._notebook
        tab_frame = ttk.Frame(nb)
        tab_frame.columnconfigure(0, weight=1)
        nb.add(tab_frame, text=label)

        cur_row = 0
        if note_text:
            tk.Label(tab_frame, text=note_text, bg=BG, fg=ACC,
                     font=('Segoe UI', 11), wraplength=900, justify='left', anchor='w'
                     ).grid(row=cur_row, column=0, sticky='ew', padx=10, pady=(4, 2))
            cur_row += 1

        tab_frame.rowconfigure(cur_row + 1, weight=1)

        # Banner frame
        banner_outer = tk.Frame(tab_frame, bg=PANEL, height=48)
        banner_outer.grid(row=cur_row, column=0, sticky='ew')
        banner_outer.grid_propagate(False)
        banner_outer.columnconfigure(0, weight=1)
        self._banner_outer[key] = banner_outer
        cur_row += 1

        lbl = tk.Label(banner_outer, text="Ready", bg=PANEL, fg=BANNER_NS_FG,
                       font=('Segoe UI', 13, 'bold'), anchor='w', padx=12)
        lbl.grid(row=0, column=0, sticky='ew', pady=4)
        self._banner_label[key] = lbl

        pb = ttk.Progressbar(banner_outer, mode='indeterminate', length=200)
        # Don't grid it yet — shown only during query
        self._progress_bar[key] = pb

        # Treeview
        tree_frame = ttk.Frame(tab_frame)
        tree_frame.grid(row=cur_row, column=0, sticky='nsew')
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)

        col_ids = [c[0] for c in cols]
        tree = ttk.Treeview(tree_frame, columns=col_ids, show='headings',
                            selectmode='browse')
        tree.tag_configure('odd',  background='#25253a')
        tree.tag_configure('even', background=BG)

        for cid, cheader, cwidth in cols:
            tree.heading(cid, text=cheader,
                         command=lambda c=cid, k=key: self._sort(k, c))
            tree.column(cid, width=cwidth, minwidth=50, stretch=False)

        vsb = ttk.Scrollbar(tree_frame, orient='vertical',   command=tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal',  command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        tree.bind('<<TreeviewSelect>>', lambda e, k=key: self._on_double_click(k, e))

        self._trees[key] = tree
        self._tab_cols[key] = cols

    # ──────────────────────────────────────────────────────────
    # Banner methods
    # ──────────────────────────────────────────────────────────

    def _stop_spinner(self, key):
        if self._anim_job[key] is not None:
            try:
                self.after_cancel(self._anim_job[key])
            except Exception:
                pass
            self._anim_job[key] = None

    def _start_spinner(self, key, msg):
        self._anim_idx[key] = 0

        def tick():
            frame = SPINNER_FRAMES[self._anim_idx[key] % len(SPINNER_FRAMES)]
            self._anim_idx[key] += 1
            try:
                self._banner_label[key].config(text=f"{frame}  {msg}")
                self._anim_job[key] = self.after(100, tick)
            except Exception:
                pass

        tick()

    def _banner_working(self, key, msg="Querying..."):
        self._stop_spinner(key)
        outer = self._banner_outer[key]
        outer.config(bg=BANNER_WORK_BG)
        lbl = self._banner_label[key]
        lbl.config(bg=BANNER_WORK_BG, fg=BANNER_WORK_FG)

        pb = self._progress_bar[key]
        pb.grid(row=0, column=1, padx=(0, 12), pady=8, sticky='e')
        pb.start(15)

        self._start_spinner(key, msg)

    def _banner_done(self, key, n):
        self._stop_spinner(key)

        if n > 0:
            outer = self._banner_outer[key]
            outer.config(bg=BANNER_OK_BG)
            lbl = self._banner_label[key]
            lbl.config(bg=BANNER_OK_BG, fg=BANNER_OK_FG,
                       text=f"✓  {n} result{'s' if n != 1 else ''}")
        else:
            outer = self._banner_outer[key]
            outer.config(bg=BANNER_ERR_BG)
            lbl = self._banner_label[key]
            lbl.config(bg=BANNER_ERR_BG, fg=BANNER_ERR_FG, text="✗  No results")

    def _banner_ready(self, key):
        self._stop_spinner(key)
        pb = self._progress_bar[key]
        pb.stop()
        pb.grid_remove()
        outer = self._banner_outer[key]
        outer.config(bg=PANEL)
        lbl = self._banner_label[key]
        lbl.config(bg=PANEL, fg=BANNER_NS_FG, text="Ready")

    def _banner_not_searched(self, key):
        self._stop_spinner(key)
        pb = self._progress_bar[key]
        pb.stop()
        pb.grid_remove()
        outer = self._banner_outer[key]
        outer.config(bg=BANNER_NS_BG)
        lbl = self._banner_label[key]
        lbl.config(bg=BANNER_NS_BG, fg=BANNER_NS_FG, text="Not searched")

    def _banner_error(self, key, msg):
        self._stop_spinner(key)
        outer = self._banner_outer[key]
        outer.config(bg=BANNER_ERR_BG)
        lbl = self._banner_label[key]
        # Truncate long error messages
        short_msg = msg[:80] + '...' if len(msg) > 80 else msg
        lbl.config(bg=BANNER_ERR_BG, fg=BANNER_ERR_FG, text=f"✗  Error: {short_msg}")
        self._on_query_complete()

    # ──────────────────────────────────────────────────────────
    # Mode toggle
    # ──────────────────────────────────────────────────────────

    def _toggle_mode(self):
        mode = self._mode_var.get()
        if mode == 'name':
            self._name_entry.config(state='normal')
            self._name_hint.config(fg=BANNER_NS_FG)
            self._ra_entry.config(state='disabled')
            self._dec_entry.config(state='disabled')
            self._radius_entry.config(state='normal')   # always editable
        else:
            self._name_entry.config(state='disabled')
            self._name_hint.config(fg='#555577')
            self._ra_entry.config(state='normal')
            self._dec_entry.config(state='normal')
            self._radius_entry.config(state='normal')

    # ──────────────────────────────────────────────────────────
    # Query execution
    # ──────────────────────────────────────────────────────────

    def _get_filters(self):
        def _f(var):
            v = var.get().strip()
            return _try_float(v)

        return {
            'period_min': _f(self._period_min_var),
            'period_max': _f(self._period_max_var),
            'mag_min':    _f(self._mag_min_var),
            'mag_max':    _f(self._mag_max_var),
            'otype':      self._otype_var.get(),
        }

    def _run_query(self):
        if self._batch_index >= 0:
            if not messagebox.askyesno("Clear Batch",
                    "Starting a manual search will discard the current batch. Continue?"):
                return
            self._close_batch(confirm=False)

        mode = self._mode_var.get()
        flt  = self._get_filters()

        if mode == 'name':
            name = self._name_var.get().strip()
            if not name:
                messagebox.showwarning("Input Required", "Please enter a star name.")
                return
            history = [n for n in self._cfg.get('name_history', []) if n != name]
            history.insert(0, name)
            history = history[:5]
            self._cfg['name_history'] = history
            _save_config({'name_history': history})
            self._name_entry['values'] = history
            ra_deg = dec_deg = None
            try:
                radius = float(self._radius_var.get().strip())
            except (ValueError, TypeError):
                radius = 1.0
        else:
            ra_str     = self._ra_var.get().strip()
            dec_str    = self._dec_str = self._dec_var.get().strip()
            radius_str = self._radius_var.get().strip()
            if not ra_str or not dec_str:
                messagebox.showwarning("Input Required", "Please enter RA and Dec.")
                return
            ra_deg = parse_ra(ra_str)
            dec_deg = parse_dec(dec_str)
            if ra_deg is None or dec_deg is None:
                messagebox.showerror("Parse Error", "Could not parse RA or Dec. Please check format.")
                return
            try:
                radius = float(radius_str)
            except ValueError:
                radius = 30.0
            _save_config({'last_radius': radius_str})
            name = None

        self._fire_query(mode, name, ra_deg, dec_deg, radius, flt)

    def _fire_query(self, mode, name, ra_deg, dec_deg, radius, flt=None, gaia_extra=None):
        """Dispatch queries to all selected data sources."""
        if flt is None:
            flt = self._get_filters()
        if gaia_extra is None:
            gaia_extra = []
            for opt_key, _, col_list in GAIA_PARAM_OPTS:
                if self._gaia_param_vars[opt_key].get():
                    gaia_extra.extend(col_list)

        self._set_status("Running query...")
        self._resolved_var.set('')
        self._last_query_mode = mode
        self._wildcard_mode = (mode == 'name' and bool(name) and '*' in name)
        if self._wildcard_mode:
            self._resolved_var.set("Wildcard search — up to 100 results per source")

        # Count threads about to be fired (for batch completion tracking)
        pending = 0
        for key, label, cols in self.TABS:
            if not self._src_vars[key].get():
                continue
            pending += len(self.VIZIER_SUBTABS) if key == 'vizier' else 1
        self._query_pending = pending

        for key, label, cols in self.TABS:
            if not self._src_vars[key].get():
                if key == 'vizier':
                    for vk, vlabel, _ in self.VIZIER_SUBTABS:
                        self._clear_tab(vk)
                        self._banner_not_searched(vk)
                        self._update_tab_text(vk, vlabel, None)
                else:
                    self._clear_tab(key)
                    self._banner_not_searched(key)
                    self._update_tab_text(key, label, None)
                continue

            if key == 'vizier':
                for vk, vlabel, _ in self.VIZIER_SUBTABS:
                    self._clear_tab(vk, silent=True)
                    self._banner_working(vk)
                    self._update_tab_text(vk, vlabel, None)
                    threading.Thread(
                        target=self._query_worker,
                        args=(vk, vlabel, mode, name, ra_deg, dec_deg, radius, flt, gaia_extra),
                        daemon=True,
                    ).start()
            else:
                self._clear_tab(key, silent=True)
                self._banner_working(key)
                self._update_tab_text(key, label, None)
                threading.Thread(
                    target=self._query_worker,
                    args=(key, label, mode, name, ra_deg, dec_deg, radius, flt, gaia_extra),
                    daemon=True,
                ).start()

    def _query_worker(self, key, label, mode, name, ra_deg, dec_deg, radius, flt, gaia_extra=None):
        pm  = flt['period_min']
        pM  = flt['period_max']
        mm  = flt['mag_min']
        mM  = flt['mag_max']
        ot  = flt['otype']

        def status_cb(msg):
            self.after(0, self._set_status, msg)

        try:
            if key == 'simbad':
                if mode == 'name':
                    results = query_simbad_by_name(name, pm, pM, mm, mM, status_cb, otype_filter=ot)
                else:
                    results = query_simbad(ra_deg, dec_deg, radius, ot, pm, pM, mm, mM, status_cb)

            elif key == 'vsx':
                if mode == 'name':
                    results = query_vsx_by_name(name, pm, pM, mm, mM, status_cb)
                else:
                    results = query_vsx(ra_deg, dec_deg, radius, pm, pM, mm, mM, status_cb)

            elif key == 'tmass':
                if mode == 'name':
                    results = query_tmass_by_name(name, radius, status_cb)
                else:
                    results = query_tmass(ra_deg, dec_deg, radius, status_cb)

            elif key == 'wise':
                if mode == 'name':
                    results = query_wise_by_name(name, radius, status_cb)
                else:
                    results = query_wise(ra_deg, dec_deg, radius, status_cb)

            elif key == 'apass':
                if mode == 'name':
                    results = query_apass_by_name(name, radius, status_cb)
                else:
                    results = query_apass(ra_deg, dec_deg, radius, status_cb)

            elif key == 'tycho2':
                if mode == 'name':
                    results = query_tycho2_by_name(name, radius, status_cb)
                else:
                    results = query_tycho2(ra_deg, dec_deg, radius, status_cb)

            elif key == 'wds':
                if mode == 'name':
                    results = query_wds_by_name(name, radius, status_cb)
                else:
                    results = query_wds(ra_deg, dec_deg, radius, status_cb)

            elif key == 'orb6':
                if mode == 'name':
                    results = query_orb6_by_name(name, radius, status_cb)
                else:
                    results = query_orb6(ra_deg, dec_deg, radius, status_cb)

            elif key == 'gaia':
                if mode == 'name':
                    results = query_gaia_by_name(name, radius, status_cb, extra_cols=gaia_extra)
                else:
                    results = query_gaia(ra_deg, dec_deg, radius, status_cb, extra_cols=gaia_extra)

            elif key == 'nea':
                if mode == 'name':
                    results = query_nea_by_name(name, status_cb)
                else:
                    results = query_nea(ra_deg, dec_deg, radius, status_cb)
                self.after(0, self._populate_nea, results)
                self.after(0, self._update_tab_text, 'nea', 'NEA', len(results))
                return

            else:
                results = []

            self.after(0, self._populate, key, results)
            self.after(0, self._update_tab_text, key, label, len(results))

        except Exception as e:
            self.after(0, self._banner_error, key, str(e))
            self.after(0, self._update_tab_text, key, label, 0)
            self.after(0, self._set_status, f"Error querying {label}: {e}")

    # ──────────────────────────────────────────────────────────
    # Populate / sort / clear
    # ──────────────────────────────────────────────────────────

    def _populate(self, key, results):
        self._results[key] = results
        self._refresh_tree(key)
        self._banner_done(key, len(results))
        self._set_status(f"{key.upper()}: {len(results)} result(s) loaded.")
        if key == 'simbad' and getattr(self, '_last_query_mode', '') == 'name' and results:
            if not self._wildcard_mode:
                main_id = results[0].get('Name', '')
                if main_id:
                    self._resolved_var.set(f"SIMBAD → {main_id}")
        if key == 'simbad' and results and not getattr(self, '_wildcard_mode', False):
            for result_row in results[:10]:
                mid = result_row.get('Name', '')
                if mid and result_row.get('_extended_detail') is None:
                    result_row['_extended_detail'] = '_fetching'
                    threading.Thread(
                        target=self._fetch_simbad_detail,
                        args=(mid, '', result_row),
                        daemon=True,
                    ).start()
        self._auto_size_pane()
        if len(results) == 1:
            tree = self._trees[key]
            children = tree.get_children()
            if children:
                tree.selection_set(children[0])
                tree.focus(children[0])
                try:
                    active = self._active_tab_key()
                except Exception:
                    active = None
                if active == key:
                    self._show_row_detail(key, 0)
        self._on_query_complete()

    def _populate_nea(self, results):
        """Populate the NEA treeview."""
        self._results['nea'] = results
        self._refresh_tree('nea')
        n = len(results)
        self._banner_done('nea', n)
        self._set_status(f"NEA: {n} planet(s) found.")
        self._auto_size_pane()
        if n == 1 and self._active_tab_key() == 'nea':
            tree = self._trees['nea']
            children = tree.get_children()
            if children:
                tree.selection_set(children[0])
                tree.focus(children[0])
                self._show_row_detail('nea', 0)
        self._on_query_complete()

    def _refresh_tree(self, key):
        tree = self._trees[key]
        cols = self._tab_cols[key]
        col_ids = [c[0] for c in cols]

        tree.delete(*tree.get_children())

        results = self._results[key]
        for i, row in enumerate(results):
            values = [row.get(cid, '') for cid in col_ids]
            tag = 'odd' if i % 2 == 1 else 'even'
            tree.insert('', 'end', iid=str(i), values=values, tags=(tag,))

    def _clear_tab(self, key, silent=False):
        tree = self._trees[key]
        tree.delete(*tree.get_children())
        self._results[key] = []
        self._sort_col[key] = None
        self._sort_asc[key] = True
        # Reset headings (remove sort arrows)
        for cid, cheader, _ in self._tab_cols[key]:
            self._trees[key].heading(cid, text=cheader)
        # Clear detail
        self._set_detail('')
        if not silent:
            self._banner_ready(key)

    def _clear_all(self):
        for key, label, cols in self.TABS:
            if key == 'vizier':
                for vk, vlabel, _ in self.VIZIER_SUBTABS:
                    self._clear_tab(vk)
                    self._update_tab_text(vk, vlabel, None)
            else:
                self._clear_tab(key)
                self._update_tab_text(key, label, None)
        self._set_status("Results cleared.")

    def _sort(self, key, col):
        if self._sort_col[key] == col:
            self._sort_asc[key] = not self._sort_asc[key]
        else:
            self._sort_col[key] = col
            self._sort_asc[key] = True

        asc = self._sort_asc[key]
        is_numeric = col in NUMERIC_COLS

        def sort_key(row):
            val = row.get(col, '')
            if is_numeric:
                try:
                    return (0, float(val))
                except (ValueError, TypeError):
                    return (1, 0.0)  # empty/invalid sorts last
            else:
                return (0 if val else 1, str(val).lower())

        self._results[key].sort(key=sort_key, reverse=not asc)
        self._refresh_tree(key)

        # Update headings to show sort arrow
        for cid, cheader, _ in self._tab_cols[key]:
            if cid == col:
                arrow = '▲' if asc else '▼'
                self._trees[key].heading(cid, text=f"{cheader} {arrow}")
            else:
                self._trees[key].heading(cid, text=cheader)

    def _update_tab_text(self, key, base_label, n):
        vizier_keys = {k for k, _, _ in self.VIZIER_SUBTABS}

        def _tab_text(text, min_w):
            return text.ljust(min_w)

        if key in vizier_keys:
            sub_nb = self._vizier_notebook
            for i, (k, lbl, _) in enumerate(self.VIZIER_SUBTABS):
                if k == key:
                    raw = lbl if n is None else f"{lbl} ({n})"
                    sub_nb.tab(i, text=_tab_text(raw, 14))
                    break
            total = sum(len(self._results[k]) for k, _, _ in self.VIZIER_SUBTABS)
            nb = self._notebook
            for i, (k, lbl, _) in enumerate(self.TABS):
                if k == 'vizier':
                    nb.tab(i, text=_tab_text(f"VizieR ({total})", 16))
                    break

        else:
            nb = self._notebook
            for i, (k, lbl, _) in enumerate(self.TABS):
                if k == key:
                    raw = base_label if n is None else f"{base_label} ({n})"
                    nb.tab(i, text=_tab_text(raw, 16))
                    break

    def _active_tab_key(self):
        nb = self._notebook
        idx = nb.index(nb.select())
        key = self.TABS[idx][0]
        if key == 'vizier' and self._vizier_notebook is not None:
            try:
                sub_idx = self._vizier_notebook.index(self._vizier_notebook.select())
                return self.VIZIER_SUBTABS[sub_idx][0]
            except Exception:
                return self.VIZIER_SUBTABS[0][0]
        return key

    # ──────────────────────────────────────────────────────────
    # Detail panel
    # ──────────────────────────────────────────────────────────

    def _set_detail(self, text):
        self._detail_text.config(state='normal')
        self._detail_text.delete('1.0', 'end')
        if text:
            self._detail_text.insert('1.0', text)
        self._detail_text.config(state='disabled')

    def _show_row_detail(self, key, idx):
        if idx >= len(self._results[key]):
            return
        row = self._results[key][idx]
        basic_text = self._format_detail(key, row)
        if key == 'simbad':
            self._active_simbad_row = row
            cached = row.get('_extended_detail')
            if isinstance(cached, dict):
                self._render_simbad_detail(row)
            elif cached == '_fetching':
                self._set_detail(basic_text + "\n\n  Loading extended data…")
            else:
                self._set_detail(basic_text + "\n\n  Loading extended data…")
                main_id = row.get('Name', '')
                if main_id:
                    threading.Thread(
                        target=self._fetch_simbad_detail,
                        args=(main_id, basic_text, row),
                        daemon=True,
                    ).start()
        elif key == 'nea':
            self._active_simbad_row = None
            self._render_nea_detail(row)
        elif key == 'vsx':
            self._active_simbad_row = None
            self._active_vsx_row    = row
            self._render_vsx_detail(row)
        elif key in self._VIZIER_CAT_REFS:
            self._active_simbad_row = None
            self._render_vizier_detail(key, row)
        else:
            self._active_simbad_row = None
            self._set_detail(basic_text)

    def _on_double_click(self, key, event):
        """Called on <<TreeviewSelect>> (single click)."""
        try:
            if self._active_tab_key() != key:
                return   # selection change on a background tab — ignore
        except Exception:
            pass
        tree = self._trees[key]
        sel  = tree.selection()
        if not sel:
            return
        try:
            idx = int(sel[0])
        except ValueError:
            return
        self._show_row_detail(key, idx)

    def _on_vizier_subtab_changed(self, event=None):
        """Update detail panel when the user switches VizieR sub-tabs."""
        self.after(10, self._auto_size_pane)
        try:
            sub_idx = self._vizier_notebook.index(self._vizier_notebook.select())
            key = self.VIZIER_SUBTABS[sub_idx][0]
        except Exception:
            return
        tree = self._trees[key]
        sel  = tree.selection()
        if sel:
            try:
                self._show_row_detail(key, int(sel[0]))
            except (ValueError, IndexError):
                self._set_detail('')
        elif len(self._results[key]) == 1:
            children = tree.get_children()
            if children:
                tree.selection_set(children[0])
                self._show_row_detail(key, 0)
        else:
            self._set_detail('')

    def _on_tab_changed(self, event=None):
        """Update detail panel when the user switches tabs."""
        self.after(10, self._auto_size_pane)
        try:
            nb  = self._notebook
            idx = nb.index(nb.select())
            key = self.TABS[idx][0]
        except Exception:
            return
        if key == 'vizier':
            self._on_vizier_subtab_changed()
            return
        tree = self._trees[key]
        sel  = tree.selection()
        if sel:
            try:
                self._show_row_detail(key, int(sel[0]))
            except (ValueError, IndexError):
                self._set_detail('')
        elif len(self._results[key]) == 1:
            children = tree.get_children()
            if children:
                tree.selection_set(children[0])
                self._show_row_detail(key, 0)
        else:
            self._set_detail('')

    def _fetch_simbad_detail(self, main_id, basic_text, result_row=None):
        """Background: fetch extended SIMBAD data and store as raw data dict."""
        safe_id = main_id.replace("'", "''")
        params_base = {'REQUEST': 'doQuery', 'LANG': 'ADQL', 'FORMAT': 'json'}

        def _q(adql):
            try:
                r = requests.post(SIMBAD_TAP,
                                  data={**params_base, 'QUERY': adql}, timeout=30)
                r.raise_for_status()
                return r.json().get('data', [])
            except Exception:
                return []

        adql_ext = f"""
SELECT b.sp_type, b.plx_value, b.plx_err, b.pmra, b.pmdec,
       b.rvz_radvel, b.rvz_err, b.nbref
FROM basic b
WHERE b.main_id = '{safe_id}'
""".strip()

        adql_flux = f"""
SELECT m.filter, m.flux, m.flux_err, m.bibcode
FROM flux m
JOIN basic b ON m.oidref = b.oid
WHERE b.main_id = '{safe_id}'
""".strip()

        adql_sptype = f"""
SELECT m.sptype, m.bibcode
FROM mesSpT m
JOIN basic b ON m.oidref = b.oid
WHERE b.main_id = '{safe_id}'
""".strip()

        adql_plx = f"""
SELECT m.plx, m.plx_err, m.bibcode
FROM mesPLX m
JOIN basic b ON m.oidref = b.oid
WHERE b.main_id = '{safe_id}'
ORDER BY m.plx_err ASC
""".strip()

        adql_pm = f"""
SELECT m.pmra, m.pmde, m.pmra_err, m.pmde_err, m.bibcode
FROM mesPM m
JOIN basic b ON m.oidref = b.oid
WHERE b.main_id = '{safe_id}'
""".strip()

        adql_rv = f"""
SELECT m.velvalue, m.veltype, m.meanerror, m.bibcode
FROM mesVelocities m
JOIN basic b ON m.oidref = b.oid
WHERE b.main_id = '{safe_id}'
""".strip()

        adql_feh = f"""
SELECT m.teff, m.log_g, m.fe_h, m.bibcode
FROM mesFe_H m
JOIN basic b ON m.oidref = b.oid
WHERE b.main_id = '{safe_id}'
""".strip()

        adql_rot = f"""
SELECT m.vsini, m.vsini_err, m.bibcode
FROM mesRot m
JOIN basic b ON m.oidref = b.oid
WHERE b.main_id = '{safe_id}'
""".strip()

        adql_dist = f"""
SELECT m.dist, m.dist_err, m.unit, m.bibcode
FROM mesDist m
JOIN basic b ON m.oidref = b.oid
WHERE b.main_id = '{safe_id}'
""".strip()

        adql_ident = f"""
SELECT i.id FROM ident i
JOIN basic b ON i.oidref = b.oid
WHERE b.main_id = '{safe_id}'
""".strip()

        ext_data    = _q(adql_ext)
        flux_data   = _q(adql_flux)
        sptype_data = _q(adql_sptype)
        plx_data    = _q(adql_plx)
        pm_data     = _q(adql_pm)
        rv_data     = _q(adql_rv)
        feh_data    = _q(adql_feh)
        rot_data    = _q(adql_rot)
        dist_data   = _q(adql_dist)
        ident_data  = _q(adql_ident)

        # Full bibliography + titles via ref + has_ref TAP join
        # (r.oidbib links ref → has_ref.oidbibref; r.oidbibref does NOT exist)
        refs_set   = set()
        ref_titles = {}
        _adql_refs = (
            f"SELECT r.bibcode, r.title "
            f"FROM ref r "
            f"JOIN has_ref h ON r.oidbib = h.oidbibref "
            f"JOIN basic b ON h.oidref = b.oid "
            f"WHERE b.main_id = '{safe_id}'"
        )
        for _row in _q(_adql_refs):
            if _row[0]:
                _bc = str(_row[0]).strip()
                refs_set.add(_bc)
                if _row[1]:
                    ref_titles[_bc] = str(_row[1]).strip()

        if not refs_set:
            # Fallback: ASCII scrape (no titles)
            _BIBPAT = re.compile(r'[12]\d{3}[A-Za-z&][A-Za-z0-9&.]{13}[A-Z]')
            try:
                _simid_resp = requests.get(
                    "https://simbad.cds.unistra.fr/simbad/sim-id"
                    f"?output.format=ASCII&Ident={urllib.parse.quote(main_id)}",
                    timeout=20
                )
                _simid_resp.raise_for_status()
                refs_set = set(_BIBPAT.findall(_simid_resp.text))
            except Exception:
                pass
        refs_set.discard('')

        # nbref from basic
        nbref = None
        if ext_data and ext_data[0][7] is not None:
            try:
                nbref = int(ext_data[0][7])
            except (ValueError, TypeError):
                pass

        # Parse spectral type measurements
        sptype_rows = []
        for r in sptype_data:
            sp = str(r[0]).strip() if r[0] else None
            bc = str(r[1]).strip() if r[1] else None
            if sp:
                sptype_rows.append((sp, bc))

        # Parse parallax measurements (sorted best-first by plx_err)
        plx_rows = []
        for r in plx_data:
            try:
                plx  = float(r[0])
                err  = float(r[1]) if r[1] is not None else None
                bc   = str(r[2]).strip() if r[2] else None
                dist = 1000.0 / plx if plx > 0 else None
                plx_rows.append((plx, err, dist, bc))
            except (ValueError, TypeError):
                pass

        # Parse proper motion measurements
        pm_rows = []
        for r in pm_data:
            try:
                pmra = float(r[0]);  pmde = float(r[1])
                era  = float(r[2]) if r[2] is not None else None
                ede  = float(r[3]) if r[3] is not None else None
                bc   = str(r[4]).strip() if r[4] else None
                pm_rows.append((pmra, pmde, era, ede, bc))
            except (ValueError, TypeError):
                pass

        # Parse radial velocity measurements
        rv_rows = []
        for r in rv_data:
            try:
                vel   = float(r[0])
                vtype = str(r[1]).strip() if r[1] else ''
                err   = float(r[2]) if r[2] is not None else None
                bc    = str(r[3]).strip() if r[3] else None
                rv_rows.append((vel, vtype, err, bc))
            except (ValueError, TypeError):
                pass

        # Parse individual flux measurements (filter, flux, err, bibcode)
        flux_rows = []
        for r in flux_data:
            try:
                band = str(r[0]).strip() if r[0] else None
                val  = float(r[1]) if r[1] is not None else None
                err  = float(r[2]) if r[2] is not None else None
                bc   = str(r[3]).strip() if r[3] else None
                if band and val is not None:
                    flux_rows.append((band, val, err, bc))
            except (ValueError, TypeError):
                pass

        # Parse mesFe_H
        fe_h_rows = []
        for r in feh_data:
            teff    = float(r[0]) if r[0] is not None else None
            log_g   = float(r[1]) if r[1] is not None else None
            fe_h    = float(r[2]) if r[2] is not None else None
            bibcode = str(r[3]).strip() if r[3] else None
            if any(v is not None for v in (teff, log_g, fe_h)):
                fe_h_rows.append((teff, log_g, fe_h, bibcode))

        # Parse mesRot
        rot_rows = []
        for r in rot_data:
            try:
                vsini = float(r[0]) if r[0] is not None else None
                err   = float(r[1]) if r[1] is not None else None
                bc    = str(r[2]).strip() if r[2] else None
                if vsini is not None:
                    rot_rows.append((vsini, err, bc))
            except (ValueError, TypeError):
                pass

        # Parse mesDist
        dist_rows = []
        for r in dist_data:
            try:
                dist = float(r[0]) if r[0] is not None else None
                err  = float(r[1]) if r[1] is not None else None
                unit = str(r[2]).strip() if r[2] else 'pc'
                bc   = str(r[3]).strip() if r[3] else None
                if dist is not None:
                    dist_rows.append((dist, err, unit, bc))
            except (ValueError, TypeError):
                pass

        # Parse identifiers and refs
        identifiers = sorted({str(r[0]).strip() for r in ident_data if r[0]} - {''})
        refs = sorted(refs_set, reverse=True)

        data = {
            'nbref':       nbref,
            'sptype_rows': sptype_rows,
            'plx_rows':    plx_rows,
            'pm_rows':     pm_rows,
            'rv_rows':     rv_rows,
            'flux_rows':   flux_rows,
            'fe_h_rows':   fe_h_rows,
            'rot_rows':    rot_rows,
            'dist_rows':   dist_rows,
            'identifiers': identifiers,
            'refs':        refs,
            'ref_titles':  ref_titles,
        }

        if result_row is not None:
            result_row['_extended_detail'] = data
        self.after(0, self._render_simbad_detail, result_row)

    def _dash(self, val):
        return val if val else '—'

    def _format_detail(self, key, row):
        d = self._dash
        if key == 'simbad':
            return (
                f"Name: {d(row.get('Name'))}    RA: {_ra_with_deg(d(row.get('RA_hms')))}    "
                f"Dec: {_dec_with_deg(d(row.get('Dec_dms')))}    Type: {d(row.get('OType_label'))}    "
                f"Var Sub-type: {d(row.get('VarType'))}\n"
                f"Period: {d(row.get('Period'))} d    "
                f"Max Mag: {d(row.get('MaxMag'))}    Min Mag: {d(row.get('MinMag'))}    "
                f"Band: {d(row.get('MagBand'))}    "
                f"Dist: {d(row.get('Dist_arcsec'))}\"    N refs: {d(row.get('N_refs'))}"
            )
        elif key == 'vsx':
            W = 18
            def _fld(lbl, val, unit=''):
                v = d(val)
                return f"{(lbl + ':'):<{W}}{v}" + (f" {unit}" if unit and v != '—' else '')
            others = row.get('OtherNames', [])
            others_str = '   '.join(others) if others else '—'
            lines = [
                _fld('Name',          row.get('Name')),
                _fld('AUID',          row.get('AUID')),
                _fld('Constellation', row.get('Constellation')),
                _fld('RA',            _ra_with_deg(row.get('RA_hms'))),
                _fld('Dec',           _dec_with_deg(row.get('Dec_dms'))),
                _fld('Dist',          row.get('Dist_arcsec'), '"'),
                '',
                _fld('Var type',      row.get('VarType')),
                _fld('Spectral type', row.get('SpectralType')),
                _fld('Period',        row.get('Period'), 'd' if row.get('Period') else ''),
                _fld('Max mag',       row.get('MaxMag')),
                _fld('Min mag',       row.get('MinMag')),
                _fld('Band',          row.get('MagBand')),
                _fld('Epoch (HJD)',   row.get('Epoch')),
                _fld('Rise/Ecl dur',  row.get('RiseDuration')),
                _fld('Discoverer',    row.get('Discoverer')),
                '',
                _fld('Other names',   others_str),
            ]
            return '\n'.join(lines)
        elif key == 'tmass':
            W = 18
            def _fld(lbl, val, unit=''):
                v = d(val)
                return f"{(lbl + ':'):<{W}}{v}" + (f" {unit}" if unit and v != '—' else '')
            def _phot(band, mag_key, err_key):
                v = d(row.get(mag_key))
                e = row.get(err_key, '')
                return f"{(band + ':'):<{W}}{v}" + (f" ± {e}" if e else '')
            lines = [
                _fld('Name',  row.get('Name')),
                _fld('RA',    _ra_with_deg(row.get('RA_hms'))),
                _fld('Dec',   _dec_with_deg(row.get('Dec_dms'))),
                _fld('Dist',  row.get('Dist_arcsec'), '"'),
                '',
                _phot('J', 'Jmag', 'e_Jmag'),
                _phot('H', 'Hmag', 'e_Hmag'),
                _phot('K', 'Kmag', 'e_Kmag'),
            ]
            if any(row.get(k) for k in ('Qflg', 'Rflg', 'Cflg')):
                lines.append('')
                if row.get('Qflg'): lines.append(_fld('Quality',      row['Qflg']))
                if row.get('Rflg'): lines.append(_fld('Read flags',   row['Rflg']))
                if row.get('Cflg'): lines.append(_fld('Contamination',row['Cflg']))
            return '\n'.join(lines)
        elif key == 'wise':
            W = 18
            def _fld(lbl, val, unit=''):
                v = d(val)
                return f"{(lbl + ':'):<{W}}{v}" + (f" {unit}" if unit and v != '—' else '')
            def _phot(band, mag_key, err_key):
                v = d(row.get(mag_key))
                e = row.get(err_key, '')
                return f"{(band + ':'):<{W}}{v}" + (f" ± {e}" if e else '')
            lines = [
                _fld('Name', row.get('Name')),
                _fld('RA',   _ra_with_deg(row.get('RA_hms'))),
                _fld('Dec',  _dec_with_deg(row.get('Dec_dms'))),
                _fld('Dist', row.get('Dist_arcsec'), '"'),
                '',
                _phot('W1', 'W1mag', 'e_W1mag'),
                _phot('W2', 'W2mag', 'e_W2mag'),
                _phot('W3', 'W3mag', 'e_W3mag'),
                _phot('W4', 'W4mag', 'e_W4mag'),
            ]
            if any(row.get(k) for k in ('J2m', 'H2m', 'K2m')):
                lines += [
                    '',
                    _fld('2MASS J (X-ref)', row.get('J2m')),
                    _fld('2MASS H (X-ref)', row.get('H2m')),
                    _fld('2MASS K (X-ref)', row.get('K2m')),
                ]
            if any(row.get(k) for k in ('qph', 'ccf', 'ex', 'var')):
                lines.append('')
                if row.get('qph'): lines.append(_fld('Quality',  row['qph']))
                if row.get('ccf'): lines.append(_fld('CC flags', row['ccf']))
                if row.get('ex'):  lines.append(_fld('Ext flag', row['ex']))
                if row.get('var'): lines.append(_fld('Var flag', row['var']))
            if row.get('pmRA') or row.get('pmDE'):
                lines += [
                    '',
                    _fld('PM RA',  row.get('pmRA'),  'mas/yr'),
                    _fld('PM Dec', row.get('pmDE'),  'mas/yr'),
                ]
            return '\n'.join(lines)
        elif key == 'apass':
            W = 18
            def _fld(lbl, val, unit=''):
                v = d(val)
                return f"{(lbl + ':'):<{W}}{v}" + (f" {unit}" if unit and v != '—' else '')
            def _phot(band, mag_key, err_key):
                v = d(row.get(mag_key))
                e = row.get(err_key, '')
                return f"{(band + ':'):<{W}}{v}" + (f" ± {e}" if e else '')
            lines = [
                _fld('Name', row.get('Name')),
                _fld('RA',   _ra_with_deg(row.get('RA_hms'))),
                _fld('Dec',  _dec_with_deg(row.get('Dec_dms'))),
                _fld('Dist', row.get('Dist_arcsec'), '"'),
                '',
                _phot('V',   'Vmag',  'e_Vmag'),
                _phot('B',   'Bmag',  'e_Bmag'),
                _phot('B-V', 'BV',    'e_BV'),
            ]
            if row.get('nobs') or row.get('mobs'):
                lines += [
                    '',
                    _fld('Fields observed', row.get('nobs')),
                    _fld('Measurements',    row.get('mobs')),
                ]
            return '\n'.join(lines)
        elif key == 'tycho2':
            W = 18
            def _fld(lbl, val, unit=''):
                v = d(val)
                return f"{(lbl + ':'):<{W}}{v}" + (f" {unit}" if unit and v != '—' else '')
            def _phot(band, mag_key, err_key):
                v = d(row.get(mag_key))
                e = row.get(err_key, '')
                return f"{(band + ':'):<{W}}{v}" + (f" ± {e}" if e else '')
            def _pm(lbl, val_key, err_key):
                v = d(row.get(val_key))
                e = row.get(err_key, '')
                return f"{(lbl + ':'):<{W}}{v}" + (f" ± {e}" if e else '') + (' mas/yr' if v != '—' else '')
            lines = [
                _fld('Name', row.get('Name')),
                _fld('RA',   _ra_with_deg(row.get('RA_hms'))),
                _fld('Dec',  _dec_with_deg(row.get('Dec_dms'))),
                _fld('Dist', row.get('Dist_arcsec'), '"'),
                '',
                _phot('BT', 'BTmag', 'e_BTmag'),
                _phot('VT', 'VTmag', 'e_VTmag'),
                '',
                _pm('PM RA',  'pmRA', 'e_pmRA'),
                _pm('PM Dec', 'pmDE', 'e_pmDE'),
            ]
            if row.get('HIP') or row.get('prox'):
                lines.append('')
                if row.get('HIP'):  lines.append(_fld('HIP',       row['HIP']))
                if row.get('prox'): lines.append(_fld('Proximity', row['prox'], "(0.1')"))
            return '\n'.join(lines)
        elif key == 'wds':
            W = 18
            def _fld(lbl, val, unit=''):
                v = d(val)
                return f"{(lbl + ':'):<{W}}{v}" + (f" {unit}" if unit and v != '—' else '')
            lines = [
                _fld('WDS',          ('WDS J' + row['WDS']) if row.get('WDS') else row.get('Name', '')),
                _fld('Components',   row.get('Comp')),
                _fld('Discoverer',   row.get('Disc')),
                _fld('RA',           _ra_with_deg(row.get('RA_hms'))),
                _fld('Dec',          _dec_with_deg(row.get('Dec_dms'))),
                _fld('Dist',         row.get('Dist_arcsec'), '"'),
                '',
                _fld('First obs',    row.get('Obs1')),
                _fld('Last obs',     row.get('Obs2')),
                _fld('N obs',        row.get('Nobs')),
                '',
                _fld('PA (first)',   row.get('PA1')),
                _fld('PA (last)',    row.get('PA2')),
                '',
                _fld('Sep (first)',  row.get('Sep1'), '"'),
                _fld('Sep (last)',   row.get('Sep2'), '"'),
                '',
                _fld('Mag primary',  row.get('Mag1')),
                _fld('Mag secondary',row.get('Mag2')),
            ]
            if row.get('SpType'):
                lines.extend(['', _fld('Spectral type', row.get('SpType'))])
            return '\n'.join(lines)
        elif key == 'orb6':
            W = 20
            def _fld(lbl, val, unit=''):
                v = d(val)
                return f"{(lbl + ':'):<{W}}{v}" + (f" {unit}" if unit and v != '—' else '')
            def _meas(lbl, val, err, unit=''):
                v = d(val)
                s = f"{(lbl + ':'):<{W}}{v}"
                if err and str(err).strip() and v != '—':
                    s += f" ± {str(err).strip()}"
                if unit and v != '—':
                    s += f" {unit}"
                return s
            wds        = row.get('WDS', '')
            grade_code = row.get('Grade', '')
            grade_label = _ORB6_GRADES.get(grade_code, grade_code)
            p_unit     = row.get('P_unit', 'yr') or 'yr'
            lines = [
                _fld('WDS', ('WDS J' + wds) if wds else row.get('Name', '')),
                _fld('Discoverer', row.get('Disc')),
                _fld('RA',  _ra_with_deg(row.get('RA_hms'))),
                _fld('Dec', _dec_with_deg(row.get('Dec_dms'))),
            ]
            if row.get('ADS'): lines.append(_fld('ADS', row['ADS']))
            if row.get('HD'):  lines.append(_fld('HD',  row['HD']))
            if row.get('HIP'): lines.append(_fld('HIP', row['HIP']))
            lines += [
                '',
                _fld('Mag 1', row.get('Mag1')),
                _fld('Mag 2', row.get('Mag2')),
                '',
                '─── ORBITAL ELEMENTS ───────────────────────────────────────',
                _meas('Period',        row.get('P'),     row.get('e_P'),     p_unit),
                _meas('Epoch T₀',      row.get('T0'),    row.get('e_T0'),    'yr'),
                _meas('Semi-major axis', row.get('a'),   row.get('e_a'),     '"'),
                _meas('Inclination',   row.get('i'),     row.get('e_i'),     '°'),
                _meas('Ω (asc. node)', row.get('Omega'), row.get('e_Omega'), '°'),
                _meas('Eccentricity',  row.get('e_ecc'), row.get('e_e_ecc'), ''),
                _meas('ω (arg. peri.)', row.get('omega'), row.get('e_omega'), '°'),
            ]
            if row.get('Equinox'):
                lines.append(_fld('Equinox', row['Equinox']))
            lines += [
                '',
                _fld('Last obs', row.get('LastObs')),
                _fld('Grade', (f"{grade_code} — {grade_label}") if grade_label and grade_label != grade_code else grade_code),
                _fld('Reference', row.get('Ref')),
            ]
            return '\n'.join(lines)
        elif key == 'gaia':
            W = 20
            def _fld(lbl, val, unit=''):
                v = d(val)
                return f"{(lbl + ':'):<{W}}{v}" + (f" {unit}" if unit and v != '—' else '')
            def _meas(lbl, val_key, err_key, unit):
                v = d(row.get(val_key))
                e = row.get(err_key, '')
                s = f"{(lbl + ':'):<{W}}{v}"
                if e and v != '—':
                    s += f" ± {e}"
                if v != '—':
                    s += f" {unit}"
                return s
            lines = [
                _fld('Source ID', row.get('source_id')),
                _fld('RA',        _ra_with_deg(row.get('RA_hms'))),
                _fld('Dec',       _dec_with_deg(row.get('Dec_dms'))),
                _fld('Dist',      row.get('Dist_arcsec'), '"'),
                '',
                _meas('G mag',    'GMag',  '',         'mag'),
            ]
            if row.get('BPMag'):  lines.append(_meas('BP mag',  'BPMag', '',        'mag'))
            if row.get('RPMag'):  lines.append(_meas('RP mag',  'RPMag', '',        'mag'))
            if row.get('BPRP'):   lines.append(_fld('BP-RP',    row.get('BPRP')))
            if row.get('Parallax'):
                lines += ['', _meas('Parallax', 'Parallax', 'PlxErr', 'mas')]
            if row.get('PMRA') or row.get('PMDec'):
                lines += [
                    '' if not row.get('Parallax') else None,
                    _meas('PM RA',  'PMRA',  'PMRAErr',  'mas/yr'),
                    _meas('PM Dec', 'PMDec', 'PMDecErr', 'mas/yr'),
                ]
                lines = [l for l in lines if l is not None]
            if row.get('RV'):
                lines += ['', _meas('Radial velocity', 'RV', 'RVErr', 'km/s')]
            if row.get('RUWE'):
                lines.append(_fld('RUWE', row.get('RUWE')))
            return '\n'.join(lines)
        elif key == 'nea':
            return self._format_detail_nea(row)
        return ''

    def _format_detail_nea(self, row):
        d = self._dash
        def _v(k, unit='', fmt=None):
            v = row.get(k)
            if not v:
                return '\u2014'
            try:
                fv = float(v)
                return (format(fv, fmt) if fmt else str(v)) + (f' {unit}' if unit else '')
            except (ValueError, TypeError):
                return str(v) + (f' {unit}' if unit else '')

        deg_sym  = '\u00b0'
        gcm3_sym = 'g/cm\u00b3'
        rsun_sym = 'R\u2609'
        msun_sym = 'M\u2609'
        lsun_sym = 'L\u2609'
        hr_orb = '\u2500\u2500\u2500 ORBITAL PARAMETERS \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500'
        hr_pl  = '\u2500\u2500\u2500 PLANET PARAMETERS \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500'
        hr_st  = '\u2500\u2500\u2500 STELLAR PARAMETERS \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500'
        hr_sys = '\u2500\u2500\u2500 SYSTEM \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500'
        lines = [
            f"Planet: {d(row.get('pl_name'))}    Host: {d(row.get('hostname'))}",
            f"Discovery: {d(row.get('discoverymethod'))} | {d(row.get('disc_year'))} | {d(row.get('disc_facility'))}",
            "",
            hr_orb,
            f"Period:       {_v('pl_orbper', 'd')}    T0 (BJD):  {_v('pl_tranmid')}",
            f"Semi-major:   {_v('pl_orbsmax', 'AU')}    Incl:      {_v('pl_orbincl', deg_sym)}",
            f"Eccentricity: {_v('pl_orbeccen')}    Impact:    {_v('pl_imppar')}",
            "",
            hr_pl,
            f"Radius:  {_v('pl_rade', 'RE')}    Mass:    {_v('pl_bmasse', 'ME')}",
            f"Teq:     {_v('pl_eqt', 'K')}    Density: {_v('pl_dens', gcm3_sym)}",
            f"Trans Depth: {_v('pl_trandep')}    Duration: {_v('pl_trandur', 'h')}",
            "",
            hr_st,
            f"Teff:  {_v('st_teff', 'K')}    log g: {_v('st_logg')}    [Fe/H]: {_v('st_met')}",
            f"R*:    {_v('st_rad', rsun_sym)}    M*:    {_v('st_mass', msun_sym)}    L*: {_v('st_lum', lsun_sym)}",
            f"Age:   {_v('st_age', 'Gyr')}",
            "",
            hr_sys,
            f"Distance: {_v('sy_dist', 'pc')}    # Stars: {_v('sy_snum')}    # Planets: {_v('sy_pnum')}",
            f"V mag: {_v('sy_vmag')}    K mag: {_v('sy_kmag')}    Gaia mag: {_v('sy_gaiamag')}",
        ]
        return '\n'.join(lines)

    @staticmethod
    def _nea_ref_groups(row):
        """Return OrderedDict of ref_text -> {'params': [...], 'url': str|None}."""
        from collections import OrderedDict
        groups = OrderedDict()
        urls = row.get('_reflink_urls', {})
        for col, param_label in _NEA_REF_LABELS:
            ref_text = row.get(col)
            if ref_text:
                if ref_text not in groups:
                    groups[ref_text] = {'params': [], 'url': urls.get(col)}
                groups[ref_text]['params'].append(param_label)
        return groups

    # Catalog-level ADS references for tabs without per-row ref columns
    _VIZIER_CAT_REFS = {
        'tmass':  ('Cutri et al. 2003',                    '2003yCat.2246....0C'),
        'wise':   ('Cutri et al. 2013',                    '2013yCat.2328....0C'),
        'apass':  ('Henden et al. 2016',                   '2016yCat.2336....0H'),
        'tycho2': ('Høg et al. 2000',                      '2000A&A...355L..27H'),
        'gaia':   ('Gaia Collaboration (Vallenari+) 2023', '2023A&A...674A...1G'),
        'wds':    ('Mason et al. 2001',                    '2001AJ....122.3466M'),
        'orb6':   ('Hartkopf et al. 2001',                 '2001AJ....122.3472H'),
    }

    def _render_vsx_detail(self, result_row):
        """Render VSX detail: columnar text + remarks + references (fetched in background)."""
        if self._active_tab_key() != 'vsx':
            return
        row = result_row
        plain = self._format_detail('vsx', row)
        txt = self._detail_text
        txt.config(state='normal')
        txt.delete('1.0', 'end')
        txt.insert('end', plain)

        fetched = row.get('_vsx_refs') is not None  # True once background fetch completes

        def _link_seg(label, url, bibcode=''):
            tag = f"_vsxr_{abs(hash(label + (url or ''))  % 0xFFFFFF)}"
            txt.tag_configure(tag, foreground=ACC, underline=True)
            txt.tag_bind(tag, '<Button-1>', lambda e, u=url: webbrowser.open(u))
            txt.tag_bind(tag, '<Enter>',    lambda e: txt.config(cursor='hand2'))
            txt.tag_bind(tag, '<Leave>',    lambda e: txt.config(cursor='arrow'))
            txt.insert('end', label, tag)
            if bibcode:
                txt.insert('end', f'  ({bibcode})')
            txt.insert('end', '\n')

        if not fetched:
            txt.insert('end', '\n\n─── REMARKS / REFERENCES (VSX) ─────────────────────────────\n')
            txt.insert('end', '  Loading…')
        else:
            # ── Remarks ──────────────────────────────────────────────────
            remarks = row.get('_vsx_remarks', [])
            txt.insert('end', '\n\n─── REMARKS (VSX) ───────────────────────────────────────────\n')
            if not remarks:
                txt.insert('end', '  None.\n')
            else:
                for rem in remarks:
                    submitter = rem.get('submitter', '')
                    text      = rem.get('text', '')
                    if submitter:
                        txt.insert('end', f'{submitter}:\n')
                    txt.insert('end', f'  {text}\n\n')

            # ── References ───────────────────────────────────────────────
            refs = row.get('_vsx_refs', [])
            txt.insert('end', '─── REFERENCES (VSX) ───────────────────────────────────────\n')
            if not refs:
                txt.insert('end', '  None.\n')
            else:
                for ref in refs:
                    citation = ref['citation']
                    bibcode  = ref.get('bibcode', '')
                    url      = ref.get('url')
                    if url:
                        _link_seg(citation, url, bibcode)
                    else:
                        txt.insert('end', citation)
                        if bibcode:
                            txt.insert('end', f'  ({bibcode})')
                        txt.insert('end', '\n')

        # VSX star page link
        oid = row.get('_vsx_oid', '')
        if oid:
            star_url = f"https://www.aavso.org/vsx/index.php?view=detail.top&oid={oid}"
        else:
            star_url = f"https://www.aavso.org/vsx/index.php?view=results.get&ident={urllib.parse.quote(row.get('Name', ''))}"
        ptag = f"_vsxp_{abs(hash(star_url)) % 0xFFFFFF}"
        txt.tag_configure(ptag, foreground=ACC, underline=True)
        txt.tag_bind(ptag, '<Button-1>', lambda e, u=star_url: webbrowser.open(u))
        txt.tag_bind(ptag, '<Enter>',    lambda e: txt.config(cursor='hand2'))
        txt.tag_bind(ptag, '<Leave>',    lambda e: txt.config(cursor='arrow'))
        txt.insert('end', '\nAAVSO VSX star page', ptag)
        txt.config(state='disabled')

        # Start background fetch if not yet loaded
        if not fetched:
            oid = row.get('_vsx_oid', '')
            if oid:
                threading.Thread(
                    target=self._fetch_vsx_page_data,
                    args=(oid, row),
                    daemon=True,
                ).start()

    def _fetch_vsx_page_data(self, oid, result_row):
        """Background thread: fetch VSX star page and parse Remarks + References."""
        try:
            url = f"https://www.aavso.org/vsx/index.php?view=detail.top&oid={oid}"
            r = requests.get(url, timeout=30)
            r.raise_for_status()
            html = r.text
            result_row['_vsx_remarks'] = _parse_vsx_page_remarks(html)
            result_row['_vsx_refs']    = _parse_vsx_page_refs(html)
        except Exception:
            result_row['_vsx_remarks'] = []
            result_row['_vsx_refs']    = []
        if self._active_vsx_row is result_row:
            self.after(0, self._render_vsx_detail, result_row)

    def _render_vizier_detail(self, key, row):
        """Render a VizieR/Gaia-tab detail panel with columnar text + hyperlinked catalog ref."""
        plain = self._format_detail(key, row)
        txt = self._detail_text
        txt.config(state='normal')
        txt.delete('1.0', 'end')
        txt.insert('end', plain)
        ref_info = self._VIZIER_CAT_REFS.get(key)
        if ref_info:
            label, bib = ref_info
            txt.insert('end', '\n\n─── CATALOG REFERENCE ──────────────────────────────────\n')
            link_tag = self._make_link_tag(txt, bib)
            txt.insert('end', label, link_tag)
            txt.insert('end', f'  ({bib})')
        if key == 'orb6':
            pngfile = row.get('_pngfile', '').strip()
            if pngfile:
                plot_url = _ORB6_PLOT_BASE + pngfile
                txt.insert('end', '\n\n─── ORBITAL PLOT ───────────────────────────────────────\n')
                ptag = f"_orb6plot_{abs(hash(plot_url)) % 0xFFFFFF}"
                txt.tag_configure(ptag, foreground=ACC, underline=True)
                txt.tag_bind(ptag, '<Button-1>', lambda e, u=plot_url: webbrowser.open(u))
                txt.tag_bind(ptag, '<Enter>', lambda e: txt.config(cursor='hand2'))
                txt.tag_bind(ptag, '<Leave>', lambda e: txt.config(cursor='arrow'))
                txt.insert('end', 'View orbital plot (GSU/WDS)', ptag)
        txt.config(state='disabled')

    def _render_nea_detail(self, row):
        """Render NEA detail with hyperlinked references into the detail Text widget."""
        plain = self._format_detail_nea(row)
        txt = self._detail_text
        txt.config(state='normal')
        txt.delete('1.0', 'end')
        txt.insert('end', plain)

        ref_groups = self._nea_ref_groups(row)
        if ref_groups:
            hr_ref = '\n─── REFERENCES (PSCP) ────────────────────────────────────\n'
            txt.insert('end', hr_ref)
            for ref_text, info in ref_groups.items():
                url = info['url']
                params = info['params']
                if url:
                    tag_name = f"_nea_{abs(hash(url)) % 0xFFFFFF}"
                    txt.tag_configure(tag_name, foreground=ACC, underline=True)
                    txt.tag_bind(tag_name, '<Button-1>',
                                 lambda e, u=url: webbrowser.open(u))
                    txt.tag_bind(tag_name, '<Enter>',
                                 lambda e: txt.config(cursor='hand2'))
                    txt.tag_bind(tag_name, '<Leave>',
                                 lambda e: txt.config(cursor='arrow'))
                    txt.insert('end', ref_text, tag_name)
                else:
                    txt.insert('end', ref_text)
                txt.insert('end', '\n')
                # Wrap param list at ~70 chars
                joined, line, limit = [], '', 70
                for p in params:
                    seg = (' · ' if line else '') + p
                    if line and len(line) + len(seg) > limit:
                        joined.append(line)
                        line = p
                    else:
                        line += seg
                if line:
                    joined.append(line)
                txt.insert('end', '  ' + '\n  '.join(joined) + '\n\n')

        txt.config(state='disabled')

    # ──────────────────────────────────────────────────────────
    # Export
    # ──────────────────────────────────────────────────────────

    def _export_excel(self):
        if not HAS_OPENPYXL:
            messagebox.showerror("Missing Library", "openpyxl is not installed.\nRun: pip install openpyxl")
            return

        key = self._active_tab_key()
        results = self._results[key]
        if not results:
            messagebox.showinfo("No Results", "No results to export for the active tab.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel files', '*.xlsx'), ('All files', '*.*')],
            title="Export to Excel",
        )
        if not path:
            return

        cols = self._tab_cols[key]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = key.upper()
        ws.append([cheader for _, cheader, _ in cols])
        for row in results:
            ws.append([row.get(cid, '') for cid, _, _ in cols])
        wb.save(path)
        self._set_status(f"Exported {len(results)} rows to {os.path.basename(path)}")

    def _export_csv(self):
        import csv

        key = self._active_tab_key()
        results = self._results[key]
        if not results:
            messagebox.showinfo("No Results", "No results to export for the active tab.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension='.csv',
            filetypes=[('CSV files', '*.csv'), ('All files', '*.*')],
            title="Export to CSV",
        )
        if not path:
            return

        cols = self._tab_cols[key]
        _csv_note = ("Note: Catalog matches are positional — nearest source within the "
                     "search radius is returned. Verify results in crowded fields or for close binaries.")
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([cheader for _, cheader, _ in cols])
            for row in results:
                writer.writerow([row.get(cid, '') for cid, _, _ in cols])
            writer.writerow([])
            writer.writerow([_csv_note])

        self._set_status(f"Exported {len(results)} rows to {os.path.basename(path)}")

    # ── Help menu ─────────────────────────────────────────────
    def _build_menu(self):
        mb_kw   = dict(bg=PANEL, fg=FG, tearoff=False,
                       activebackground=ACC, activeforeground=BG)
        item_kw = dict(bg=BG, fg=FG, tearoff=False,
                       activebackground=ACC, activeforeground=BG)

        menubar = tk.Menu(self, **mb_kw)
        self.config(menu=menubar)

        hm = tk.Menu(menubar, **item_kw)
        menubar.add_cascade(label="Help", menu=hm)
        hm.add_command(label="User Guide",       command=self._show_user_guide)
        hm.add_command(label="Revision History", command=self._show_revision_history)
        hm.add_separator()
        hm.add_command(label="About SOQyT…",     command=self._show_about)

    def _show_about(self):
        win = tk.Toplevel(self)
        win.title("About Stellar Object Query Tool")
        win.configure(bg=BG)
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        tk.Label(win, text="Stellar Object Query Tool", bg=PANEL, fg=ACC,
                 font=("Segoe UI", 14, "bold"), pady=10).pack(fill="x")

        tk.Label(win, text=f"Version {VERSION}", bg=BG,
                 font=("Segoe UI", 11, "bold"), fg=FG).pack(pady=(16, 2))
        tk.Label(win, text="Multi-Database Stellar Object Query Tool", bg=BG,
                 font=("Segoe UI", 11), fg=FG).pack()

        tk.Frame(win, bg=ACC, height=1).pack(fill="x", padx=36, pady=14)

        tk.Label(win, text="Art Trail", bg=BG,
                 font=("Segoe UI", 11, "bold"), fg=FG).pack()
        tk.Label(win, text=COPYRIGHT, bg=BG,
                 font=("Segoe UI", 9), fg=BANNER_NS_FG).pack(pady=(2, 0))
        em_lbl = tk.Label(win, text="art.trail@icloud.com", bg=BG,
                          font=("Segoe UI", 9, "underline"), fg=ACC, cursor="hand2")
        em_lbl.pack(pady=(2, 0))
        em_lbl.bind("<Button-1>", lambda e: webbrowser.open("mailto:art.trail@icloud.com"))
        gh_lbl = tk.Label(win, text="github.com/ArtTrail/SOQyT", bg=BG,
                          font=("Segoe UI", 9, "underline"), fg=ACC, cursor="hand2")
        gh_lbl.pack(pady=(2, 0))
        gh_lbl.bind("<Button-1>", lambda e: webbrowser.open("https://github.com/ArtTrail/SOQyT"))

        tk.Frame(win, bg=ACC, height=1).pack(fill="x", padx=36, pady=14)

        tk.Button(win, text="Close", command=win.destroy,
                  bg=PANEL, fg=FG, relief="flat",
                  font=("Segoe UI", 11), cursor="hand2", padx=16, pady=4
                  ).pack(pady=(0, 16))

        win.update_idletasks()
        px, py = self.winfo_x(), self.winfo_y()
        pw, ph = self.winfo_width(), self.winfo_height()
        ww, wh = win.winfo_width(), win.winfo_height()
        win.geometry(f"+{px + (pw - ww) // 2}+{py + (ph - wh) // 2}")

    # ──────────────────────────────────────────────────────────────
    # Batch import / navigate / export
    # ──────────────────────────────────────────────────────────────

    def _on_query_complete(self):
        """Called by _populate, _populate_nea, and _banner_error when a tab finishes."""
        self._query_pending = max(0, self._query_pending - 1)
        if self._query_pending == 0:
            all_query_keys = ([k for k, _, _ in self.TABS if k != 'vizier'] +
                              [k for k, _, _ in self.VIZIER_SUBTABS])
            for k in all_query_keys:
                pb = self._progress_bar.get(k)
                if pb:
                    pb.stop()
                    pb.grid_remove()
            if self._batch_index >= 0:
                if self._auto_run:
                    self._auto_run_after_id = self.after(1000, self._auto_run_next)
                else:
                    self._update_batch_nav()

    def _import_batch(self):
        import csv as _csv
        path = filedialog.askopenfilename(
            title="Import Batch File",
            filetypes=[('All files', '*.*'), ('CSV files', '*.csv'),
                       ('Excel files', '*.xlsx *.xls')],
        )
        if not path:
            return
        try:
            ext = os.path.splitext(path)[1].lower()
            if ext == '.xlsx':
                if not HAS_OPENPYXL:
                    messagebox.showerror("Missing Library",
                        "openpyxl is required to read .xlsx files.\nRun: pip install openpyxl")
                    return
                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb.active
                raw = list(ws.iter_rows(values_only=True))
                if not raw:
                    messagebox.showerror("Empty File", "The selected file contains no data.")
                    return
                headers = [str(c) if c is not None else '' for c in raw[0]]
                data_rows = [
                    {headers[i]: (str(v) if v is not None else '') for i, v in enumerate(row)}
                    for row in raw[1:] if any(v is not None for v in row)
                ]
            elif ext == '.xls':
                try:
                    import xlrd
                except ImportError:
                    messagebox.showerror("Missing Library",
                        "xlrd is required to read .xls files.\n"
                        "Run: pip install xlrd\n"
                        "Or save the file as .xlsx and re-import.")
                    return
                def _xv(v):
                    if isinstance(v, float) and v == int(v):
                        return str(int(v))
                    return str(v) if v != '' else ''
                wb = xlrd.open_workbook(path)
                ws = wb.sheet_by_index(0)
                if ws.nrows == 0:
                    messagebox.showerror("Empty File", "The selected file contains no data.")
                    return
                headers = [_xv(ws.cell_value(0, c)) for c in range(ws.ncols)]
                data_rows = [
                    {headers[i]: _xv(ws.cell_value(r, i)) for i in range(ws.ncols)}
                    for r in range(1, ws.nrows)
                    if any(ws.cell_value(r, c) != '' for c in range(ws.ncols))
                ]
            else:
                with open(path, newline='', encoding='utf-8-sig') as f:
                    reader = _csv.DictReader(f)
                    headers = list(reader.fieldnames or [])
                    data_rows = [dict(r) for r in reader]
            if not headers or not data_rows:
                messagebox.showerror("Empty File", "No data found in the file.")
                return
            self._batch_filename = os.path.splitext(os.path.basename(path))[0]
            self._show_batch_col_dialog(headers, data_rows, os.path.basename(path))
        except Exception as e:
            messagebox.showerror("Import Error", f"Could not read file:\n{e}")

    def _show_batch_col_dialog(self, headers, rows, filename):
        win = tk.Toplevel(self)
        win.title("Configure Batch Import")
        win.configure(bg=BG)
        win.resizable(False, False)
        win.grab_set()

        tk.Label(win, text="Configure Batch Import", bg=PANEL, fg=ACC,
                 font=('Segoe UI', 11, 'bold'), pady=8).pack(fill='x')
        tk.Label(win, text=f"File: {filename}", bg=BG, fg=FG,
                 font=('Segoe UI', 10)).pack(padx=14, pady=(8, 0), anchor='w')
        tk.Label(win, text=f"{len(rows)} row(s) found", bg=BG, fg=FG,
                 font=('Segoe UI', 10)).pack(padx=14, pady=(0, 4), anchor='w')

        ttk.Separator(win, orient='horizontal').pack(fill='x', padx=14, pady=6)

        tk.Label(win, text="Search mode:", bg=BG, fg=FG,
                 font=('Segoe UI', 10, 'bold')).pack(padx=14, anchor='w')
        mode_var = tk.StringVar(value='name')
        rb_f = tk.Frame(win, bg=BG)
        rb_f.pack(padx=14, fill='x', pady=4)
        tk.Radiobutton(rb_f, text="By Star Name", variable=mode_var, value='name',
                       bg=BG, fg=FG, selectcolor=ENT, activebackground=BG,
                       font=('Segoe UI', 11)).pack(side='left', padx=(0, 20))
        tk.Radiobutton(rb_f, text="By Coordinates", variable=mode_var, value='coords',
                       bg=BG, fg=FG, selectcolor=ENT, activebackground=BG,
                       font=('Segoe UI', 11)).pack(side='left')

        # Name mode widgets
        name_frame = tk.Frame(win, bg=BG)
        tk.Label(name_frame, text="Star Name column:", bg=BG, fg=FG,
                 font=('Segoe UI', 10)).pack(anchor='w')
        name_col_var = tk.StringVar(value=headers[0] if headers else '')
        ttk.Combobox(name_frame, textvariable=name_col_var, values=headers,
                     state='readonly', width=30).pack(anchor='w', pady=2)
        tk.Label(name_frame, text="Search radius (arcmin):", bg=BG, fg=FG,
                 font=('Segoe UI', 10)).pack(anchor='w', pady=(4, 0))
        name_radius_var = tk.StringVar(value=self._radius_var.get())
        ttk.Entry(name_frame, textvariable=name_radius_var, width=10).pack(anchor='w', pady=2)

        # Coords mode widgets
        coords_frame = tk.Frame(win, bg=BG)
        tk.Label(coords_frame, text="RA column:", bg=BG, fg=FG,
                 font=('Segoe UI', 10)).pack(anchor='w')
        ra_col_var = tk.StringVar(value=headers[0] if headers else '')
        ttk.Combobox(coords_frame, textvariable=ra_col_var, values=headers,
                     state='readonly', width=30).pack(anchor='w', pady=2)
        tk.Label(coords_frame, text="Dec column:", bg=BG, fg=FG,
                 font=('Segoe UI', 10)).pack(anchor='w', pady=(4, 0))
        dec_col_var = tk.StringVar(value=headers[1] if len(headers) > 1 else (headers[0] if headers else ''))
        ttk.Combobox(coords_frame, textvariable=dec_col_var, values=headers,
                     state='readonly', width=30).pack(anchor='w', pady=2)
        tk.Label(coords_frame, text="Search radius (arcmin):", bg=BG, fg=FG,
                 font=('Segoe UI', 10)).pack(anchor='w', pady=(4, 0))
        radius_var = tk.StringVar(value=self._radius_var.get())
        ttk.Entry(coords_frame, textvariable=radius_var, width=10).pack(anchor='w', pady=2)
        tk.Label(coords_frame, text="Target name column (optional):", bg=BG, fg=FG,
                 font=('Segoe UI', 10)).pack(anchor='w', pady=(4, 0))
        label_col_var = tk.StringVar(value='')
        ttk.Combobox(coords_frame, textvariable=label_col_var,
                     values=['(none)'] + headers,
                     state='readonly', width=30).pack(anchor='w', pady=2)
        label_col_var.set('(none)')

        def _toggle(*_):
            if mode_var.get() == 'name':
                coords_frame.pack_forget()
                name_frame.pack(padx=14, fill='x', pady=4)
            else:
                name_frame.pack_forget()
                coords_frame.pack(padx=14, fill='x', pady=4)
        mode_var.trace_add('write', _toggle)
        name_frame.pack(padx=14, fill='x', pady=4)  # start in name mode

        auto_var = tk.BooleanVar(value=False)
        auto_chk_f = tk.Frame(win, bg=BG)
        auto_chk_f.pack(padx=14, fill='x', pady=(8, 0))
        tk.Checkbutton(auto_chk_f, text="Auto-Run all targets",
                       variable=auto_var, bg=BG, fg=FG, selectcolor=ENT,
                       activebackground=BG,
                       font=('Segoe UI', 11)).pack(side='left')
        tk.Label(auto_chk_f,
                 text="  (queries every row, then prompts to export)",
                 bg=BG, fg='#888', font=('Segoe UI', 9)).pack(side='left')

        ttk.Separator(win, orient='horizontal').pack(fill='x', padx=14, pady=8)

        btn_f = tk.Frame(win, bg=BG)
        btn_f.pack(padx=14, pady=(0, 12), fill='x')

        def _load():
            m = mode_var.get()
            if m == 'name':
                nc = name_col_var.get()
                if not nc:
                    messagebox.showwarning("Column Required",
                        "Please select a Star Name column.", parent=win)
                    return
                try:
                    name_rad = float(name_radius_var.get())
                except ValueError:
                    name_rad = 1.0
                win.destroy()
                self._start_batch(rows, 'name', name_col=nc, radius=name_rad, auto_run=auto_var.get())
            else:
                rc, dc = ra_col_var.get(), dec_col_var.get()
                if not rc or not dc:
                    messagebox.showwarning("Column Required",
                        "Please select RA and Dec columns.", parent=win)
                    return
                try:
                    rad = float(radius_var.get())
                except ValueError:
                    rad = 1.0
                lc = label_col_var.get()
                lc = lc if lc and lc != '(none)' else None
                win.destroy()
                self._start_batch(rows, 'coords', ra_col=rc, dec_col=dc, radius=rad,
                                  label_col=lc, auto_run=auto_var.get())

        ttk.Button(btn_f, text="Load Batch", style='Accent.TButton',
                   command=_load).pack(side='left', padx=(0, 8))
        ttk.Button(btn_f, text="Cancel", command=win.destroy).pack(side='left')

        # Center over the main window
        win.update_idletasks()
        px, py = self.winfo_x(), self.winfo_y()
        pw, ph = self.winfo_width(), self.winfo_height()
        ww, wh = win.winfo_width(), win.winfo_height()
        win.geometry(f"+{px + (pw - ww) // 2}+{py + (ph - wh) // 2}")

    def _start_batch(self, rows, mode, name_col=None, ra_col=None, dec_col=None, radius=1.0, label_col=None, auto_run=False):
        if self._batch_index >= 0:
            self._close_batch(confirm=False)
        self._batch_rows      = rows
        self._batch_index     = 0
        self._batch_results   = {}
        self._batch_mode      = mode
        self._batch_name_col  = name_col
        self._batch_ra_col    = ra_col
        self._batch_dec_col   = dec_col
        self._batch_label_col = label_col
        self._batch_radius    = radius
        self._auto_run          = auto_run
        self._auto_run_after_id = None
        self._batch_nav_frame.pack(fill='x', padx=0, pady=(0, 4))
        if auto_run:
            self._batch_cancel_auto_btn.pack(padx=10, pady=(0, 2),
                                             before=self._batch_close_btn)
        self._update_batch_nav()
        self._batch_query_current()

    def _batch_label(self, index):
        if index < 0 or index >= len(self._batch_rows):
            return ''
        row = self._batch_rows[index]
        if self._batch_mode == 'name':
            return str(row.get(self._batch_name_col, f'Row {index+1}')).strip() or f'Row {index+1}'
        if self._batch_label_col:
            val = str(row.get(self._batch_label_col, '')).strip()
            if val:
                return val
        ra  = str(row.get(self._batch_ra_col,  '')).strip()
        dec = str(row.get(self._batch_dec_col, '')).strip()
        return f"{ra}, {dec}" if ra and dec else f'Row {index+1}'

    def _batch_save_current(self):
        if self._batch_index < 0:
            return
        label = self._batch_label(self._batch_index)
        all_keys = ([k for k, _, _ in self.TABS if k != 'vizier'] +
                    [k for k, _, _ in self.VIZIER_SUBTABS])
        self._batch_results[label] = {k: list(self._results.get(k, [])) for k in all_keys}

    def _batch_load_stored(self, label):
        stored = self._batch_results.get(label, {})
        all_keys = ([k for k, _, _ in self.TABS if k != 'vizier'] +
                    [k for k, _, _ in self.VIZIER_SUBTABS])
        key_to_label = {k: lbl for k, lbl, _ in self.TABS if k != 'vizier'}
        key_to_label.update({k: lbl for k, lbl, _ in self.VIZIER_SUBTABS})

        # First pass: restore all results and treeviews
        for key in all_keys:
            results = stored.get(key, [])
            self._results[key] = results
            self._refresh_tree(key)
            if results:
                self._banner_done(key, len(results))
            else:
                self._banner_ready(key)

        # Second pass: update tab labels (all self._results now current)
        for key in all_keys:
            self._update_tab_text(key, key_to_label.get(key, key), len(self._results[key]))

        self._set_detail('')
        self._resolved_var.set('')

        # Auto-select single-result rows and show detail for the active tab
        try:
            active_key = self._active_tab_key()
        except Exception:
            active_key = None
        for key in all_keys:
            if len(self._results[key]) == 1:
                tree = self._trees[key]
                children = tree.get_children()
                if children:
                    tree.selection_set(children[0])
                    tree.focus(children[0])
                    if key == active_key:
                        self._show_row_detail(key, 0)

        self._auto_size_pane()
        self._set_status(f"Batch: {label}")

    def _batch_query_current(self):
        if self._batch_index < 0 or self._batch_index >= len(self._batch_rows):
            return
        label = self._batch_label(self._batch_index)
        row = self._batch_rows[self._batch_index]
        if self._batch_mode == 'name':
            name = str(row.get(self._batch_name_col, '')).strip()
            if not name:
                self._set_status(f"Batch: skipping empty name at row {self._batch_index + 1}")
                self._query_pending = 0
                self._update_batch_nav()
                if self._auto_run:
                    self._auto_run_after_id = self.after(200, self._auto_run_next)
                return
            self._fire_query('name', name, None, None, self._batch_radius)
        else:
            ra_str  = str(row.get(self._batch_ra_col,  '')).strip()
            dec_str = str(row.get(self._batch_dec_col, '')).strip()
            ra_deg  = parse_ra(ra_str)
            dec_deg = parse_dec(dec_str)
            if ra_deg is None or dec_deg is None:
                self._set_status(
                    f"Batch: could not parse coordinates for row {self._batch_index + 1}")
                self._query_pending = 0
                self._update_batch_nav()
                if self._auto_run:
                    self._auto_run_after_id = self.after(200, self._auto_run_next)
                return
            self._fire_query('coords', None, ra_deg, dec_deg, self._batch_radius)

    def _batch_go(self, delta):
        if self._query_pending > 0:
            return
        self._batch_save_current()
        new_idx = self._batch_index + delta
        if new_idx < 0 or new_idx >= len(self._batch_rows):
            return
        self._batch_index = new_idx
        self._update_batch_nav()
        label = self._batch_label(new_idx)
        if label in self._batch_results:
            self._batch_load_stored(label)
            self._query_pending = 0
            self._update_batch_nav()
        else:
            self._batch_query_current()

    def _update_batch_nav(self):
        if self._batch_index < 0:
            return
        n     = len(self._batch_rows)
        label = self._batch_label(self._batch_index)
        if self._auto_run:
            self._batch_nav_label.config(
                text=f"Auto-running:  {self._batch_index + 1} / {n}  —  {label}")
        else:
            self._batch_nav_label.config(text=f"{self._batch_index + 1} / {n}:  {label}")
        busy     = self._query_pending > 0 or self._auto_run
        prev_ok  = not busy and self._batch_index > 0
        next_ok  = not busy and self._batch_index < n - 1
        self._batch_prev_btn.config(state='normal' if prev_ok else 'disabled',
                                     fg=FG if prev_ok else '#555')
        self._batch_next_btn.config(state='normal' if next_ok else 'disabled',
                                     fg=FG if next_ok else '#555')
        self._batch_export_btn.config(
            state='normal' if self._batch_results else 'disabled')

    def _auto_run_next(self):
        self._auto_run_after_id = None
        if not self._auto_run or self._batch_index < 0:
            return
        self._batch_save_current()
        next_idx = self._batch_index + 1
        n = len(self._batch_rows)
        if next_idx >= n:
            self._auto_run = False
            self._batch_cancel_auto_btn.pack_forget()
            self._update_batch_nav()
            self._set_status(f"Auto-run complete — {n} target(s) queried")
            if self._batch_results:
                if messagebox.askyesno("Auto-Run Complete",
                        f"All {n} target(s) queried.\n\nExport results now?",
                        parent=self):
                    self._export_batch()
            return
        self._batch_index = next_idx
        # Server cooldown: pause every _COOLDOWN_EVERY targets
        if next_idx > 0 and next_idx % _COOLDOWN_EVERY == 0:
            self._update_batch_nav()
            self._auto_run_cooldown(_COOLDOWN_SECS)
            return
        label = self._batch_label(next_idx)
        if label in self._batch_results:
            self._update_batch_nav()
            self._auto_run_after_id = self.after(100, self._auto_run_next)
        else:
            self._update_batch_nav()
            self._batch_query_current()

    def _auto_run_cooldown(self, secs_remaining):
        if not self._auto_run:
            return
        if secs_remaining <= 0:
            self._set_status("Server cooldown complete — resuming auto-run…")
            self._auto_run_after_id = self.after(500, self._auto_run_next)
            return
        self._set_status(
            f"Server cooldown — resuming in {secs_remaining}s  "
            f"({self._batch_index} / {len(self._batch_rows)} targets completed)")
        self._auto_run_after_id = self.after(
            1000, self._auto_run_cooldown, secs_remaining - 1)

    def _cancel_auto_run(self):
        self._auto_run = False
        if self._auto_run_after_id:
            self.after_cancel(self._auto_run_after_id)
            self._auto_run_after_id = None
        self._batch_cancel_auto_btn.pack_forget()
        self._update_batch_nav()
        self._set_status("Auto-run cancelled")

    def _close_batch(self, confirm=True):
        if self._auto_run:
            self._cancel_auto_run()
        if confirm and self._batch_results:
            if not messagebox.askyesno("Close Batch",
                    "Discard accumulated batch results and close?"):
                return
        self._batch_rows     = []
        self._batch_index    = -1
        self._batch_results  = {}
        self._batch_mode     = None
        self._batch_name_col  = None
        self._batch_ra_col    = None
        self._batch_dec_col   = None
        self._batch_label_col = None
        self._batch_nav_frame.pack_forget()
        self._set_status("Batch closed.")

    def _export_batch(self):
        if not HAS_OPENPYXL:
            messagebox.showerror("Missing Library",
                "openpyxl is required to export Excel files.\nRun: pip install openpyxl")
            return
        if not self._batch_results:
            messagebox.showinfo("No Results", "No batch results to export yet.")
            return
        import datetime
        _ts        = datetime.datetime.now().strftime('%m_%d_%Y_%H_%M_%S')
        _mode_str  = 'Name' if self._batch_mode == 'name' else 'Coord'
        _stem      = self._batch_filename or 'batch'
        _default   = f"{_stem}_SOQyT_Query_by_{_mode_str}_{_ts}"
        path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel files', '*.xlsx'), ('All files', '*.*')],
            title="Export Batch Results",
            initialfile=_default,
        )
        if not path:
            return

        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # (header_hex, alt_row_hex) per export key
        _PALETTE = {
            'simbad':       ('4472C4', 'DCE6F1'),
            'simbad_meas':  ('2E75B6', 'DEEAF1'),
            'simbad_phot':  ('1F4E79', 'D6E4F0'),
            'simbad_stpar': ('0070C0', 'CCE5F4'),
            'simbad_ident': ('005B8E', 'CCE0F0'),
            'simbad_refs':  ('002060', 'D0DEF0'),
            'vsx':          ('ED7D31', 'FCE4D6'),
            'tmass':        ('C55A11', 'F8CBAD'),
            'wise':         ('375623', 'E2EFDA'),
            'apass':        ('70AD47', 'EBF5E0'),
            'tycho2':       ('7030A0', 'E2D0F5'),
            'wds':          ('005B8E', 'CCE5FF'),
            'orb6':         ('BF8F00', 'FFF2CC'),
            'gaia':         ('C00000', 'FFDDD9'),
            'nea':          ('375623', 'E2EFDA'),
        }

        def _coerce_val(val):
            """Convert numeric strings to native numbers; preserve leading-zero codes."""
            if val is None or val == '':
                return ''
            if isinstance(val, (int, float)):
                return val
            s = str(val)
            # Preserve codes like '000', '042' (leading zero, no decimal)
            if len(s) > 1 and s[0] == '0' and s[1] != '.':
                return s
            try:
                f = float(s)
                if '.' not in s and 'e' not in s.lower():
                    try:
                        return int(f)
                    except (OverflowError, ValueError):
                        pass
                return f
            except (ValueError, TypeError):
                return s

        def _get_val(row, k):
            # Pull actual reference count from extended detail when available
            if k == 'N_refs':
                ext = row.get('_extended_detail')
                if isinstance(ext, dict) and ext.get('nbref') is not None:
                    return ext['nbref']
            # Gaia source IDs are 18-19 digit integers that exceed Excel's float precision;
            # force text so the full ID is preserved exactly.
            if k == 'source_id':
                return str(row.get(k, ''))
            return _coerce_val(row.get(k, ''))

        def _style_ws(ws, palette_key):
            hdr_hex, alt_hex = _PALETTE.get(palette_key, ('4472C4', 'DCE6F1'))
            hdr_fill = PatternFill('solid', fgColor=hdr_hex)
            alt_fill = PatternFill('solid', fgColor=alt_hex)
            hdr_font = Font(bold=True, color='FFFFFF', name='Segoe UI', size=10)
            dat_font = Font(name='Segoe UI', size=10)
            ctr      = Alignment(horizontal='center', vertical='center')
            thin_sd  = Side(style='thin', color='D0D0D0')
            hdr_brd  = Border(bottom=thin_sd)
            for cell in ws[1]:
                cell.fill      = hdr_fill
                cell.font      = hdr_font
                cell.alignment = ctr
                cell.border    = hdr_brd
            for row_idx, ws_row in enumerate(ws.iter_rows(min_row=2), start=2):
                fill = alt_fill if row_idx % 2 == 0 else None
                for cell in ws_row:
                    cell.font      = dat_font
                    cell.alignment = ctr
                    if fill:
                        cell.fill = fill
            ws.freeze_panes = 'B2'
            for col_cells in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col_cells), default=8)
                ws.column_dimensions[
                    get_column_letter(col_cells[0].column)
                ].width = min(max(max_len + 3, 10), 60)

        _NOTE_TEXT = (
            "Note: Catalog matches are positional — nearest source within the search "
            "radius is returned. Verify results in crowded fields or for close binaries."
        )

        def _add_note_row(ws):
            ncols = ws.max_column or 1
            ws.append([''] * ncols)
            note_row = ws.max_row
            ws.cell(note_row, 1).value = _NOTE_TEXT
            if ncols > 1:
                ws.merge_cells(start_row=note_row, start_column=1,
                               end_row=note_row, end_column=ncols)
            nc = ws.cell(note_row, 1)
            nc.fill      = PatternFill('solid', fgColor='FFF2CC')
            nc.font      = Font(name='Segoe UI', size=9, italic=True, color='7F6000')
            nc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.row_dimensions[note_row].height = 28

        def _apply_url_col(ws, col_idx):
            url_font = Font(name='Segoe UI', size=10, color='0563C1', underline='single')
            for ws_row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                cell = ws_row[0]
                if cell.value:
                    cell.hyperlink = str(cell.value)
                    cell.font = url_font

        def _apply_bibcode_cols(ws):
            """Hyperlink every 'Bibcode' column to NASA ADS (abs/{bibcode})."""
            bib_font = Font(name='Segoe UI', size=10, color='0563C1', underline='single')
            bib_col_indices = [c.column for c in ws[1] if c.value == 'Bibcode']
            for col_idx in bib_col_indices:
                for ws_row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    cell = ws_row[0]
                    if cell.value:
                        bib = str(cell.value).strip()
                        cell.hyperlink = (
                            f"https://ui.adsabs.harvard.edu/abs/"
                            f"{urllib.parse.quote(bib)}"
                        )
                        cell.font = bib_font

        tab_labels = {k: lbl for k, lbl, _ in self.TABS if k != 'vizier'}
        tab_labels.update({k: lbl for k, lbl, _ in self.VIZIER_SUBTABS})

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        written = 0

        # Positional photometric catalogs: export only the nearest match per target.
        # (Results are already sorted nearest-first by _sort_by_dist.)
        _NEAREST_ONLY = {'tmass', 'wise', 'apass', 'tycho2'}

        for key, col_defs in self._BATCH_EXPORT_COLS.items():
            all_rows = []
            for target_label, tab_data in self._batch_results.items():
                rows_for_target = tab_data.get(key, [])
                if key in _NEAREST_ONLY and rows_for_target:
                    rows_for_target = rows_for_target[:1]
                for row in rows_for_target:
                    all_rows.append((target_label, row))
            if not all_rows:
                continue
            sheet_name = tab_labels.get(key, key.upper())[:31]
            ws = wb.create_sheet(title=sheet_name)
            ws.append(['Target'] + [h for _, h in col_defs])
            for target_label, row in all_rows:
                ws.append([target_label] + [_get_val(row, k) for k, _ in col_defs])
            _style_ws(ws, key)
            if key == 'orb6':
                url_col_idx = next(
                    (i + 2 for i, (k, _) in enumerate(col_defs) if k == 'PlotURL'),
                    None,
                )
                if url_col_idx:
                    _apply_url_col(ws, url_col_idx)
            _add_note_row(ws)
            written += 1

        # ── SIMBAD extended sheets ────────────────────────────────────────────
        # Always written after the main loop so they appear even when the main
        # SIMBAD query returned no rows (e.g. coord-mode on non-variable stars).
        ext_rows = []
        for tgt_lbl, tab_data in self._batch_results.items():
            for sim_row in tab_data.get('simbad', []):
                ext = sim_row.get('_extended_detail')
                if isinstance(ext, dict):
                    ext_rows.append((tgt_lbl, sim_row.get('Name', ''), ext))

        def _write_meas_sheet(title, headers, rows):
            nonlocal written
            ws_x = wb.create_sheet(title=title)
            ws_x.append(headers)
            if rows:
                for r in rows:
                    ws_x.append(r)
            else:
                ncols = len(headers)
                ws_x.append(['No data available for these targets in this SIMBAD measurement table.']
                             + [''] * (ncols - 1))
                nd_row = ws_x.max_row
                if ncols > 1:
                    ws_x.merge_cells(start_row=nd_row, start_column=1,
                                     end_row=nd_row, end_column=ncols)
                nd_cell = ws_x.cell(nd_row, 1)
                nd_cell.font      = Font(name='Segoe UI', size=10, italic=True, color='888888')
                nd_cell.alignment = Alignment(horizontal='center', vertical='center')
                ws_x.row_dimensions[nd_row].height = 20
            _style_ws(ws_x, 'simbad_meas')
            _apply_bibcode_cols(ws_x)
            _add_note_row(ws_x)
            written += 1

        # ── Spectral Type ────────────────────────────────────────────────────
        spt_out = []
        for tgt, mid, ext in ext_rows:
            for sp, bc in ext.get('sptype_rows', []):
                spt_out.append([tgt, mid, sp, bc or ''])
        _write_meas_sheet('SIMBAD SpT',
                          ['Target', 'Main ID', 'SpType', 'Bibcode'],
                          spt_out)

        # ── Parallax ─────────────────────────────────────────────────────────
        plx_out = []
        for tgt, mid, ext in ext_rows:
            for plx, err, dist_pc, bc in ext.get('plx_rows', []):
                plx_out.append([tgt, mid, plx,
                                err if err is not None else '',
                                dist_pc if dist_pc is not None else '',
                                bc or ''])
        _write_meas_sheet('SIMBAD Plx',
                          ['Target', 'Main ID', 'Plx (mas)', 'e_Plx (mas)',
                           'Dist (pc)', 'Bibcode'],
                          plx_out)

        # ── Distance ─────────────────────────────────────────────────────────
        dist_out = []
        for tgt, mid, ext in ext_rows:
            for d, derr, unit, bc in ext.get('dist_rows', []):
                dist_out.append([tgt, mid, d,
                                 derr if derr is not None else '',
                                 unit or '', bc or ''])
        _write_meas_sheet('SIMBAD Dist',
                          ['Target', 'Main ID', 'Dist', 'e_Dist', 'Unit', 'Bibcode'],
                          dist_out)

        # ── Proper Motion ─────────────────────────────────────────────────────
        pm_out = []
        for tgt, mid, ext in ext_rows:
            for pmra, pmde, era, ede, bc in ext.get('pm_rows', []):
                pm_out.append([tgt, mid, pmra,
                               era if era is not None else '',
                               pmde,
                               ede if ede is not None else '',
                               bc or ''])
        _write_meas_sheet('SIMBAD PM',
                          ['Target', 'Main ID', 'pmRA (mas/yr)', 'e_pmRA',
                           'pmDec (mas/yr)', 'e_pmDec', 'Bibcode'],
                          pm_out)

        # ── Radial Velocity ───────────────────────────────────────────────────
        rv_out = []
        for tgt, mid, ext in ext_rows:
            for vel, vtype, err, bc in ext.get('rv_rows', []):
                rv_out.append([tgt, mid, vel,
                               err if err is not None else '',
                               vtype or '', bc or ''])
        _write_meas_sheet('SIMBAD RV',
                          ['Target', 'Main ID', 'RV (km/s)', 'e_RV', 'Type', 'Bibcode'],
                          rv_out)

        # ── Rotation ─────────────────────────────────────────────────────────
        rot_out = []
        for tgt, mid, ext in ext_rows:
            for vsini, err, bc in ext.get('rot_rows', []):
                rot_out.append([tgt, mid, vsini,
                                err if err is not None else '',
                                bc or ''])
        _write_meas_sheet('SIMBAD Rot',
                          ['Target', 'Main ID', 'v sin i (km/s)', 'e_v sin i', 'Bibcode'],
                          rot_out)

        # ── Photometry (conditional) ──────────────────────────────────────────
        phot = []
        for tgt, mid, ext in ext_rows:
            for band, flux, err, bc in ext.get('flux_rows', []):
                phot.append([tgt, mid, band, flux,
                             (err if err is not None else ''), bc or ''])
        if phot:
            ws_p = wb.create_sheet(title='SIMBAD Photometry')
            ws_p.append(['Target', 'Main ID', 'Filter', 'Flux', 'e_Flux', 'Bibcode'])
            for r in phot: ws_p.append(r)
            _style_ws(ws_p, 'simbad_phot')
            _apply_bibcode_cols(ws_p)
            _add_note_row(ws_p)
            written += 1

        # ── Stellar Parameters (conditional) ─────────────────────────────────
        stpar = []
        for tgt, mid, ext in ext_rows:
            for teff, logg, feh, bc in ext.get('fe_h_rows', []):
                stpar.append([tgt, mid,
                              (teff if teff is not None else ''),
                              (logg if logg is not None else ''),
                              (feh  if feh  is not None else ''),
                              bc or ''])
        if stpar:
            ws_sp = wb.create_sheet(title='SIMBAD Stell Params')
            ws_sp.append(['Target', 'Main ID', 'Teff (K)', 'log g', '[Fe/H]', 'Bibcode'])
            for r in stpar: ws_sp.append(r)
            _style_ws(ws_sp, 'simbad_stpar')
            _apply_bibcode_cols(ws_sp)
            _add_note_row(ws_sp)
            written += 1

        # ── Identifiers (conditional) ─────────────────────────────────────────
        idents = []
        for tgt, mid, ext in ext_rows:
            for ident in ext.get('identifiers', []):
                idents.append([tgt, mid, ident])
        if idents:
            ws_i = wb.create_sheet(title='SIMBAD Identifiers')
            ws_i.append(['Target', 'Main ID', 'Identifier'])
            for r in idents: ws_i.append(r)
            _style_ws(ws_i, 'simbad_ident')
            _add_note_row(ws_i)
            written += 1

        # ── References (conditional) ──────────────────────────────────────────
        refs = []
        for tgt, mid, ext in ext_rows:
            ref_titles = ext.get('ref_titles', {})
            for bc in ext.get('refs', []):
                refs.append([tgt, mid, bc, ref_titles.get(bc, '')])
        if refs:
            ws_r = wb.create_sheet(title='SIMBAD References')
            ws_r.append(['Target', 'Main ID', 'Bibcode', 'Title'])
            for r in refs: ws_r.append(r)
            _style_ws(ws_r, 'simbad_refs')
            _apply_bibcode_cols(ws_r)
            _add_note_row(ws_r)
            written += 1

        if written == 0:
            messagebox.showinfo("No Data", "All catalogs returned empty results.")
            return
        # Reorder sheets: all SIMBAD sheets together first, then other catalogs
        _SHEET_ORDER = [
            'SIMBAD', 'SIMBAD SpT', 'SIMBAD Plx', 'SIMBAD Dist',
            'SIMBAD PM', 'SIMBAD RV', 'SIMBAD Rot',
            'SIMBAD Photometry', 'SIMBAD Stell Params',
            'SIMBAD Identifiers', 'SIMBAD References',
            'AAVSO VSX', '2MASS', 'AllWISE', 'APASS', 'Tycho-2',
            'WDS', 'Orb6', 'Gaia DR3', 'NEA',
        ]
        existing = {ws.title: ws for ws in wb.worksheets}
        ordered  = [existing[t] for t in _SHEET_ORDER if t in existing]
        for ws in wb.worksheets:
            if ws not in ordered:
                ordered.append(ws)
        wb._sheets = ordered

        for ws in wb.worksheets:
            ws.freeze_panes = 'B2'
        wb.save(path)
        total = sum(
            len(rows)
            for td in self._batch_results.values()
            for rows in td.values()
        )
        self._set_status(
            f"Batch export: {total} total rows across {written} sheets → "
            f"{os.path.basename(path)}")

    def _show_revision_history(self):
        win = tk.Toplevel(self)
        win.title("Revision History — Stellar Object Query Tool")
        win.configure(bg=BG)
        win.resizable(True, True)
        win.geometry("680x500")

        tk.Label(win, text="Revision History", bg=PANEL, fg=ACC,
                 font=("Segoe UI", 11, "bold"), pady=8).pack(fill="x")

        tf = tk.Frame(win, bg=BG)
        tf.pack(fill="both", expand=True, padx=12, pady=8)

        txt = tk.Text(tf, bg=ENT, fg=FG, font=("Segoe UI", 11),
                      relief="flat", wrap="word", cursor="arrow", padx=10, pady=8)
        vsb = ttk.Scrollbar(tf, orient="vertical", command=txt.yview)
        txt.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        txt.pack(fill="both", expand=True)

        txt.tag_configure("ver",    foreground=ACC, font=("Segoe UI", 11, "bold"))
        txt.tag_configure("bullet", foreground=FG,  font=("Segoe UI", 11))

        def entry(version, date, bullets):
            txt.insert("end", f"{version}  —  {date}\n", "ver")
            for b in bullets:
                txt.insert("end", f"  •  {b}\n\n", "bullet")
            txt.insert("end", "\n")

        entry("v1.2.0", "2026-05-21", [
            "SIMBAD extended data (spectral type, parallax, proper motion, radial velocity, "
            "photometry, stellar parameters, identifiers, references) now pre-fetched in the "
            "background as soon as SIMBAD results arrive — no longer requires opening the "
            "SIMBAD tab first.",
            "Batch backward navigation fixed: left arrow now correctly restores all results "
            "and tab counts for previously visited targets.",
            "Progress indicators now run continuously until every catalog query has "
            "completed, rather than stopping after the first few tabs finish.",
            "Batch file browser now defaults to All Files instead of CSV.",
            "Batch column setup dialog now opens centred over the app window.",
            "WDS results pane now auto-sizes to fit the row count (no scrolling required "
            "for typical result sets).",
            "SOQyT acronym added to the title bar.",
            "User Guide restyled: proportional Segoe UI font, word-wrap, live search bar "
            "with Find Next / Clear, coloured section headers matching the app accent colour.",
            "Batch export: all numeric values (magnitudes, parallaxes, periods, etc.) now "
            "written as native Excel numbers, eliminating green-arrow warnings.",
            "Batch export: N Refs column now shows the true total SIMBAD reference count "
            "(nbref) instead of the variability-bibcode count from the initial query.",
            "Batch export: five additional SIMBAD sheets — Meas (spectral type, parallax, "
            "distance, proper motion, radial velocity, v sin i), Photometry, Stell Params "
            "(Teff / log g / [Fe/H]), Identifiers, and References — written immediately "
            "after the main SIMBAD sheet when extended data is available.",
            "Batch export: bibcode cells in SIMBAD extended sheets are hyperlinked to "
            "NASA ADS (ui.adsabs.harvard.edu/abs/…).",
            "Batch export: per-sheet styling — distinct coloured headers with white bold "
            "text, alternating row shading, frozen header row, and auto-fit column widths.",
            "Batch export: Orb6 sheet gains a Plot URL column; each cell is a clickable "
            "hyperlink to the GSU orbital-plot PNG.",
            "Name resolution: 'Gaia DR3 XXXXXX' names now fall back to a direct Gaia "
            "source_id position lookup if SIMBAD resolution fails, so Orb6 and WDS cone "
            "searches succeed in name-mode batch runs even when SIMBAD is unresponsive.",
            "About dialog: GitHub URL is now a clickable hyperlink.",
            "About dialog: contact email (art.trail@icloud.com) added as a clickable "
            "mailto link; dialog now opens centred over the app window.",
            "Auto-Run batch mode: new checkbox in the batch import dialog runs all "
            "targets sequentially without user interaction, then prompts to export. "
            "A Cancel Auto-Run button is shown in the batch nav bar during the run; "
            "a 1-second courtesy delay separates targets.",
            "VizieR and WDS cone-search results are now sorted by angular distance from "
            "the query centre — the closest catalog source is always returned first, "
            "eliminating cases where an off-axis neighbour was silently substituted for "
            "the intended target.",
            "Batch export and single-tab CSV export: a note row appended to every sheet "
            "warns that positional matches return the nearest source within the search "
            "radius, and to verify results in crowded fields or for close binaries.",
            "Batch export filename now defaults to the original input filename plus "
            "_SOQyT_Query_by_Name/Coord and a date/time stamp.",
            "SIMBAD Object Type Code (OType) column fixed — was always blank due to a "
            "missing dict key in the result row builder.",
            "Batch export: SIMBAD measurement data split into six typed sheets (SpT, Plx, "
            "Dist, PM, RV, Rot) with proper SIMBAD column names; all six are always written "
            "— empty sheets show a 'no data' message rather than being omitted.",
            "Batch export: all sheets have frozen header row and column A, with center-aligned "
            "data cells.",
            "Batch export: Gaia DR3 Source ID now written as a text string to preserve full "
            "18-19 digit precision (previously truncated by float conversion).",
            "Batch export: positional photometric catalogs (2MASS, AllWISE, APASS, Tycho-2) "
            "export only the nearest match per target when multiple sources fall within the "
            "search radius.",
            "Batch export: SIMBAD sheets are always grouped together at the front of the "
            "workbook regardless of query mode.",
            "Coord-mode SIMBAD query changed from INNER JOIN to LEFT JOIN mesVar — now "
            "returns all stellar objects, not variables only.",
            "Batch config dialog: search radius field added to Name mode (previously used "
            "the main panel radius silently).",
            f"Auto-Run server cooldown: a {_COOLDOWN_SECS}-second pause is automatically "
            f"inserted every {_COOLDOWN_EVERY} targets during Auto-Run to prevent catalog "
            f"server rate-limit errors. A countdown is shown in the status bar.",
        ])
        entry("v1.1.0", "2026-05-19", [
            "Name entry hint expanded to show catalog-specific format examples "
            "(RR Lyr, WASP-24, GJ 3470, HD 216963, Gaia DR3 …).",
            "Resolved name display: after a name-mode search, the SIMBAD canonical "
            "main_id is shown below the search field (e.g. 'SIMBAD → HD 216963').",
            "AAVSO VSX detail panel: Remarks section (submitter + text) fetched from "
            "the VSX star page alongside References; both displayed with clickable links.",
            "AAVSO VSX detail panel: reorganised to columnar label:value layout; "
            "added clickable AAVSO VSX star page link.",
            "Gaia DR3 detail panel: reorganised to columnar label:value layout; "
            "added clickable catalog reference (Gaia Collaboration 2023, ADS).",
            "VizieR detail panels (2MASS, AllWISE, APASS DR9, Tycho-2): reorganised to "
            "columnar layout with clickable catalog references (ADS).",
            "VizieR queries expanded: 2MASS adds e_JHK, quality/contamination flags; "
            "AllWISE adds errors, 2MASS cross-ref, quality/PM flags; APASS adds B-V and "
            "errors; Tycho-2 adds errors, HIP cross-ID, proximity flag.",
            "SIMBAD Other Names: fixed-width column layout (3 per row, aligned on "
            "longest name) replaces ragged pipe-separated format.",
            "NEA detail tab source note font increased to match tab label size.",
            "Added Help menu with User Guide, Revision History, and About.",
            "Added version number to title bar and copyright to footer.",
            "Added SOQyT GitHub repo link in About dialog.",
            "SIMBAD detail panel expanded: per-measurement source citations for spectral "
            "type, parallax, proper motion, radial velocity, photometry, and stellar "
            "parameters — each value shows its bibcode as a hyperlink to NASA ADS.",
            "SIMBAD references: full bibliography (71+ entries) via ref + has_ref TAP join "
            "with paper titles shown inline; titles word-wrap aligned under the title start. "
            "Bibcodes are hyperlinks to NASA ADS.",
            "Reference year-range filter (debounced) with default upper bound = current year.",
            "Recent star name history: last 5 searches saved to config and recalled via "
            "Combobox dropdown.",
            "Results pane auto-sizes to fit the active tab row count; detail panel "
            "fills remaining space (replaces fixed 50/50 split).",
            "Left panel declutter: DATA SOURCES and FILTERS sections are now collapsible "
            "(▼/▶ toggle, both open by default).",
            "Gaia DR3 parameters moved to a ⚙ popup (click gear icon on Gaia DR3 row); "
            "added BP-RP color as a selectable Gaia parameter.",
            "SIMBAD detail measurement sections (SpT, Plx, Distance, PM, Radial Velocity, "
            "Rotation (v sin i), Photometry, Stellar Parameters) individually togglable "
            "via ⚙ popup (click gear icon on SIMBAD row).",
            "SIMBAD: fixed detail panel overwriting other tabs when a background fetch "
            "completes while a non-SIMBAD tab is active.",
            "AAVSO VSX: fixed RA parsing for decimal-degree format; added Constellation, "
            "Spectral Type, Discoverer, Epoch, Rise Duration, and Other Names to detail panel.",
            "NEA: replaced 5 sub-tabs (Overview / Orbital / Planet / Star / System) with a "
            "single flat results tab; all data remains accessible in the detail panel.",
            "NEA: added per-parameter references from pscomppars _reflink columns (22 "
            "parameters: discovery, orbital, planet, stellar, system). References appear in "
            "the detail panel grouped by unique paper with parameter lists.",
            "VizieR: added Washington Double Star Catalog (WDS, B/wds/wds) as a fifth "
            "sub-tab. Shows components, discoverer, first/last obs year, position angle, "
            "separation, magnitudes, and spectral type.",
            "VizieR: added Sixth Orbit Catalog (orb6) as a sixth sub-tab. Data fetched "
            "directly from the GSU/WDS website. Detail panel shows full orbital elements "
            "(period, epoch, semi-major axis, inclination, node, eccentricity, argument of "
            "periastron), grade, last obs year, and reference. If an orbital plot is "
            "available, a clickable link opens it in the browser.",
            "Star Name field: wildcard searches using * (e.g. RR*) trigger LIKE queries "
            "on SIMBAD and NEA, capped at 100 results per source.",
            "Star Name field: WDS identifiers auto-normalised — 'WDS 00001+5400' is "
            "silently corrected to 'WDS J00001+5400' before querying.",
            "SIMBAD filters (Object Type, Period range, Max Mag range, Ref year range) "
            "moved under the SIMBAD row in DATA SOURCES; standalone FILTERS section removed.",
            "Star Name cursor colour changed to white for visibility on dark background.",
            "Batch import: load a CSV or XLSX target list and query all catalogs line by "
            "line. Import via '📂 Import Batch…' button; choose a Star Name column or RA "
            "and Dec columns. Navigate targets with ◀ / ▶ buttons — results for each "
            "target are cached so you can step back without re-querying. Export all "
            "accumulated results to a multi-sheet XLSX (one sheet per catalog, with a "
            "'Target' column prepended) using '⬇ Export Batch Results'.",
            "Detail panels: RA and Dec now show decimal degrees in parentheses alongside "
            "the sexagesimal value (e.g. 20 17 34.66  (304.3944°)). Applies to all tabs: "
            "SIMBAD, VSX, 2MASS, AllWISE, APASS DR9, Tycho-2, WDS, Orb6, and Gaia DR3. "
            "Orb6 detail panel also gains RA and Dec fields (previously absent).",
        ])
        entry("v1.0.0", "2026-05-10", [
            "Initial release. Queries SIMBAD, AAVSO VSX, VizieR (2MASS, AllWISE, "
            "APASS DR9, Tycho-2), Gaia DR3, and NASA Exoplanet Archive. Search by "
            "name or coordinates. Filters for object type, period, and magnitude. "
            "Export to Excel and CSV.",
        ])

        txt.config(state="disabled")

        tk.Button(win, text="Close", command=win.destroy,
                  bg=PANEL, fg=FG, relief="flat",
                  font=("Segoe UI", 11), cursor="hand2", padx=12, pady=4
                  ).pack(pady=(0, 8))

    def _show_user_guide(self):
        win = tk.Toplevel(self)
        win.title("User Guide — Stellar Object Query Tool")
        win.configure(bg=BG)
        win.resizable(True, True)
        win.geometry("860x980")

        tk.Label(win, text="Stellar Object Query Tool — User Guide", bg=PANEL, fg=ACC,
                 font=("Segoe UI", 14, "bold"), pady=8).pack(fill="x")

        # Search bar
        sb = tk.Frame(win, bg=PANEL, padx=8, pady=5)
        sb.pack(fill="x")
        tk.Label(sb, text="Search:", bg=PANEL, fg=FG,
                 font=("Segoe UI", 12)).pack(side="left", padx=(0, 4))
        search_var = tk.StringVar()
        search_entry = tk.Entry(sb, textvariable=search_var, bg=ENT, fg=FG,
                                insertbackground=FG, font=("Segoe UI", 12),
                                width=28, relief="flat")
        search_entry.pack(side="left", padx=(0, 6))
        _find_btn = tk.Button(sb, text="Find Next",
                              bg=PANEL, fg=FG, activebackground=ACC, activeforeground=BG,
                              relief="flat", font=("Segoe UI", 11),
                              cursor="hand2", padx=8, pady=2)
        _find_btn.pack(side="left", padx=(0, 4))
        _clr_btn = tk.Button(sb, text="Clear",
                             bg=PANEL, fg=FG, activebackground=ACC, activeforeground=BG,
                             relief="flat", font=("Segoe UI", 11),
                             cursor="hand2", padx=8, pady=2)
        _clr_btn.pack(side="left", padx=(0, 8))
        search_lbl = tk.Label(sb, text="", bg=PANEL, fg=FG,
                              font=("Segoe UI", 11), width=14, anchor="w")
        search_lbl.pack(side="left")

        tf = tk.Frame(win, bg=BG)
        tf.pack(fill="both", expand=True, padx=8, pady=8)

        txt = tk.Text(tf, bg=ENT, fg=FG, font=("Segoe UI", 14),
                      relief="flat", wrap="word", cursor="arrow",
                      padx=14, pady=10, spacing1=2, spacing2=2)
        vsb = ttk.Scrollbar(tf, orient="vertical", command=txt.yview)
        txt.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        txt.pack(fill="both", expand=True)

        txt.tag_configure("h2",   foreground=ACC, font=("Segoe UI", 14, "bold"),
                          spacing1=14, spacing3=4)
        txt.tag_configure("bold", font=("Segoe UI", 14, "bold"))
        txt.tag_configure("i1",   lmargin1=22, lmargin2=22)
        txt.tag_configure("i2",   lmargin1=44, lmargin2=44)
        txt.tag_configure("shi",  background="#4a6b9f", foreground="white")
        txt.tag_configure("scur", background=ACC, foreground=BG)

        search_matches = []
        search_idx = [0]

        def _do_search(*_):
            txt.tag_remove("shi",  "1.0", "end")
            txt.tag_remove("scur", "1.0", "end")
            search_matches.clear()
            q = search_var.get().strip()
            if not q:
                search_lbl.config(text="")
                return
            start = "1.0"
            while True:
                pos = txt.search(q, start, stopindex="end", nocase=True)
                if not pos:
                    break
                end_pos = f"{pos}+{len(q)}c"
                txt.tag_add("shi", pos, end_pos)
                search_matches.append(pos)
                start = end_pos
            if search_matches:
                search_idx[0] = 0
                _highlight_current()
                search_lbl.config(text=f"1 / {len(search_matches)}")
            else:
                search_lbl.config(text="Not found")

        def _find_next(*_):
            if not search_matches:
                _do_search()
                return
            search_idx[0] = (search_idx[0] + 1) % len(search_matches)
            _highlight_current()
            search_lbl.config(text=f"{search_idx[0]+1} / {len(search_matches)}")

        def _highlight_current():
            txt.tag_remove("scur", "1.0", "end")
            if not search_matches:
                return
            pos = search_matches[search_idx[0]]
            q = search_var.get()
            end_pos = f"{pos}+{len(q)}c"
            txt.tag_add("scur", pos, end_pos)
            txt.see(pos)

        def _clear_search(*_):
            search_var.set("")
            search_matches.clear()
            search_idx[0] = 0
            search_lbl.config(text="")

        _find_btn.config(command=_find_next)
        _clr_btn.config(command=_clear_search)
        search_entry.bind("<Return>", _find_next)
        search_var.trace_add("write", _do_search)

        sections = [
            ("1.  Search Modes",     "sec_modes"),
            ("2.  Data Sources",     "sec_sources"),
            ("3.  Filters",          "sec_filters"),
            ("4.  Results & Detail", "sec_results"),
            ("5.  Exporting",        "sec_export"),
            ("6.  Batch Import",     "sec_batch"),
        ]

        txt.insert("end", "TABLE OF CONTENTS\n", "h2")
        txt.insert("end", "\n")
        for label, mark in sections:
            tag = f"toc_{mark}"
            txt.insert("end", f"  {label}\n", tag)
            txt.tag_configure(tag, foreground=ACC, font=("Segoe UI", 14),
                              underline=True, lmargin1=16, lmargin2=16)
            txt.tag_bind(tag, "<Button-1>", lambda e, m=mark: txt.see(m))
            txt.tag_bind(tag, "<Enter>",    lambda e: txt.config(cursor="hand2"))
            txt.tag_bind(tag, "<Leave>",    lambda e: txt.config(cursor="arrow"))
        txt.insert("end", "\n")

        def _h2(title, mark):
            txt.mark_set(mark, "end")
            txt.mark_gravity(mark, "left")
            txt.insert("end", title + "\n", "h2")

        # Section 1 — Search Modes
        _h2("1.  SEARCH MODES", "sec_modes")
        txt.insert("end", "• ", "i1"); txt.insert("end", "By Name", ("i1", "bold"))
        txt.insert("end", " — enter a star name or identifier (e.g. \"RR Lyr\","
                   " \"V* AB Aur\", \"HIP 91262\", \"20176+2622\")."
                   " Searches all selected databases for that object.\n", "i1")
        txt.insert("end", "• ", "i1"); txt.insert("end", "By Coordinates", ("i1", "bold"))
        txt.insert("end", " — enter RA and Dec (degrees or hms/dms) and a search radius"
                   " in arcminutes. Returns all objects within that cone from all"
                   " selected databases.\n", "i1")
        txt.insert("end", "\n")

        # Section 2 — Data Sources
        _h2("2.  DATA SOURCES", "sec_sources")
        txt.insert("end", "• ", "i1"); txt.insert("end", "SIMBAD", ("i1", "bold"))
        txt.insert("end", " — CDS Strasbourg; broad object types, spectral types,"
                   " parallax, proper motion, and cross-identifiers.\n", "i1")
        txt.insert("end", "• ", "i1"); txt.insert("end", "AAVSO VSX", ("i1", "bold"))
        txt.insert("end", " — Variable Star Index; period, variability type, magnitude.\n", "i1")
        txt.insert("end", "• ", "i1"); txt.insert("end", "VizieR", ("i1", "bold"))
        txt.insert("end", " — Six catalogues across separate sub-tabs:\n", "i1")
        for _cname, _cdesc in [
            ("2MASS",     "J, H, K near-infrared magnitudes."),
            ("AllWISE",   "W1–W4 mid-infrared magnitudes."),
            ("APASS DR9", "V and B optical magnitudes."),
            ("Tycho-2",   "BT, VT magnitudes and proper motions."),
            ("WDS",       "Washington Double Star Catalog (position angle, separation,"
                          " component magnitudes, discoverer)."),
            ("Orb6",      "Sixth Orbit Catalog; computed visual binary orbits (period,"
                          " semi-major axis, inclination, eccentricity)."
                          " Clicking a row shows a link to the orbital plot."),
        ]:
            txt.insert("end", "  • ", "i2"); txt.insert("end", _cname, ("i2", "bold"))
            txt.insert("end", f" — {_cdesc}\n", "i2")
        txt.insert("end", "  VizieR and WDS results are positional — the app performs a cone"
                   " search centred on the resolved coordinates and returns all sources within"
                   " the search radius, sorted by angular distance (closest first). The"
                   " nearest source is therefore always listed at the top, but it may not"
                   " be the intended target in crowded fields or for close binaries. Always"
                   " check the Dist(\") column and verify identifications where separation"
                   " from the query centre is larger than expected. SIMBAD, Gaia DR3 (name"
                   " mode), VSX, and NEA use name- or ID-based resolution and are not"
                   " subject to this ambiguity.\n", "i1")
        txt.insert("end", "\n")
        txt.insert("end", "• ", "i1"); txt.insert("end", "Gaia DR3", ("i1", "bold"))
        txt.insert("end", " — ESA; parallax, proper motion, BP/RP magnitudes, RUWE,"
                   " radial velocity. Select columns via the Gaia ⚙ button.\n", "i1")
        txt.insert("end", "• ", "i1"); txt.insert("end", "NEA", ("i1", "bold"))
        txt.insert("end", " — NASA Exoplanet Archive; planet and host-star parameters"
                   " (orbital, physical, stellar, system, references).\n", "i1")
        txt.insert("end", "\n")

        # Section 3 — Filters
        _h2("3.  FILTERS", "sec_filters")
        txt.insert("end", "• ", "i1"); txt.insert("end", "Object Type", ("i1", "bold"))
        txt.insert("end", " (SIMBAD only) — restrict results to a SIMBAD object type"
                   " code (e.g. \"RR*\" for RR Lyrae, \"V*\" for variable stars).\n", "i1")
        txt.insert("end", "• ", "i1"); txt.insert("end", "Period range", ("i1", "bold"))
        txt.insert("end", " — filter AAVSO VSX results by period in days.\n", "i1")
        txt.insert("end", "• ", "i1"); txt.insert("end", "Max Mag range", ("i1", "bold"))
        txt.insert("end", " — filter AAVSO VSX results by maximum magnitude.\n", "i1")
        txt.insert("end", "All filters are optional and applied per-database where relevant.\n")
        txt.insert("end", "\n")

        # Section 4 — Results & Detail
        _h2("4.  RESULTS & DETAIL", "sec_results")
        txt.insert("end", "Each database result appears in its own tab."
                   " Click any row to show full details in the Detail panel below the results.\n")
        txt.insert("end", "\n")
        txt.insert("end", "SIMBAD extended data (spectral type, parallax, proper motion, radial"
                   " velocity, photometry, stellar parameters, identifiers, and references) is"
                   " fetched in the background as soon as SIMBAD results arrive. Clicking a row"
                   " shows the detail immediately if ready, or displays a loading indicator if"
                   " still fetching. Subsequent clicks reuse the cached data. Use the SIMBAD ⚙"
                   " button to toggle which measurement sections are shown.\n")
        txt.insert("end", "\n")
        txt.insert("end", "RA and Dec are shown in both sexagesimal and decimal degrees"
                   " (in parentheses) throughout the Detail panel.\n")
        txt.insert("end", "\n")
        txt.insert("end", "When a query returns exactly one result it is auto-selected and its"
                   " detail is shown immediately (if that tab is active).\n")
        txt.insert("end", "\n")
        txt.insert("end", "Switching tabs updates the Detail panel to the selected row of"
                   " the newly active tab.\n")
        txt.insert("end", "\n")

        # Section 5 — Exporting
        _h2("5.  EXPORTING", "sec_export")
        txt.insert("end", "• ", "i1"); txt.insert("end", "Export Excel", ("i1", "bold"))
        txt.insert("end", " — saves the active tab's results to an .xlsx file.\n", "i1")
        txt.insert("end", "• ", "i1"); txt.insert("end", "Export CSV", ("i1", "bold"))
        txt.insert("end", " — saves the active tab's results to a .csv file.\n", "i1")
        txt.insert("end", "Both exports include all result columns for the active tab.\n")
        txt.insert("end", "\n")
        txt.insert("end", "• ", "i1"); txt.insert("end", "Export Batch", ("i1", "bold"))
        txt.insert("end", " — available during a batch session (see section 6)."
                   " Saves all accumulated results to a styled multi-sheet Excel file."
                   " Each catalog gets its own colour-coded sheet with a frozen header row,"
                   " center-aligned cells, and auto-fit columns. SIMBAD sheets are grouped"
                   " first: main SIMBAD results followed by six measurement sheets (SpT,"
                   " Plx, Dist, PM, RV, Rot) — always present, showing \"no data\" when"
                   " empty — plus Photometry, Stell Params, Identifiers, and References"
                   " when data is available. All bibcodes are hyperlinked to NASA ADS."
                   " The Orb6 sheet includes a clickable Plot URL column.\n", "i1")
        txt.insert("end", "\n")

        # Section 6 — Batch Import
        _h2("6.  BATCH IMPORT", "sec_batch")
        txt.insert("end", "Import a CSV or Excel file to query a list of targets one by one.\n")
        txt.insert("end", "\n")
        txt.insert("end", "HOW TO USE\n", "bold")
        txt.insert("end", "1. Click \"Import Batch…\" in the left panel.\n", "i1")
        txt.insert("end", "2. Select a .csv or .xlsx file.\n", "i1")
        txt.insert("end", "3. In the Column Setup dialog:\n", "i1")
        txt.insert("end", "   • Name mode — choose the column that contains star names"
                   " or WDS designations.\n", "i2")
        txt.insert("end", "   • Coord mode — choose the columns for RA and Dec, and set"
                   " a search radius (arcmin).\n", "i2")
        txt.insert("end", "   • Auto-Run — check this box to query every row automatically"
                   " without interaction. The app steps through all targets in sequence"
                   " (1-second pause between each), then prompts to export when finished."
                   " A Cancel Auto-Run button appears in the batch bar during the run.\n", "i2")
        txt.insert("end", "4. Click \"Load Batch\". The first target is queried automatically.\n", "i1")
        txt.insert("end", "5. Use Prev / Next to move between targets. Results for each"
                   " target are stored as you navigate.\n", "i1")
        txt.insert("end", "6. Click \"Export Batch Results\" to save all accumulated results to a"
                   " multi-sheet Excel file (one sheet per database). The file is named"
                   " after the original input file with _SOQyT_Query_by_Name/Coord and a"
                   " date/time stamp appended by default.\n", "i1")
        txt.insert("end", "7. Click \"Close Batch\" to exit batch mode and clear results.\n", "i1")
        txt.insert("end", "\n")
        txt.insert("end", "NOTES\n", "bold")
        txt.insert("end", "• The header bar shows \"Auto-running: N / M\" during Auto-Run,"
                   " or \"N / M: target\" during manual navigation.\n", "i1")
        txt.insert("end", "• ", "i1"); txt.insert("end", "Coord mode runs faster than Name mode.", ("i1", "bold"))
        txt.insert("end", " Name mode must first resolve each name to coordinates via SIMBAD"
                   " before the other catalog queries can fire, and it pre-fetches SIMBAD"
                   " extended data (up to ~10 additional queries per star) in the background."
                   " Coord mode skips name resolution and fires all queries in parallel"
                   " immediately.\n", "i1")
        txt.insert("end", "• ", "i1"); txt.insert("end", "Server cooldown pauses", ("i1", "bold"))
        txt.insert("end", f" are inserted automatically during Auto-Run every"
                   f" {_COOLDOWN_EVERY} targets. A {_COOLDOWN_SECS}-second countdown"
                   f" is shown in the status bar (\"Server cooldown — resuming in Ns\")."
                   f" This prevents catalog server rate-limit errors during large batch"
                   f" runs.\n", "i1")
        txt.insert("end", "• Navigating away before a query finishes discards in-flight results"
                   " for that target.\n", "i1")
        txt.insert("end", "• Starting a manual search while in batch mode asks for confirmation"
                   " before clearing batch results.\n", "i1")
        txt.insert("end", "• Only results that have been loaded (by visiting the target or"
                   " running Auto-Run) are included in the batch export.\n", "i1")
        txt.insert("end", "\n")

        txt.config(state="disabled")

        tk.Button(win, text="Close", command=win.destroy,
                  bg=PANEL, fg=FG, relief="flat",
                  font=("Segoe UI", 14), cursor="hand2", padx=12, pady=4
                  ).pack(pady=(0, 8))

    def _render_simbad_detail(self, result_row):
        """Render SIMBAD detail from cached data dict, applying current year filter."""
        if result_row is None:
            return
        if self._active_tab_key() != 'simbad':
            return  # data already cached in result_row; renders on next SIMBAD tab selection
        data = result_row.get('_extended_detail')
        if not isinstance(data, dict):
            return

        try:
            yr_from = int(self._ref_year_from_var.get())
        except (ValueError, AttributeError):
            yr_from = None
        try:
            yr_to = int(self._ref_year_to_var.get())
        except (ValueError, AttributeError):
            yr_to = None

        d = lambda v: v if v else '—'
        meas = self._simbad_meas_vars   # shorthand
        segs = []

        # Basic fields from main query
        segs.append((
            f"Name: {d(result_row.get('Name'))}    "
            f"RA: {_ra_with_deg(d(result_row.get('RA_hms')))}    "
            f"Dec: {_dec_with_deg(d(result_row.get('Dec_dms')))}    "
            f"Type: {d(result_row.get('OType_label'))}    "
            f"Var Sub-type: {d(result_row.get('VarType'))}\n",
            None
        ))
        segs.append((
            f"Period: {d(result_row.get('Period'))} d    "
            f"Max Mag: {d(result_row.get('MaxMag'))}    "
            f"Min Mag: {d(result_row.get('MinMag'))}    "
            f"Band: {d(result_row.get('MagBand'))}    "
            f"Dist: {d(result_row.get('Dist_arcsec'))}\"    "
            f"Var refs: {d(result_row.get('N_refs'))}\n",
            None
        ))

        def _bib_seg(bc, prefix='  '):
            """Append an inline bibcode link at end of a measurement line."""
            if bc:
                segs.append((f'{prefix}', None))
                segs.append((bc, 'link'))

        # Spectral Type
        sptype_rows = data.get('sptype_rows', [])
        if sptype_rows and meas['spt'].get():
            segs.append(('\n', None))
            segs.append(('Spectral Type:\n', None))
            for sp, bc in sptype_rows:
                segs.append((f'  {sp}', None))
                _bib_seg(bc, '  ')
                segs.append(('\n', None))

        # Parallax
        plx_rows = data.get('plx_rows', [])
        if plx_rows and meas['plx'].get():
            segs.append(('\n', None))
            segs.append(('Parallax:\n', None))
            for plx, err, dist, bc in plx_rows:
                line = f'  {plx:.3f}'
                line += f' ± {err:.3f}' if err is not None else ''
                line += ' mas'
                if dist is not None:
                    line += f'  (≈ {dist:.0f} pc)'
                segs.append((line, None))
                _bib_seg(bc, '  ')
                segs.append(('\n', None))

        # Proper Motion
        pm_rows = data.get('pm_rows', [])
        if pm_rows and meas['pm'].get():
            segs.append(('\n', None))
            segs.append(('Proper Motion:\n', None))
            for pmra, pmde, era, ede, bc in pm_rows:
                line = f'  RA: {pmra:+.3f}'
                line += f' ± {era:.3f}' if era is not None else ''
                line += f'   Dec: {pmde:+.3f}'
                line += f' ± {ede:.3f}' if ede is not None else ''
                line += '  mas/yr'
                segs.append((line, None))
                _bib_seg(bc, '  ')
                segs.append(('\n', None))

        # Radial Velocity
        rv_rows = data.get('rv_rows', [])
        if rv_rows and meas['rv'].get():
            segs.append(('\n', None))
            segs.append(('Radial Velocity:\n', None))
            for vel, vtype, err, bc in rv_rows:
                line = f'  {vel:.2f}'
                line += f' ± {err:.2f}' if err is not None else ''
                line += ' km/s'
                if vtype:
                    line += f'  ({vtype})'
                segs.append((line, None))
                _bib_seg(bc, '  ')
                segs.append(('\n', None))

        # Photometry (individual flux measurements with bibcodes)
        flux_rows = data.get('flux_rows', [])
        if flux_rows and meas['flux'].get():
            segs.append(('\n', None))
            segs.append(('Photometry:\n', None))
            for band, val, err, bc in flux_rows:
                line = f'  {band}={val:.3f}'
                line += f' ± {err:.3f}' if err is not None else ''
                segs.append((line, None))
                _bib_seg(bc, '  ')
                segs.append(('\n', None))

        # Rotation (mesRot)
        rot_rows = data.get('rot_rows', [])
        if rot_rows and meas['rot'].get():
            segs.append(('\n', None))
            segs.append(('Rotation:\n', None))
            for vsini, err, bc in rot_rows:
                line = f'  v sin i = {vsini:.1f}'
                line += f' ± {err:.1f}' if err is not None else ''
                line += ' km/s'
                segs.append((line, None))
                _bib_seg(bc, '  ')
                segs.append(('\n', None))

        # Distance (mesDist)
        dist_rows = data.get('dist_rows', [])
        if dist_rows and meas['dist'].get():
            segs.append(('\n', None))
            segs.append(('Distance:\n', None))
            for dist, err, unit, bc in dist_rows:
                line = f'  {dist:.1f}'
                line += f' ± {err:.1f}' if err is not None else ''
                line += f' {unit}'
                segs.append((line, None))
                _bib_seg(bc, '  ')
                segs.append(('\n', None))

        # Stellar parameters (mesFe_H)
        fe_h_rows = data.get('fe_h_rows', [])
        if fe_h_rows and meas['fe_h'].get():
            segs.append(('\n', None))
            segs.append(('Stellar Parameters:\n', None))
            for teff, log_g, fe_h, bibcode in fe_h_rows:
                parts = []
                if teff  is not None: parts.append(f"Teff={teff:.0f} K")
                if log_g is not None: parts.append(f"log g={log_g:.2f}")
                if fe_h  is not None: parts.append(f"[Fe/H]={fe_h:+.2f}")
                segs.append(('  ' + '   '.join(parts), None))
                _bib_seg(bibcode, '  ')
                segs.append(('\n', None))

        if data.get('nbref') is not None:
            segs.append(('\n', None))
            segs.append((f"Total refs in SIMBAD: {data['nbref']}\n", None))

        # Identifiers — fixed-width columns (3 per row, width = longest name)
        identifiers = data.get('identifiers', [])
        if identifiers:
            segs.append(('\n', None))
            segs.append(('Other Names:\n', None))
            col_w = max(len(n) for n in identifiers)
            for i in range(0, len(identifiers), 3):
                chunk = identifiers[i:i+3]
                row_str = '  ' + '   '.join(n.ljust(col_w) for n in chunk).rstrip() + '\n'
                segs.append((row_str, None))

        # References (year-filtered)
        refs      = data.get('refs', [])
        nbref     = data.get('nbref')
        ref_titles = data.get('ref_titles', {})
        if refs:
            if yr_from is not None or yr_to is not None:
                filtered = []
                for bib in refs:
                    try:
                        year = int(bib[:4])
                        if yr_from is not None and year < yr_from: continue
                        if yr_to   is not None and year > yr_to:   continue
                        filtered.append(bib)
                    except (ValueError, IndexError):
                        filtered.append(bib)
                yr_parts = []
                if yr_from is not None: yr_parts.append(f"from {yr_from}")
                if yr_to   is not None: yr_parts.append(f"to {yr_to}")
                count_str = f"{len(filtered)} of {len(refs)} collected"
                if nbref is not None: count_str += f" / {nbref} total in SIMBAD"
                hdr = f"References ({', '.join(yr_parts)})  [{count_str}]:\n"
            else:
                filtered = refs
                count_str = f"{len(refs)} collected"
                if nbref is not None: count_str += f" / {nbref} total in SIMBAD"
                hdr = f"References  [{count_str}]:\n"
            segs.append(('\n', None))
            segs.append((hdr, None))
            for bib in filtered:
                wrapped = self._wrap_ref_title(ref_titles.get(bib, ''))
                segs.append(((bib, wrapped), 'ref_entry'))

        self._set_detail_rich(segs)

    def _make_link_tag(self, txt, bib):
        """Configure and return a unique hyperlink tag for a bibcode."""
        link_tag = f"_lnk_{bib}"
        url = f"https://ui.adsabs.harvard.edu/abs/{urllib.parse.quote(bib)}"
        txt.tag_configure(link_tag, foreground=ACC, underline=True)
        txt.tag_bind(link_tag, '<Button-1>', lambda e, u=url: webbrowser.open(u))
        txt.tag_bind(link_tag, '<Enter>',    lambda e: txt.config(cursor='hand2'))
        txt.tag_bind(link_tag, '<Leave>',    lambda e: txt.config(cursor='arrow'))
        return link_tag

    def _wrap_ref_title(self, title, prefix_chars=24):
        """Word-wrap a ref title so continuation lines align under the title start."""
        if not title:
            return ''
        txt = self._detail_text
        w = txt.winfo_width()
        if w < 100:
            w = 800
        if not hasattr(self, '_ref_char_w'):
            f = tkfont.Font(font=txt.cget('font'))
            self._ref_char_w = max(1, f.measure('M'))
        # Subtract ~20px for widget internal padding + scrollbar
        chars_per_line = max(40, (w - 20) // self._ref_char_w)
        avail = max(15, chars_per_line - prefix_chars)
        words = title.split()
        if not words:
            return ''
        lines, current, current_len = [], [], 0
        for word in words:
            wl = len(word)
            if not current:
                current.append(word)
                current_len = wl
            elif current_len + 1 + wl <= avail:
                current.append(word)
                current_len += 1 + wl
            else:
                lines.append(' '.join(current))
                current = [word]
                current_len = wl
        if current:
            lines.append(' '.join(current))
        return ('\n' + ' ' * prefix_chars).join(lines)

    def _set_detail_rich(self, segments):
        """Insert (text, tag) segments into the detail panel.
        tag='link'      → clickable ADS bibcode link
        tag='ref_entry' → text=(bib, pre-wrapped-title) tuple
        """
        txt = self._detail_text
        txt.config(state='normal')
        txt.delete('1.0', 'end')
        for text, tag in segments:
            if tag == 'ref_entry':
                bib, title = text
                link_tag = self._make_link_tag(txt, bib)
                txt.insert('end', '  ')
                txt.insert('end', bib, link_tag)
                if title:
                    txt.insert('end', f'   {title}')
                txt.insert('end', '\n')
            elif tag == 'link':
                link_tag = self._make_link_tag(txt, text)
                txt.insert('end', text, link_tag)
            else:
                txt.insert('end', text)
        txt.config(state='disabled')

    def _auto_size_pane(self):
        """Shrink the results pane to fit the active tab's row count; detail fills the rest."""
        if not hasattr(self, '_right_pane'):
            return
        try:
            key = self._active_tab_key()
        except Exception:
            key = None
        n = len(self._results.get(key, [])) if key else 0
        _vizier_sub = {k for k, _, _ in self.VIZIER_SUBTABS}
        if key == 'nea':
            OVERHEAD = 220   # NEA has an extra note label (~90px)
        elif key in _vizier_sub:
            OVERHEAD = 168   # VizieR sub-tabs carry both main and sub-notebook tab bars
        else:
            OVERHEAD = 128
        ROW_H    = 44    # matches rowheight=44 in Treeview style
        target   = OVERHEAD + n * ROW_H
        win_h = self.winfo_height()
        if win_h < 50:
            win_h = 900
        target = max(80, min(target, int(win_h * 0.65)))
        self._right_pane.sash_place(0, 0, target)

    def _on_ref_year_filter_change(self, *args):
        if self._year_filter_job:
            self.after_cancel(self._year_filter_job)
        self._year_filter_job = self.after(400, self._apply_year_filter_rerender)

    def _apply_year_filter_rerender(self):
        self._year_filter_job = None
        row = self._active_simbad_row
        if row is not None and isinstance(row.get('_extended_detail'), dict):
            self._render_simbad_detail(row)

    # ──────────────────────────────────────────────────────────
    # Status bar
    # ──────────────────────────────────────────────────────────

    def _set_status(self, msg):
        self._status_var.set(msg)


# ──────────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────────

if __name__ == '__main__':
    app = StarQueryApp()
    app.mainloop()
